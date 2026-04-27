# -*- coding: utf-8 -*-
"""
Листы «ДДС 20xx» в финансовой книге: помесячные доходы/расходы по статьям.
Также: лист «ЗП …» — ежедневные начисления ЗП (дата, должность, сумма).
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from date_parse import parse_to_iso
from io_utils import read_excel_sheet
from parse_kzt import parse_amount


def _norm(s: Any) -> str:
    return str(s or "").strip().lower().replace("ё", "е")


def _parse_row_numeric(row: List[Any], year: int, month_nums: List[int]) -> Tuple[Dict[str, float], float]:
    """Колонки 1..12 — месяцы, колонка 13 (индекс 13) — «Итого», если есть."""
    by_month: Dict[str, float] = {}
    for j, mnum in enumerate(month_nums):
        ci = 1 + j
        key = f"{year:04d}-{int(mnum):02d}"
        by_month[key] = float(parse_amount(row[ci] if ci < len(row) else 0))
    total = float(parse_amount(row[13] if len(row) > 13 else 0))
    return by_month, total


def _read_month_nums_from_row0(rows: List[List[Any]]) -> List[int]:
    default = list(range(1, 13))
    if not rows or len(rows[0]) < 13:
        return default
    r0 = rows[0]
    out: List[int] = []
    for i in range(1, 13):
        v = r0[i]
        try:
            out.append(int(float(parse_amount(v))))
        except (ValueError, TypeError):
            return default
    return out if len(out) == 12 else default


def _year_from_sheet_name(sheet: str) -> Optional[int]:
    m = re.search(r"(20\d{2})", str(sheet))
    return int(m.group(1)) if m else None


def parse_dds_sheet(rows: List[List[Any]], sheet_name: str) -> Optional[Dict[str, Any]]:
    if len(rows) < 18:
        return None
    year = _year_from_sheet_name(sheet_name)
    if year is None:
        return None
    month_nums = _read_month_nums_from_row0(rows)

    month_labels: List[str] = []
    inc_header = -1
    for i, r in enumerate(rows[:80]):
        if not r:
            continue
        if _norm(r[0]) == "доходы":
            inc_header = i
            for c in range(1, 13):
                month_labels.append(str(r[c] if c < len(r) else "").strip())
            break
    if inc_header < 0:
        return None

    income_rows: List[Dict[str, Any]] = []
    income_itogo: Optional[Dict[str, Any]] = None
    r = inc_header + 1
    while r < len(rows):
        row = rows[r]
        if not row:
            r += 1
            continue
        c0 = str(row[0] or "").strip()
        if not c0:
            r += 1
            continue
        low = c0.lower()
        if low == "итого":
            bm, tot = _parse_row_numeric(row, year, month_nums)
            income_itogo = {"name": "Итого (доходы)", "byMonth": bm, "total": tot}
            r += 1
            break
        if _norm(c0) == "расходы":
            break
        bm, tot = _parse_row_numeric(row, year, month_nums)
        income_rows.append({"name": c0, "byMonth": bm, "total": tot})
        r += 1

    while r < len(rows):
        row = rows[r]
        if row and _norm(row[0]) == "расходы":
            break
        r += 1
    if r >= len(rows):
        return {
            "year": year,
            "sheet": sheet_name,
            "currency": "KZT",
            "monthNumbers": month_nums,
            "monthLabels": month_labels,
            "incomeRows": income_rows,
            "incomeItogo": income_itogo,
            "expenseRows": [],
            "expenseItogo": None,
        }

    expense_rows: List[Dict[str, Any]] = []
    expense_itogo: Optional[Dict[str, Any]] = None
    r += 1
    while r < len(rows):
        row = rows[r]
        if not row:
            r += 1
            continue
        c0 = str(row[0] or "").strip()
        if not c0:
            r += 1
            continue
        low = c0.lower()
        if low == "итого":
            bm, tot = _parse_row_numeric(row, year, month_nums)
            expense_itogo = {"name": "Итого (расходы)", "byMonth": bm, "total": tot}
            break
        bm, tot = _parse_row_numeric(row, year, month_nums)
        expense_rows.append({"name": c0, "byMonth": bm, "total": tot})
        r += 1

    return {
        "year": year,
        "sheet": sheet_name,
        "currency": "KZT",
        "monthNumbers": month_nums,
        "monthLabels": month_labels,
        "incomeRows": income_rows,
        "incomeItogo": income_itogo,
        "expenseRows": expense_rows,
        "expenseItogo": expense_itogo,
    }


def load_all_dds(path: Union[str, Path]) -> List[Dict[str, Any]]:
    path = Path(path)
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    out: List[Dict[str, Any]] = []
    for name in names:
        if "ддс" not in _norm(name):
            continue
        rows = read_excel_sheet(path, name)
        parsed = parse_dds_sheet(rows, name)
        if parsed:
            out.append(parsed)
    return sorted(out, key=lambda x: (x.get("year") or 0, x.get("sheet") or ""))


def load_payroll_daily(path: Union[str, Path]) -> List[Dict[str, Any]]:
    path = Path(path)
    from openpyxl import load_workbook

    sheet_name = None
    wb = load_workbook(path, read_only=True, data_only=True)
    for n in wb.sheetnames:
        if "зп" in _norm(n):
            sheet_name = n
            break
    wb.close()
    if not sheet_name:
        return []
    rows = read_excel_sheet(path, sheet_name)
    if len(rows) < 2:
        return []
    headers = [str(c or "").strip().lower().replace("ё", "е") for c in rows[0][:12]]
    try:
        i_date = next(i for i, h in enumerate(headers) if h.startswith("дата"))
    except StopIteration:
        i_date = 0
    i_sum = next((i for i, h in enumerate(headers) if "сумма" in h), 5)
    i_role = next((i for i, h in enumerate(headers) if "должност" in h or h == "должность"), 1)
    i_name = next((i for i, h in enumerate(headers) if h in ("имя", "фио")), 2)
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if not r or i_date >= len(r):
            continue
        ds = parse_to_iso(r[i_date])
        if len(ds) < 10:
            continue
        amt = float(parse_amount(r[i_sum] if i_sum < len(r) else 0))
        if amt == 0:
            continue
        role = str(r[i_role] if i_role < len(r) else "").strip()
        nm = str(r[i_name] if i_name < len(r) else "").strip()
        rec: Dict[str, Any] = {"date": ds[:10], "amount": round(amt, 2)}
        if role:
            rec["role"] = role
        if nm:
            rec["employeeName"] = nm
        out.append(rec)
    return out


def load_finance_extensions(path: Union[str, Path]) -> Dict[str, Any]:
    """Всё дополнительное из книги финансов для merge (ДДС, ЗП)."""
    ext: Dict[str, Any] = {}
    dds = load_all_dds(path)
    if dds:
        ext["dds"] = dds
    pay = load_payroll_daily(path)
    if pay:
        ext["financePayrollDaily"] = pay
    return ext


if __name__ == "__main__":
    import json
    import sys

    p = Path(sys.argv[1])
    print(json.dumps(load_finance_extensions(p), ensure_ascii=False, indent=2)[:8000])
