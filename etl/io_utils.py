# -*- coding: utf-8 -*-
"""Чтение CSV (UTF-8 / UTF-8-sig / cp1251), Excel xlsx/xls, таблиц из PDF (pdfplumber)."""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Iterator, List, Optional, Sequence, Union

import pdfplumber


def read_bytes(path: Path) -> bytes:
    return path.read_bytes()


def decode_csv_bytes(raw: bytes) -> str:
    """Пробуем кодировки по очереди (кириллица из Excel в СНГ часто в cp1251)."""
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def sniff_csv_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample[:4096], delimiters=";,")
    except csv.Error:
        class _Semi(csv.Dialect):
            delimiter = ";"
            quotechar = '"'
            escapechar = None
            doublequote = True
            skipinitialspace = True
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL

        return _Semi()


def read_csv_rows(path: Union[str, Path]) -> List[List[str]]:
    """Возвращает строки CSV как списки строк (для маппинга колонок)."""
    p = Path(path)
    text = decode_csv_bytes(read_bytes(p))
    dialect = sniff_csv_dialect(text)
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    return [list(row) for row in reader]


def read_excel_sheet(path: Union[str, Path], sheet: Union[int, str] = 0) -> List[List[Any]]:
    """xlsx через openpyxl; xls через xlrd."""
    p = Path(path)
    suf = p.suffix.lower()
    if suf == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(p))
        sh = book.sheet_by_index(sheet) if isinstance(sheet, int) else book.sheet_by_name(str(sheet))
        return [sh.row_values(rx) for rx in range(sh.nrows)]
    from openpyxl import load_workbook

    wb = load_workbook(str(p), read_only=True, data_only=True)
    names = wb.sheetnames
    if isinstance(sheet, int):
        ws = wb[names[sheet]]
    else:
        ws = wb[sheet]
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def extract_pdf_tables(path: Union[str, Path], pages: Optional[str] = None) -> List[List[List[str]]]:
    """
    Извлекает таблицы со страниц PDF. pages — зарезервировано.
    Возвращает список таблиц (каждая таблица — список строк).
    """
    out: List[List[List[str]]] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for tbl in tables:
                if tbl:
                    out.append([[("" if c is None else str(c)).strip() for c in row] for row in tbl])
    return out


def rows_to_records(rows: Sequence[Sequence[Any]], header_row: int = 0) -> List[dict]:
    if not rows:
        return []
    headers = [str(c or "").strip() for c in rows[header_row]]
    recs = []
    for r in rows[header_row + 1 :]:
        if not any(x is not None and str(x).strip() for x in r):
            continue
        recs.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    return recs
