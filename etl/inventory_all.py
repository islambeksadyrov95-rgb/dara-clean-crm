# -*- coding: utf-8 -*-
"""
Рекурсивный обход папок: метаданные + заголовки CSV + листы xlsx + число таблиц в PDF.
Выход: etl/inventory_report.json
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

_ETL = Path(__file__).resolve().parent
if str(_ETL) not in sys.path:
    sys.path.insert(0, str(_ETL))

SKIP_DIR_NAMES = {".git", "__pycache__", "node_modules", ".venv", "venv"}


def safe_read_csv_header(path: Path, max_len: int = 500) -> Optional[str]:
    try:
        from io_utils import decode_csv_bytes

        raw = path.read_bytes()
        text = decode_csv_bytes(raw)[:max_len]
        line = text.splitlines()[0] if text else ""
        return line[:2000]
    except Exception as e:
        return f"<error: {e}>"


def inspect_xlsx(path: Path) -> Dict[str, Any]:
    out: Dict[str, Any] = {"sheets": [], "first_row_first_sheet": None}
    suf = path.suffix.lower()
    try:
        if suf == ".xls":
            import xlrd

            book = xlrd.open_workbook(str(path))
            out["sheets"] = book.sheet_names()
            if book.sheet_names():
                sh = book.sheet_by_index(0)
                out["first_row_first_sheet"] = [str(sh.cell_value(0, c)) for c in range(min(sh.ncols, 20))]
        else:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            out["sheets"] = list(wb.sheetnames)
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
                row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                out["first_row_first_sheet"] = [str(x) if x is not None else "" for x in (row or ())][:20]
            wb.close()
    except Exception as e:
        out["error"] = str(e)
    return out


def inspect_pdf(path: Path, extract_tables: bool) -> Dict[str, Any]:
    """extract_tables=False: только page_count (быстро). True: полный подсчёт таблиц (медленно на больших PDF)."""
    out: Dict[str, Any] = {"page_count": 0, "table_count": None, "tables_extracted": extract_tables}
    try:
        import pdfplumber

        with pdfplumber.open(str(path)) as pdf:
            out["page_count"] = len(pdf.pages)
            if not extract_tables:
                return out
            out["table_count"] = 0
            for page in pdf.pages:
                tbls = page.extract_tables() or []
                out["table_count"] += len(tbls)
    except Exception as e:
        out["error"] = str(e)
    return out


def walk_roots(roots: List[Path], repo_root: Path, pdf_extract_tables: bool) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for root in roots:
        root = root.resolve()
        if not root.is_dir():
            continue
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            if any(part in SKIP_DIR_NAMES for part in p.parts):
                continue
            rel = str(p.relative_to(repo_root))
            st = p.stat()
            entry: Dict[str, Any] = {
                "path": rel.replace("\\", "/"),
                "name": p.name,
                "ext": p.suffix.lower(),
                "size_bytes": st.st_size,
                "mtime_iso": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat(),
            }
            suf = p.suffix.lower()
            if suf == ".csv":
                entry["csv_header_line1"] = safe_read_csv_header(p)
            elif suf in (".xlsx", ".xls"):
                entry["excel"] = inspect_xlsx(p)
            elif suf == ".pdf":
                entry["pdf"] = inspect_pdf(p, pdf_extract_tables)
            items.append(entry)
    items.sort(key=lambda x: x["path"])
    return items


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--roots",
        nargs="*",
        default=["Продажи", "Финансы", "маркетинг"],
        help="Подпапки от корня репозитория",
    )
    ap.add_argument("--repo", type=Path, default=Path(__file__).resolve().parents[1])
    ap.add_argument("-o", "--out", type=Path, default=Path(__file__).resolve().parent / "inventory_report.json")
    ap.add_argument(
        "--pdf-extract-tables",
        action="store_true",
        help="Медленно: pdfplumber extract_tables на каждой странице каждого PDF (по умолчанию только число страниц)",
    )
    args = ap.parse_args()
    repo = args.repo.resolve()
    roots = [repo / r for r in args.roots]
    items = walk_roots(roots, repo, args.pdf_extract_tables)
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "repo_root": str(repo),
        "roots": [str(r) for r in roots],
        "pdf_extract_tables": args.pdf_extract_tables,
        "file_count": len(items),
        "files": items,
    }
    args.out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {len(items)} files -> {args.out}")


if __name__ == "__main__":
    main()
