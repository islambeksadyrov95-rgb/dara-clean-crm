# -*- coding: utf-8 -*-
"""
Отчёт «по оплатам / по услугам» с блоками «Приемный пункт Машина N / Сайрам» → Excel
с колонкой приёмного пункта и отдельными листами по каждому пункту.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

_RE_DATE_HEADER = re.compile(r"^за\s+(\d{2})\.(\d{2})\.(\d{4})\s*$")
_RE_SUBTOTAL = re.compile(r"^Итого\s+за\s+(\d{2})\.(\d{2})\.(\d{4})\s*:\s*(.*)$", re.I)
_RE_SECTION_TOTAL = re.compile(r"^Итого\s+по\s+(.+?)\s*:\s*(.*)$", re.I)


def _norm(s: str) -> str:
    return (s or "").replace("\xa0", " ").strip()


def _parse_ru_num(tok: str) -> Optional[float]:
    t = _norm(tok).rstrip(",").replace(" ", "")
    if not t:
        return None
    t = t.replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return None


def _nums_from_suffix(suffix: str) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    """После двоеточия в строке «Итого по …» — три числа подряд."""
    suffix = _norm(suffix)
    if not suffix:
        return None, None, None
    toks = suffix.split()
    vals: List[float] = []
    for tok in toks:
        v = _parse_ru_num(tok)
        if v is not None:
            vals.append(v)
    if len(vals) >= 3:
        return vals[0], vals[1], vals[2]
    return None, None, None


def extract_rows_from_pdf(pdf_path: Path) -> Tuple[List[Dict[str, Any]], str]:
    with pdfplumber.open(str(pdf_path)) as pdf:
        parts = []
        for pg in pdf.pages:
            parts.append(pg.extract_text() or "")
        full = "\n".join(parts)
    lines = [_norm(L) for L in full.splitlines()]
    punkt: Optional[str] = None
    day: Optional[str] = None
    rows: List[Dict[str, Any]] = []
    report_title = ""

    for line in lines:
        if not line:
            continue
        if report_title == "" and ("Отчет" in line or "Отчёт" in line):
            report_title = line
            continue
        if "Приемный пункт" in line:
            punkt = line.replace("Приемный пункт", "").strip()
            day = None
            continue
        m = _RE_DATE_HEADER.match(line)
        if m:
            day = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            continue
        if line.startswith("Кол-во") or line.startswith("Группа изделий") or line.startswith("изд."):
            continue

        m = _RE_SECTION_TOTAL.match(line)
        if m:
            qty, nal, bez = _nums_from_suffix(m.group(2))
            rows.append(
                {
                    "приемный_пункт": punkt or "",
                    "дата": "",
                    "группа_изделий": f"ИТОГО ПО ПУНКТУ: {m.group(1).strip()}",
                    "количество_шт": qty,
                    "наличные_руб": nal,
                    "безнал_руб": bez,
                    "тип_строки": "итого_пункт",
                }
            )
            continue

        m = _RE_SUBTOTAL.match(line)
        if m:
            rest = m.group(4)
            nums = re.findall(r"[\d\s\u00a0]+,\d+", rest)
            if len(nums) >= 3:
                qty_s, nal_s, bez_s = nums[-3], nums[-2], nums[-1]
                rows.append(
                    {
                        "приемный_пункт": punkt or "",
                        "дата": day or "",
                        "группа_изделий": f"Итого за {m.group(1)}.{m.group(2)}.{m.group(3)}",
                        "количество_шт": _parse_ru_num(qty_s),
                        "наличные_руб": _parse_ru_num(nal_s),
                        "безнал_руб": _parse_ru_num(bez_s),
                        "тип_строки": "итого_день",
                    }
                )
            continue

        if line.startswith("Итого"):
            continue

        nums = re.findall(r"[\d\s\u00a0]+,\d+", line)
        if len(nums) >= 3:
            qty_s, nal_s, bez_s = nums[-3], nums[-2], nums[-1]
            pos = line.find(nums[-3])
            service = line[:pos].strip() if pos > 0 else line
            if not service:
                continue
            rows.append(
                {
                    "приемный_пункт": punkt or "",
                    "дата": day or "",
                    "группа_изделий": service,
                    "количество_шт": _parse_ru_num(qty_s),
                    "наличные_руб": _parse_ru_num(nal_s),
                    "безнал_руб": _parse_ru_num(bez_s),
                    "тип_строки": "услуга",
                }
            )
            continue

        m2 = re.match(
            r"^(.+?)\s+(\d+)\s+([\d\s\u00a0]+,\d+)\s+([\d\s\u00a0]+,\d+)\s*$",
            line,
        )
        if m2:
            rows.append(
                {
                    "приемный_пункт": punkt or "",
                    "дата": day or "",
                    "группа_изделий": m2.group(1).strip(),
                    "количество_шт": float(m2.group(2)),
                    "наличные_руб": _parse_ru_num(m2.group(3)),
                    "безнал_руб": _parse_ru_num(m2.group(4)),
                    "тип_строки": "услуга",
                }
            )

    return rows, report_title


def _sheet_name(punkt: str) -> str:
    s = re.sub(r'[\[\]:\\/*?]', "_", (punkt or "без_имени").strip())[:31]
    return s or "лист"


def write_xlsx(rows: List[Dict[str, Any]], title: str, out: Path) -> None:
    headers = [
        "приемный_пункт",
        "дата",
        "группа_изделий",
        "количество_шт",
        "наличные_руб",
        "безнал_руб",
        "тип_строки",
    ]
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = _sheet_name("Все_данные")
    if title:
        ws_all.cell(1, 1, title)
        start_r = 3
    else:
        start_r = 1
    for c, h in enumerate(headers, 1):
        ws_all.cell(start_r, c, h)
    r = start_r + 1
    for row in rows:
        for c, h in enumerate(headers, 1):
            ws_all.cell(r, c, row.get(h))
        r += 1
    wsd = wb.create_sheet(title=_sheet_name("Только_услуги"))
    for c, h in enumerate(headers, 1):
        wsd.cell(1, c, h)
    r = 2
    for row in rows:
        if row.get("тип_строки") != "услуга":
            continue
        for c, h in enumerate(headers, 1):
            wsd.cell(r, c, row.get(h))
        r += 1

    punkts = []
    seen = set()
    for row in rows:
        p = row.get("приемный_пункт") or ""
        if p and p not in seen:
            seen.add(p)
            punkts.append(p)

    for p in punkts:
        name = _sheet_name(p)
        if name in wb.sheetnames:
            n = 2
            while f"{name[:28]}_{n}" in wb.sheetnames:
                n += 1
            name = f"{name[:28]}_{n}"
        ws = wb.create_sheet(title=name)
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        rr = 2
        for row in rows:
            if row.get("приемный_пункт") != p:
                continue
            for c, h in enumerate(headers, 1):
                ws.cell(rr, c, row.get(h))
            rr += 1

    for sheet in wb.worksheets:
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18
        sheet.column_dimensions["C"].width = 48

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf", type=Path)
    ap.add_argument("-o", "--out", type=Path, default=None)
    args = ap.parse_args()
    pdf = args.pdf.resolve()
    out = args.out or (pdf.parent / (pdf.stem + "_по_пунктам.xlsx"))
    rows, t = extract_rows_from_pdf(pdf)
    write_xlsx(rows, t, out.resolve())
    print(f"OK: {len(rows)} строк -> {out}")


if __name__ == "__main__":
    main()
