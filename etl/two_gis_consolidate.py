# -*- coding: utf-8 -*-
"""
Сводка выгрузок 2GIS (connections / pagevisits / products daily) в один лист Excel.
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_ETL = Path(__file__).resolve().parent
if str(_ETL) not in sys.path:
    sys.path.insert(0, str(_ETL))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore

from io_utils import read_excel_sheet


def _parse_date_cell(v: Any) -> Optional[str]:
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None


def _find_period_header_row(rows: List[List[Any]]) -> int:
    for i, r in enumerate(rows[:60]):
        if len(r) > 1 and str(r[1] or "").strip() == "Период":
            return i
    return -1


def _sheet_to_daily_fixed_columns(
    rows: List[List[Any]],
    prefix: str,
) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    """Заголовок в строке с «Период» в колонке B, метрики с колонки C."""
    hi = _find_period_header_row(rows)
    if hi < 0:
        return {}, []
    if hi >= len(rows):
        return {}, []
    header = [str(c or "").strip() for c in rows[hi]]
    names: List[str] = []
    for j in range(2, len(header)):
        h = header[j]
        if not h:
            continue
        names.append(f"{prefix}: {h}")
    by_date: Dict[str, Dict[str, Any]] = {}
    for r in rows[hi + 1 :]:
        if len(r) < 3:
            continue
        ds = _parse_date_cell(r[1])
        if not ds:
            continue
        if ds not in by_date:
            by_date[ds] = {}
        for k, name in enumerate(names):
            j = 2 + k
            if j >= len(r):
                continue
            val = r[j]
            if val is None or str(val).strip() == "":
                by_date[ds][name] = 0
            else:
                try:
                    by_date[ds][name] = int(float(str(val).replace(" ", "").replace(",", ".")))
                except (ValueError, TypeError):
                    by_date[ds][name] = val
    return by_date, names


def merge_three_files(
    connections: Path,
    pagevisits: Path,
    products: Path,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    rc = read_excel_sheet(connections, 0)
    rp = read_excel_sheet(pagevisits, 0)
    rq = read_excel_sheet(products, 0)

    d1, n1 = _sheet_to_daily_fixed_columns(rc, "Связи")
    d2, n2 = _sheet_to_daily_fixed_columns(rp, "Страницы")
    d3, n3 = _sheet_to_daily_fixed_columns(rq, "Товары")

    all_dates = sorted(set(d1) | set(d2) | set(d3))
    seen = set()
    ordered_cols = ["Дата"]
    for part in (n1, n2, n3):
        for c in part:
            if c not in seen:
                seen.add(c)
                ordered_cols.append(c)

    rows_out: List[Dict[str, Any]] = []
    for ds in all_dates:
        row: Dict[str, Any] = {name: 0 for name in ordered_cols[1:]}
        row["Дата"] = ds
        for src in (d1, d2, d3):
            if ds in src:
                row.update(src[ds])
        rows_out.append(row)

    return ordered_cols, rows_out


def write_xlsx(path: Path, columns: List[str], rows: List[Dict[str, Any]]) -> None:
    if Workbook is None:
        raise RuntimeError("Нужен пакет openpyxl: pip install openpyxl")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "2GIS ежедневно"
    bold = Font(bold=True)
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(1, c, name)
        cell.font = bold
    for ri, rec in enumerate(rows, start=2):
        for c, name in enumerate(columns, start=1):
            ws.cell(ri, c, rec.get(name))
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Сводка 2GIS daily в один xlsx")
    root = Path(__file__).resolve().parents[1]
    g = root / "маркетинг" / "2Gis"
    ap.add_argument("--connections", type=Path, default=g / "connections-daily.xlsx")
    ap.add_argument("--pagevisits", type=Path, default=g / "pagevisits-daily.xlsx")
    ap.add_argument("--products", type=Path, default=g / "products-daily.xlsx")
    ap.add_argument("--out", type=Path, default=g / "2gis_consolidated_daily.xlsx")
    args = ap.parse_args()

    cols, data = merge_three_files(args.connections, args.pagevisits, args.products)
    write_xlsx(args.out, cols, data)
    print(f"Колонок: {len(cols)}, строк: {len(data)} -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
