# -*- coding: utf-8 -*-
"""
Финансы: ожидается лист с колонками «дата» (или помесячно) и «план» / «сумма плана».
Если структура книги другая — задайте config.yaml (см. merge_build).
По умолчанию: первый лист, строка 1 — заголовки, колонки дата + план.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from date_parse import parse_to_iso
from io_utils import read_excel_sheet, rows_to_records
from parse_kzt import parse_amount


def _month_to_daily(year: int, month: int, total_plan: float) -> List[Dict[str, Any]]:
    import calendar

    _, ndays = calendar.monthrange(year, month)
    per = total_plan / ndays if ndays else total_plan
    out = []
    for d in range(1, ndays + 1):
        iso = f"{year:04d}-{month:02d}-{d:02d}"
        out.append({"date": iso, "amount": round(per, 2)})
    return out


def _looks_like_month_sheet_with_date_row(rows: List[List[Any]]) -> bool:
    """Лист «Октябрь» и аналоги: вторая строка — колонка B = «дата»."""
    if len(rows) < 3:
        return False
    r1 = rows[1]
    if len(r1) < 2:
        return False
    return str(r1[1] or "").strip().lower() == "дата"


def _daily_plans_from_month_sheet(rows: List[List[Any]]) -> List[Dict[str, Any]]:
    """Колонка B — дата; C+D — суммы по дню (шаблон «Октябрь»: заказы + доход по услугам)."""
    daily: List[Dict[str, Any]] = []
    for r in rows[2:]:
        if len(r) < 3:
            continue
        ds = parse_to_iso(r[1])
        if len(ds) < 10:
            continue
        c = parse_amount(r[2]) if len(r) > 2 else 0.0
        d = parse_amount(r[3]) if len(r) > 3 else 0.0
        plan = c + d
        daily.append({"date": ds[:10], "amount": plan})
    return daily


def records_to_daily_plans(records: List[dict]) -> List[Dict[str, Any]]:
    daily: List[Dict[str, Any]] = []
    for row in records:
        keys = {str(k or "").strip().lower().replace("ё", "е"): v for k, v in row.items()}
        d = None
        for k in ("дата", "date", "день"):
            if k in keys and keys[k]:
                d = keys[k]
                break
        ds = parse_to_iso(d)
        if len(ds) < 10:
            continue
        plan = None
        for k in ("план", "plan", "сумма плана", "план выручки", "сумма заказов", "доход"):
            if k in keys and keys[k] is not None:
                plan = parse_amount(keys[k])
                break
        if plan is None:
            continue
        daily.append({"date": ds[:10], "amount": plan})
    return daily


def _daily_total(daily: List[Dict[str, Any]]) -> float:
    return sum(float(d.get("amount") or 0) for d in daily)


def _load_finance_one_sheet(path: Path, sheet: Union[int, str], header_row: int) -> Optional[Dict[str, Any]]:
    rows = read_excel_sheet(path, sheet)
    if _looks_like_month_sheet_with_date_row(rows):
        daily = _daily_plans_from_month_sheet(rows)
        if daily and _daily_total(daily) > 0:
            return {
                "plans": {"daily": daily},
                "meta": {"currency": "KZT", "source": "finance_etl", "sheet": str(sheet)},
            }
    recs = rows_to_records(rows, header_row)
    daily = records_to_daily_plans(recs)
    if not daily:
        for row in rows[1:]:
            if len(row) >= 2 and row[0] and row[1] is not None:
                ds = parse_to_iso(row[0])
                if len(ds) >= 10:
                    daily.append({"date": ds, "amount": parse_amount(row[1])})
    if daily and _daily_total(daily) > 0:
        return {"plans": {"daily": daily}, "meta": {"currency": "KZT", "source": "finance_etl", "sheet": str(sheet)}}
    return None


def _merge_all_month_template_plans(path: Path, names: List[str]) -> Optional[Dict[str, Any]]:
    """Склеивает все листы с шаблоном «строка заголовков — дата в B» (Октябрь, Ноябрь, …)."""
    merged: Dict[str, float] = {}
    used: List[str] = []
    for i, name in enumerate(names):
        try:
            rows = read_excel_sheet(path, i)
        except Exception:
            continue
        if not _looks_like_month_sheet_with_date_row(rows):
            continue
        daily = _daily_plans_from_month_sheet(rows)
        if not daily or _daily_total(daily) <= 0:
            continue
        used.append(name)
        for d in daily:
            ds = d["date"]
            merged[ds] = merged.get(ds, 0.0) + float(d["amount"])
    if not merged:
        return None
    daily_sorted = [{"date": k, "amount": round(merged[k], 2)} for k in sorted(merged.keys())]
    return {
        "plans": {"daily": daily_sorted},
        "meta": {"currency": "KZT", "source": "finance_etl", "monthSheets": used},
    }


def load_finance_plans(path: Union[str, Path], sheet: Union[int, str] = 0, header_row: int = 0) -> Dict[str, Any]:
    path = Path(path)
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    merged = _merge_all_month_template_plans(path, names)
    if merged and merged.get("plans", {}).get("daily"):
        return merged
    try_order: List[Union[int, str]] = []
    seen: set[Union[int, str]] = set()
    if isinstance(sheet, int):
        if 0 <= sheet < len(names):
            try_order.append(sheet)
            seen.add(sheet)
        for i in range(len(names)):
            if i not in seen:
                try_order.append(i)
                seen.add(i)
    else:
        try_order.append(sheet)
        seen.add(sheet)
        for n in names:
            if n not in seen:
                try_order.append(n)
                seen.add(n)
    for sh in try_order:
        try:
            got = _load_finance_one_sheet(path, sh, header_row)
        except Exception:
            continue
        if got and got.get("plans", {}).get("daily"):
            return got
    return {"plans": {"daily": []}, "meta": {"currency": "KZT", "source": "finance_etl"}}


def load_loss_reasons_from_sheet(path: Path, sheet: Union[int, str] = 1) -> List[Dict[str, Any]]:
    """Опциональный лист: причина, количество, сумма."""
    try:
        rows = read_excel_sheet(path, sheet)
    except Exception:
        return []
    recs = rows_to_records(rows, 0)
    out: List[Dict[str, Any]] = []
    for row in recs:
        keys = {str(k or "").strip().lower(): v for k, v in row.items()}
        reason = None
        for k in ("причина", "reason", "основание"):
            if k in keys:
                reason = str(keys[k]).strip()
                break
        if not reason:
            continue
        cnt = int(parse_amount(keys.get("количество", keys.get("count", 0))))
        amt = parse_amount(keys.get("сумма", keys.get("amount", 0)))
        out.append({"reason": reason, "count": cnt, "amount": amt})
    return out


def main():
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("-o", "--out", type=Path, default=Path("finance.json"))
    args = ap.parse_args()
    data = load_finance_plans(args.workbook)
    args.out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {len(data['plans']['daily'])} дней плана → {args.out}")


if __name__ == "__main__":
    main()
