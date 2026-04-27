# -*- coding: utf-8 -*-
"""
Пакетная обработка PDF Яндекс.Метрики из папки: манифест, шаблоны CSV/XLSX для дашборда.

Важно: типичный экспорт «отчёт в PDF» из Метрики — это страницы-картинки без текстового слоя.
Тогда pdfplumber не видит таблицу. Надёжный путь — выгрузить из интерфейса Метрики отчёт «По дням»
в формате CSV/Excel и положить в structured/ или в маркетинг/Яндекс/.

Этот скрипт:
- сканирует *.pdf, считает страницы, проверяет наличие текста;
- сопоставляет имя файла с месяцем (опечатка «Сентябрт» учитывается);
- пишет manifest.json, files.csv, шаблон Excel и краткую инструкцию;
- опционально: --try-text-extract — если в PDF появится текст, вытащит строки по дням.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_ETL = Path(__file__).resolve().parent
if str(_ETL) not in sys.path:
    sys.path.insert(0, str(_ETL))

# Имя файла (stem) → номер месяца
_MONTH_STEM: Dict[str, int] = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "сентябрт": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def _month_from_filename(stem: str) -> Optional[int]:
    s = stem.strip().lower()
    return _MONTH_STEM.get(s)


def _year_for_calendar_month(month: int, season: str) -> int:
    """
    season=sep2025_mar2026: сент–дек 2025, янв–март 2026 (типичный «учётный» год отчётов).
    """
    if season == "sep2025_mar2026":
        if month >= 9:
            return 2025
        return 2026
    from datetime import datetime

    return datetime.now().year


def _inspect_pdf(path: Path) -> Dict[str, Any]:
    import pdfplumber

    out: Dict[str, Any] = {"pages": 0, "text_chars": 0, "has_text_layer": False, "sample": ""}
    with pdfplumber.open(str(path)) as pdf:
        out["pages"] = len(pdf.pages)
        for pg in pdf.pages[:3]:
            t = pg.extract_text() or ""
            out["text_chars"] += len(t)
            if len(out["sample"]) < 400:
                out["sample"] += t[:400]
        out["has_text_layer"] = out["text_chars"] > 50
    return out


def _try_extract_daily_lines(path: Path) -> List[Dict[str, Any]]:
    """Если в PDF есть текст — строки вида DD.MM.YYYY с числами."""
    import pdfplumber

    rows: List[Dict[str, Any]] = []
    _re_day = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})")
    with pdfplumber.open(str(path)) as pdf:
        for pg in pdf.pages:
            for line in (pg.extract_text() or "").splitlines():
                line = line.strip()
                if not line:
                    continue
                m = _re_day.search(line)
                if not m:
                    continue
                iso = f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
                nums = re.findall(r"[\d\s\u00a0]+,\d+|\d+", line.replace("\xa0", " "))
                if len(nums) >= 2:
                    rows.append({"date": iso, "raw": line})
    return rows


def _write_templates(out_dir: Path, files_meta: List[Dict[str, Any]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    out_dir.mkdir(parents=True, exist_ok=True)
    inv = out_dir / "yandex_metrica_manifest.json"
    inv.write_text(json.dumps({"files": files_meta}, ensure_ascii=False, indent=2), encoding="utf-8")

    fc = out_dir / "yandex_metrica_files.csv"
    with fc.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["файл", "месяц", "год", "страниц", "есть_текст_в_pdf", "комментарий"])
        for m in files_meta:
            w.writerow(
                [
                    m["file"],
                    m.get("month") or "",
                    m.get("year") or "",
                    m.get("pages", ""),
                    "да" if m.get("has_text_layer") else "нет",
                    m.get("note", ""),
                ]
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "Инструкция"
    ws["A1"] = "Яндекс.Метрика → дашборд"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "1) Откройте Метрику → нужный счётчик → Отчёты → Например «Посещаемость» → группировка «По дням».",
        "2) Период = нужный месяц. Экспорт: CSV или XLSX (не PDF для автозагрузки).",
        "3) Сохраните файл в папку: маркетинг/Яндекс/  с именем вида: Визиты_2025_12.csv",
        "4) Сборка дашборда: python etl/merge_build.py ... --yandex маркетинг/Яндекс",
        "",
        "PDF из этого набора без текстового слоя программно не разбираются (сканы).",
        "Используйте CSV из интерфейса Метрики — колонки «Дата», «Визиты» (как в etl/marketing_yandex_metrica.py).",
    ]
    for i, L in enumerate(lines, start=3):
        ws.cell(i, 1, L)
    ws.column_dimensions["A"].width = 100

    wsd = wb.create_sheet("Шаблон_по_дням")
    hdr = ["Дата", "Визиты", "Посетители", "источник_файл"]
    for c, h in enumerate(hdr, 1):
        wsd.cell(1, c, h)
    r = 2
    for m in files_meta:
        wsd.cell(r, 4, m["file"])
        r += 1

    wb.save(out_dir / "Метрика_шаблон_и_инструкция.xlsx")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--dir",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "маркетинг" / "Яндекс метрика",
        help="Папка с PDF Метрики",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Папка structured (по умолчанию --dir/structured)",
    )
    ap.add_argument(
        "--season",
        default="sep2025_mar2026",
        choices=["sep2025_mar2026", "current_year"],
        help="Как сопоставлять месяц с годом по имени файла",
    )
    ap.add_argument("--try-text-extract", action="store_true", help="Пробовать вытащить строки с датами из текстового PDF")
    args = ap.parse_args()
    base = args.dir.resolve()
    out_dir = (args.out or (base / "structured")).resolve()
    if not base.is_dir():
        print(f"Нет папки: {base}")
        sys.exit(1)

    files_meta: List[Dict[str, Any]] = []
    extracted_any: List[Dict[str, Any]] = []

    for pdf in sorted(base.glob("*.pdf")):
        stem = pdf.stem
        month = _month_from_filename(stem)
        year = _year_for_calendar_month(month, args.season) if month else None
        info = _inspect_pdf(pdf)
        note = ""
        if not info["has_text_layer"]:
            note = "PDF без текстового слоя (как скан). Нужна выгрузка CSV из Метрики."
        else:
            note = "Есть текст в PDF — можно доработать разбор под ваш шаблон."
        meta: Dict[str, Any] = {
            "file": pdf.name,
            "path": str(pdf).replace("\\", "/"),
            "month": month,
            "year": year,
            "pages": info["pages"],
            "has_text_layer": info["has_text_layer"],
            "text_sample": info["sample"][:200],
            "note": note,
        }
        if args.try_text_extract and info["has_text_layer"]:
            raw_rows = _try_extract_daily_lines(pdf)
            meta["text_lines_guessed"] = len(raw_rows)
            for x in raw_rows:
                x["source_pdf"] = pdf.name
            extracted_any.extend(raw_rows)
        files_meta.append(meta)

    _write_templates(out_dir, files_meta)

    if extracted_any:
        ex = out_dir / "extracted_from_text_pdf.csv"
        with ex.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["date", "raw", "source_pdf"], delimiter=";", extrasaction="ignore")
            w.writeheader()
            w.writerows(extracted_any)

    # Единый CSV-шаблон для ручного заполнения / вставки из Метрики
    tpl = out_dir / "yandex_metrica_daily_template.csv"
    with tpl.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Дата", "Визиты", "Посетители", "комментарий"])
        w.writerow(["ДД.ММ.ГГГГ", "", "", "Скопируйте сюда строки из отчёта Метрики «По дням»"])

    print(f"OK: {len(files_meta)} pdf -> {out_dir}")
    print(f"  {out_dir / 'yandex_metrica_manifest.json'}")
    print(f"  {out_dir / 'Метрика_шаблон_и_инструкция.xlsx'}")


if __name__ == "__main__":
    main()
