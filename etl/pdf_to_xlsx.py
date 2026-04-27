# -*- coding: utf-8 -*-
"""Извлечь таблицы из PDF (pdfplumber) → один файл .xlsx (лист на каждую таблицу)."""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

_ETL = Path(__file__).resolve().parent
if str(_ETL) not in sys.path:
    sys.path.insert(0, str(_ETL))

from io_utils import extract_pdf_tables


def _safe_sheet_name(name: str, idx: int) -> str:
    s = re.sub(r'[\[\]:\\/*?]', "_", name)[:31]
    s = s.strip() or f"T{idx}"
    return s


MAX_SHEETS = 60


def _write_merged(wb, tables: list) -> None:
    ws = wb.create_sheet(title=_safe_sheet_name("Все_таблицы", 1))
    row = 1
    for ti, tbl in enumerate(tables, start=1):
        ws.cell(row, 1, f"--- Таблица {ti} ---")
        row += 1
        for tbl_row in tbl:
            for col, val in enumerate(tbl_row, start=1):
                ws.cell(row, col, val)
            row += 1
        row += 1


def pdf_to_xlsx(pdf_path: Path, xlsx_path: Path, merge_if_many: bool = True) -> int:
    from openpyxl import Workbook

    tables = extract_pdf_tables(pdf_path)
    wb = Workbook()
    wb.remove(wb.active)
    if not tables:
        ws = wb.create_sheet(title="NoTables")
        ws.cell(1, 1, "Таблицы не найдены (pdfplumber extract_tables).")
        wb.save(xlsx_path)
        return 0
    if merge_if_many and len(tables) > MAX_SHEETS:
        _write_merged(wb, tables)
    else:
        for i, tbl in enumerate(tables, start=1):
            title = _safe_sheet_name(f"T{i}", i)
            base = title
            n = 1
            while title in wb.sheetnames:
                n += 1
                title = _safe_sheet_name(f"{base[:28]}_{n}", i)
            ws = wb.create_sheet(title=title)
            for r, row in enumerate(tbl, start=1):
                for c, val in enumerate(row, start=1):
                    ws.cell(r, c, val)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return len(tables)


def main():
    ap = argparse.ArgumentParser(description="PDF → Excel (таблицы по листам)")
    ap.add_argument("pdf", type=Path, nargs="+", help="Один или несколько PDF")
    ap.add_argument(
        "-o",
        "--out-dir",
        type=Path,
        default=None,
        help="Папка для xlsx (по умолчанию — рядом с каждым PDF)",
    )
    ap.add_argument(
        "--always-split-sheets",
        action="store_true",
        help=f"Всегда отдельный лист на таблицу (иначе при >{MAX_SHEETS} таблиц — один лист «Все_таблицы»)",
    )
    args = ap.parse_args()
    for pdf in args.pdf:
        p = pdf.resolve()
        if not p.exists():
            print(f"SKIP (нет файла): {p}")
            continue
        out_dir = args.out_dir.resolve() if args.out_dir else p.parent
        out = out_dir / (p.stem + ".xlsx")
        n = pdf_to_xlsx(p, out, merge_if_many=not args.always_split_sheets)
        mode = "merged" if (not args.always_split_sheets and n > MAX_SHEETS) else "sheets"
        print(f"OK: {n} tables ({mode}) -> {out}")


if __name__ == "__main__":
    main()
