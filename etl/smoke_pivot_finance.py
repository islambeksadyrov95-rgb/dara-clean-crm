# -*- coding: utf-8 -*-
"""Быстрые проверки: свод продаж, склейка листов финансов. Запуск: python etl/smoke_pivot_finance.py"""
from __future__ import annotations

import sys
from pathlib import Path

_ETL = Path(__file__).resolve().parent
if str(_ETL) not in sys.path:
    sys.path.insert(0, str(_ETL))

from finance_workbook import _merge_all_month_template_plans, load_finance_plans
from io_utils import read_excel_sheet as rex
from sales_from_sheet import parse_pivot_payment_groups


def main() -> int:
    root = _ETL.parent
    sales = root / "Продажи" / "Отчет по оплатам групп услуг.xlsx"
    fin = root / "Финансы" / "Финансы DaraClean 2025-2026.xlsx"
    if not sales.exists() or not fin.exists():
        print("SKIP: нет файлов продаж/финансов в проекте")
        return 0
    rows = rex(sales, 0)
    pv = parse_pivot_payment_groups(rows)
    assert len(pv) > 0, "pivot: ожидались строки"
    from openpyxl import load_workbook

    wb = load_workbook(fin, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    m = _merge_all_month_template_plans(fin, names)
    assert m and m["plans"]["daily"], "finance merge: ожидались дни плана"
    d2 = load_finance_plans(fin)
    assert len(d2["plans"]["daily"]) == len(m["plans"]["daily"])
    print("OK pivot rows:", len(pv), "| finance days:", len(m["plans"]["daily"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
