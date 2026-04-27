# -*- coding: utf-8 -*-
"""Список листов xlsx/xls — для выбора листа плана в Финансы DaraClean."""
from __future__ import annotations

import sys
from pathlib import Path


def main():
    p = Path(sys.argv[1] if len(sys.argv) > 1 else ".")
    if not p.exists():
        print("Файл не найден:", p, file=sys.stderr)
        sys.exit(1)
    suf = p.suffix.lower()
    if suf == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(p))
        for i, name in enumerate(book.sheet_names()):
            print(f"{i}\t{name}")
    else:
        from openpyxl import load_workbook

        wb = load_workbook(str(p), read_only=True)
        for i, name in enumerate(wb.sheetnames):
            print(f"{i}\t{name}")
        wb.close()


if __name__ == "__main__":
    main()
