"""
build_planning_sheets.py
Создаёт 3 новых листа в Плановый расчет.xlsx:
  1. Воронка       — маркетинговая воронка по месяцам
  2. План продаж   — дневной план на выбранный месяц
  3. КПД Логистики — эффективность 3 машин
"""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC_PATH = 'D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет.xlsx'

wb = openpyxl.load_workbook(SRC_PATH)

# ─── Стили ───────────────────────────────────────────────────────────────────
def S(l='thin',r='thin',t='thin',b='thin',lc='CBD5E1',rc='CBD5E1',tc='CBD5E1',bc='CBD5E1'):
    def sd(s,c): return Side(style=s,color=c) if s else Side(style=None)
    return Border(left=sd(l,lc),right=sd(r,rc),top=sd(t,tc),bottom=sd(b,bc))

def F(bold=False,size=9,color='1E293B',italic=False):
    return Font(bold=bold,size=size,color=color,name='Calibri',italic=italic)
def aL(i=0): return Alignment(horizontal='left',  vertical='center',indent=i, wrap_text=False)
def aR():    return Alignment(horizontal='right', vertical='center')
def aC(wrap=False): return Alignment(horizontal='center',vertical='center',wrap_text=wrap)

HEADER_FILL  = PatternFill('solid',fgColor='1E293B')
WHITE_FILL   = PatternFill('solid',fgColor='FFFFFF')
GRAY_FILL    = PatternFill('solid',fgColor='F8FAFC')
YELLOW_FILL  = PatternFill('solid',fgColor='FEFCE8')  # ввод
GREEN_FILL   = PatternFill('solid',fgColor='DCFCE7')
BLUE_FILL    = PatternFill('solid',fgColor='DBEAFE')
PURPLE_FILL  = PatternFill('solid',fgColor='EDE9FE')
ORANGE_FILL  = PatternFill('solid',fgColor='FEF3C7')
RED_FILL     = PatternFill('solid',fgColor='FEE2E2')
TOTAL_FILL   = PatternFill('solid',fgColor='E2E8F0')

def sc(ws,row,col,val=None,fml=None,font=None,fill=None,align=None,brd=None,nfmt=None):
    c = ws.cell(row=row,column=col)
    c.value = fml if fml else val
    if font:  c.font  = font
    if fill:  c.fill  = fill
    if align: c.alignment = align
    if brd:   c.border= brd
    if nfmt:  c.number_format = nfmt
    return c

MONTHS_RU = ['Январь','Февраль','Март','Апрель','Май','Июнь',
             'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ: ВОРОНКА
# ═══════════════════════════════════════════════════════════════════════════
if 'Воронка' in wb.sheetnames: del wb['Воронка']
ws = wb.create_sheet('Воронка')

# Актуальные данные за Q1 2026 (из CRM / ДДС)
# Формат: [месяц, маркетинг_расход, обращений, продаж, ср_чек, ltv_клиент]
# LTV = ср_чек × 1.8 (среднее кол-во заказов на клиента в год)
ACTUAL_DATA = {
    'Январь':  {'mkt': 1_135_000, 'leads': 390,  'sales': 155, 'avg_check': 24_100, 'ltv_orders': 1.8},
    'Февраль': {'mkt': 1_285_000, 'leads': 420,  'sales': 171, 'avg_check': 24_300, 'ltv_orders': 1.8},
    'Март':    {'mkt': 1_395_000, 'leads': 503,  'sales': 213, 'avg_check': 25_100, 'ltv_orders': 1.8},
}

# Колонки:
# A=Месяц, B=Маркетинг₸, C=Обращений, D=Продаж, E=Конверсия%, F=CAC₸,
# G=Ср чек₸, H=LTV₸(1 клиент), I=Выручка₸, J=ROMI%

ws.column_dimensions['A'].width = 14
for col_l, w in zip(['B','C','D','E','F','G','H','I','J'], [14,11,11,12,12,12,14,14,12]):
    ws.column_dimensions[col_l].width = w

r = 1
# Мега-заголовок
ws.merge_cells(f'A{r}:J{r}')
sc(ws,r,1, val='МАРКЕТИНГОВАЯ ВОРОНКА — Факт + Планирование',
   font=F(bold=True,size=13,color='FFFFFF'), fill=HEADER_FILL, align=aC())
ws.row_dimensions[r].height = 28
r+=1

# Подзаголовок
ws.merge_cells(f'A{r}:J{r}')
sc(ws,r,1,
   val='Жёлтые ячейки = вводите плановые значения. Зелёные = факт (Q1 2026). LTV = Ср.чек × Кол-во заказов/год на клиента.',
   font=F(size=9,color='64748B',italic=True),
   fill=GRAY_FILL, align=aL())
ws.row_dimensions[r].height = 14
r+=1

# ── СЕКЦИЯ ФАКТ ──────────────────────────────────────────────────────────────
ws.merge_cells(f'A{r}:J{r}')
sc(ws,r,1, val='📊 ФАКТ — Q1 2026',
   font=F(bold=True,size=10,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='065F46'), align=aL(1))
ws.row_dimensions[r].height = 18
r+=1

# Заголовки таблицы факт
HDR = ['Месяц','Маркетинг ₸','Обращений','Продаж','Конверсия %','CAC ₸',
       'Ср. чек ₸','LTV (1 клиент) ₸','Выручка ₸','ROMI %']
brd_h = S(t='medium',b='medium',tc='475569',bc='475569')
for ci, h in enumerate(HDR, start=1):
    sc(ws,r,ci, val=h, font=F(bold=True,size=9,color='FFFFFF'),
       fill=HEADER_FILL, align=aC(True), brd=brd_h)
ws.row_dimensions[r].height = 30
ws.freeze_panes = 'B5'
r+=1

fact_start = r
for i, (month, d) in enumerate(ACTUAL_DATA.items()):
    rf = GREEN_FILL if i%2==0 else PatternFill('solid',fgColor='F0FDF4')
    mkt = d['mkt']; leads = d['leads']; sales = d['sales']
    avg = d['avg_check']; ltv_k = d['ltv_orders']
    conv = sales/leads if leads else 0
    cac  = mkt/sales if sales else 0
    ltv  = avg * ltv_k
    rev  = sales * avg
    romi = (rev - mkt) / mkt if mkt else 0

    row_vals = [month, mkt, leads, sales, conv, cac, avg, ltv, rev, romi]
    fmts = [None,'#,##0','#,##0','#,##0','0.0%','#,##0','#,##0','#,##0','#,##0','0%']
    for ci, (v, nf) in enumerate(zip(row_vals, fmts), start=1):
        sc(ws, r, ci, val=v, font=F(size=9,bold=(ci==1)),
           fill=rf, align=(aL() if ci==1 else aR()), nfmt=nf)
    ws.row_dimensions[r].height = 16
    r+=1
fact_end = r-1

# Итого факт Q1
brd_t = S(l='medium',r='medium',t='medium',b='medium',lc='94A3B8',rc='94A3B8',tc='475569',bc='475569')
sc(ws,r,1, val='ИТОГО Q1', font=F(bold=True,size=9), fill=TOTAL_FILL, align=aL(), brd=brd_t)
sc(ws,r,2, fml=f'=SUM(B{fact_start}:B{fact_end})', font=F(bold=True,size=9),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
sc(ws,r,3, fml=f'=SUM(C{fact_start}:C{fact_end})', font=F(bold=True,size=9),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
sc(ws,r,4, fml=f'=SUM(D{fact_start}:D{fact_end})', font=F(bold=True,size=9),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
sc(ws,r,5, fml=f'=IF(C{r}>0,D{r}/C{r},"")', font=F(bold=True,size=9,color='065F46'),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='0.0%')
sc(ws,r,6, fml=f'=IF(D{r}>0,B{r}/D{r},"")', font=F(bold=True,size=9),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
sc(ws,r,7, fml=f'=IF(D{r}>0,I{r}/D{r},"")', font=F(bold=True,size=9),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
sc(ws,r,8, val='', fill=TOTAL_FILL, brd=brd_t)
sc(ws,r,9, fml=f'=SUM(I{fact_start}:I{fact_end})', font=F(bold=True,size=9),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
sc(ws,r,10, fml=f'=IF(B{r}>0,(I{r}-B{r})/B{r},"")', font=F(bold=True,size=9,color='1D4ED8'),
   fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='0%')
ws.row_dimensions[r].height = 18
r+=2

# ── СЕКЦИЯ ПЛАН ──────────────────────────────────────────────────────────────
ws.merge_cells(f'A{r}:J{r}')
sc(ws,r,1, val='🎯 ПЛАНОВАЯ ВОРОНКА — Апрель–Декабрь 2026 (вводите значения вручную)',
   font=F(bold=True,size=10,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='1D4ED8'), align=aL(1))
ws.row_dimensions[r].height = 18
r+=1

# Заголовки план
for ci, h in enumerate(HDR, start=1):
    sc(ws,r,ci, val=h, font=F(bold=True,size=9,color='FFFFFF'),
       fill=PatternFill('solid',fgColor='1E3A5F'), align=aC(True), brd=brd_h)
ws.row_dimensions[r].height = 30
plan_hdr_row = r
r+=1

plan_start = r
for i, month in enumerate(MONTHS_RU[3:], start=1):  # Апрель–Декабрь
    rf = YELLOW_FILL if i%2==1 else PatternFill('solid',fgColor='FEFCE8')
    # Ячейки ввода: B=маркетинг, C=обращений, G=ср_чек, H=LTV_коэф (заказов/год)
    # Расчётные: D=продаж=C*conv, E=conv (формула), F=CAC, I=выручка, J=ROMI
    cur_r = r
    sc(ws,cur_r,1, val=month, font=F(bold=True,size=9), fill=rf, align=aL())

    # B=маркетинг (ввод)
    sc(ws,cur_r,2, val=1_500_000, font=F(size=9,color='5F4C0A'),
       fill=YELLOW_FILL, align=aR(), nfmt='#,##0',
       brd=S(l='medium',r='medium',lc='CA8A04',rc='CA8A04'))
    # C=обращений (ввод)
    sc(ws,cur_r,3, val=500, font=F(size=9,color='5F4C0A'),
       fill=YELLOW_FILL, align=aR(), nfmt='#,##0',
       brd=S(l='medium',r='medium',lc='CA8A04',rc='CA8A04'))
    # D=продаж (ввод, т.к. конверсия может меняться)
    sc(ws,cur_r,4, val=255, font=F(size=9,color='5F4C0A'),
       fill=YELLOW_FILL, align=aR(), nfmt='#,##0',
       brd=S(l='medium',r='medium',lc='CA8A04',rc='CA8A04'))
    # E=конверсия (расчёт)
    sc(ws,cur_r,5, fml=f'=IF(C{cur_r}>0,D{cur_r}/C{cur_r},"")',
       font=F(size=9,color='065F46'), fill=GREEN_FILL, align=aR(), nfmt='0.0%')
    # F=CAC (расчёт)
    sc(ws,cur_r,6, fml=f'=IF(D{cur_r}>0,B{cur_r}/D{cur_r},"")',
       font=F(size=9), fill=GREEN_FILL, align=aR(), nfmt='#,##0')
    # G=ср_чек (ввод)
    sc(ws,cur_r,7, val=26_000, font=F(size=9,color='5F4C0A'),
       fill=YELLOW_FILL, align=aR(), nfmt='#,##0',
       brd=S(l='medium',r='medium',lc='CA8A04',rc='CA8A04'))
    # H=LTV ₸ (ввод: кол-во заказов/год, формула = G×H_коэф)
    # Оставим ввод LTV в ₸ напрямую
    sc(ws,cur_r,8, fml=f'=G{cur_r}*1.8',
       font=F(size=9,color='7C3AED'), fill=PURPLE_FILL, align=aR(), nfmt='#,##0')
    # I=выручка (расчёт)
    sc(ws,cur_r,9, fml=f'=D{cur_r}*G{cur_r}',
       font=F(size=9,bold=True,color='065F46'), fill=GREEN_FILL, align=aR(), nfmt='#,##0')
    # J=ROMI
    sc(ws,cur_r,10, fml=f'=IF(B{cur_r}>0,(I{cur_r}-B{cur_r})/B{cur_r},"")',
       font=F(size=9,color='1D4ED8'), fill=BLUE_FILL, align=aR(), nfmt='0%')

    ws.row_dimensions[cur_r].height = 16
    r+=1
plan_end = r-1

# Итого план
sc(ws,r,1, val='ИТОГО ПЛАН', font=F(bold=True,size=9), fill=TOTAL_FILL, align=aL(), brd=brd_t)
for ci, fml in enumerate([
    f'=SUM(B{plan_start}:B{plan_end})',
    f'=SUM(C{plan_start}:C{plan_end})',
    f'=SUM(D{plan_start}:D{plan_end})',
    f'=IF(C{r}>0,D{r}/C{r},"")',
    f'=IF(D{r}>0,B{r}/D{r},"")',
    f'=IF(D{r}>0,I{r}/D{r},"")',
    '',
    f'=SUM(I{plan_start}:I{plan_end})',
    f'=IF(B{r}>0,(I{r}-B{r})/B{r},"")',
], start=2):
    nf = '0.0%' if ci==5 else ('0%' if ci==10 else '#,##0')
    sc(ws,r,ci, fml=fml if fml else None,
       font=F(bold=True,size=9), fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt=nf)
ws.row_dimensions[r].height = 18
r+=2

# Подсказка
ws.merge_cells(f'A{r}:J{r}')
sc(ws,r,1,
   val='ℹ  Жёлтые ячейки — ввод: Маркетинг ₸, Кол-во обращений, Продаж, Ср. чек. '
       'Конверсия, CAC, LTV, Выручка, ROMI считаются автоматически. '
       'LTV = Ср.чек × 1.8 заказов/год (среднее по клиентской базе).',
   font=F(size=9,color='64748B',italic=True),
   fill=GRAY_FILL, align=aL())
ws.row_dimensions[r].height = 28

print(f'✓ Лист "Воронка" создан, строк: {r}')

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ: ПЛАН ПРОДАЖ
# ═══════════════════════════════════════════════════════════════════════════
if 'План продаж' in wb.sheetnames: del wb['План продаж']
ws = wb.create_sheet('План продаж')

# Колонки: A=Дата, B=День недели, C=Коэф нагрузки, D=Заказов план,
# E=Выручка план ₸, F=Маркетинг план ₸, G=Обращений нужно (=D/0.51)
ws.column_dimensions['A'].width = 13
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 16

r = 1
ws.merge_cells('A1:G1')
sc(ws,r,1, val='ПЛАН ПРОДАЖ — Дневная разбивка по месяцу',
   font=F(bold=True,size=13,color='FFFFFF'), fill=HEADER_FILL, align=aC())
ws.row_dimensions[r].height = 28
r+=1

# Зона параметров (строки 2-6)
ws.merge_cells('A2:B2')
sc(ws,2,1, val='Выберите месяц:', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,2,3, val='Апрель', font=F(bold=True,size=11,color='1D4ED8'),
   fill=YELLOW_FILL, align=aC(),
   brd=S(l='medium',r='medium',t='medium',b='medium',lc='3B82F6',rc='3B82F6',tc='3B82F6',bc='3B82F6'))
sc(ws,2,4, val='← выберите из списка', font=F(size=9,color='6B7280',italic=True), align=aL())

# DataValidation для месяца
dv_month = DataValidation(type='list',
    formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"',
    allow_blank=False)
ws.add_data_validation(dv_month)
dv_month.sqref = 'C2'

ws.row_dimensions[2].height = 22

ws.merge_cells('A3:B3')
sc(ws,3,1, val='Заказов в месяц (план):', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,3,3, val=290, font=F(bold=True,size=11,color='065F46'),
   fill=YELLOW_FILL, align=aC(),
   brd=S(l='medium',r='medium',t='medium',b='medium',lc='3B82F6',rc='3B82F6',tc='3B82F6',bc='3B82F6'))
sc(ws,3,4, val='шт — итого за месяц', font=F(size=9,color='6B7280',italic=True), align=aL())
ws.row_dimensions[3].height = 20

ws.merge_cells('A4:B4')
sc(ws,4,1, val='Ср. чек (план):', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,4,3, val=26_000, font=F(bold=True,size=11,color='065F46'),
   fill=YELLOW_FILL, align=aC(),
   brd=S(l='medium',r='medium',t='medium',b='medium',lc='3B82F6',rc='3B82F6',tc='3B82F6',bc='3B82F6'))
sc(ws,4,4, val='₸ — средний чек заказа', font=F(size=9,color='6B7280',italic=True), align=aL())
ws.row_dimensions[4].height = 20

ws.merge_cells('A5:B5')
sc(ws,5,1, val='Маркетинг в месяц (план):', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,5,3, val=1_500_000, font=F(bold=True,size=11,color='7C3AED'),
   fill=YELLOW_FILL, align=aC(),
   brd=S(l='medium',r='medium',t='medium',b='medium',lc='7C3AED',rc='7C3AED',tc='7C3AED',bc='7C3AED'))
sc(ws,5,4, val='₸ — рекламный бюджет', font=F(size=9,color='6B7280',italic=True), align=aL())
ws.row_dimensions[5].height = 20

ws.merge_cells('A6:B6')
sc(ws,6,1, val='Конверсия (обращение → заказ):', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,6,3, val=0.51, font=F(bold=True,size=11,color='B45309'),
   fill=YELLOW_FILL, align=aC(), nfmt='0%',
   brd=S(l='medium',r='medium',t='medium',b='medium',lc='B45309',rc='B45309',tc='B45309',bc='B45309'))
sc(ws,6,4, val='% — исторический факт (51%)', font=F(size=9,color='6B7280',italic=True), align=aL())
ws.row_dimensions[6].height = 20

# Ссылочные ячейки для формул
ORDERS_CELL  = '$C$3'
AVG_CHECK    = '$C$4'
MKT_MONTH    = '$C$5'
CONV_CELL    = '$C$6'

r = 7
# Расчётные итоги
ws.merge_cells('A7:B7')
sc(ws,7,1, val='Выручка месяц (план):', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,7,3, fml=f'={ORDERS_CELL}*{AVG_CHECK}',
   font=F(bold=True,size=11,color='065F46'),
   fill=GREEN_FILL, align=aC(), nfmt='#,##0')
sc(ws,7,4, val='₸ = Заказы × Ср.чек', font=F(size=9,color='6B7280',italic=True), align=aL())
ws.row_dimensions[7].height = 20

ws.merge_cells('A8:B8')
sc(ws,8,1, val='Обращений нужно:', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,8,3, fml=f'=ROUND({ORDERS_CELL}/{CONV_CELL},0)',
   font=F(bold=True,size=11,color='B45309'),
   fill=ORANGE_FILL, align=aC(), nfmt='#,##0')
sc(ws,8,4, val='шт = Заказы / Конверсия', font=F(size=9,color='6B7280',italic=True), align=aL())
ws.row_dimensions[8].height = 20

r = 9
# Разделитель
ws.merge_cells(f'A{r}:G{r}')
sc(ws,r,1, fill=PatternFill('solid',fgColor='1E293B'))
ws.row_dimensions[r].height = 5
r+=1

# Заголовки таблицы дней
HDR_DAY = ['Дата','День недели','Коэф.\nнагрузки','Заказов\nплан','Выручка\nплан ₸',
           'Маркетинг\nплан ₸','Обращений\nнужно']
brd_h = S(t='medium',b='medium',tc='475569',bc='475569')
for ci, h in enumerate(HDR_DAY, start=1):
    sc(ws,r,ci, val=h, font=F(bold=True,size=9,color='FFFFFF'),
       fill=HEADER_FILL, align=aC(True), brd=brd_h)
ws.row_dimensions[r].height = 30
ws.freeze_panes = f'A{r+1}'
r+=1

table_start = r

# Генерим строки для 31 дня (максимум в месяце)
# Формулы используют DATE(2026, MATCH(C2, список_месяцев, 0)+3, день)
# Но для упрощения — пишем для апреля-декабря 2026 через статику
# Используем IFERROR + DATE чтобы скрыть несуществующие дни (29-31)

MONTH_NUM_FORMULA = 'MATCH($C$2,{"Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"},0)'
DAYS_RU = ['', 'Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
# Excel WEEKDAY: 1=Sun, 2=Mon ... 7=Sat → mode 2: 1=Mon ... 7=Sun

for day_n in range(1, 32):
    cur_r = r
    # Дата: если день > кол-ва дней в месяце → пусто
    date_fml = f'=IFERROR(IF(DAY(DATE(2026,{MONTH_NUM_FORMULA},{day_n}))={day_n},DATE(2026,{MONTH_NUM_FORMULA},{day_n}),""),"")'
    sc(ws,cur_r,1, fml=date_fml,
       font=F(size=9), fill=WHITE_FILL if day_n%2==1 else GRAY_FILL,
       align=aL(), nfmt='DD.MM.YYYY')

    # День недели
    dow_fml = f'=IFERROR(IF(A{cur_r}="","",CHOOSE(WEEKDAY(A{cur_r},2),"Пн","Вт","Ср","Чт","Пт","Сб","Вс")),"")'
    # Определяем цвет: Сб/Вс = серее
    sc(ws,cur_r,2, fml=dow_fml,
       font=F(size=9), fill=WHITE_FILL if day_n%2==1 else GRAY_FILL, align=aC())

    # Коэф нагрузки: Пн-Пт=1.0, Сб=0.6, Вс=0.3
    coef_fml = (f'=IFERROR(IF(A{cur_r}="","",IF(WEEKDAY(A{cur_r},2)=6,0.6,'
                f'IF(WEEKDAY(A{cur_r},2)=7,0.3,1.0))),"")')
    sc(ws,cur_r,3, fml=coef_fml,
       font=F(size=9,color='374151'),
       fill=WHITE_FILL if day_n%2==1 else GRAY_FILL,
       align=aC(), nfmt='0.0')

    # Заказов план = Итого_месяц × коэф / SUMIF_коэф
    # Упрощённо: orders_month × coef / SUM(все коэфы)
    # SUM коэфов = сумма по всей таблице
    coef_sum_range = f'C{table_start}:C{table_start+30}'
    orders_fml = (f'=IFERROR(IF(A{cur_r}="","",ROUND({ORDERS_CELL}*C{cur_r}/'
                  f'SUMIF({coef_sum_range},">"&0),0)),"")')
    sc(ws,cur_r,4, fml=orders_fml,
       font=F(size=9,bold=True,color='065F46'),
       fill=GREEN_FILL if day_n%2==1 else PatternFill('solid',fgColor='DCFCE7'),
       align=aR(), nfmt='#,##0')

    # Выручка = заказов × ср_чек
    rev_fml = f'=IFERROR(IF(D{cur_r}="","",D{cur_r}*{AVG_CHECK}),"")'
    sc(ws,cur_r,5, fml=rev_fml,
       font=F(size=9,color='1E40AF'),
       fill=BLUE_FILL if day_n%2==1 else PatternFill('solid',fgColor='EFF6FF'),
       align=aR(), nfmt='#,##0')

    # Маркетинг = бюджет_месяц × коэф / sum_коэф
    mkt_fml = (f'=IFERROR(IF(A{cur_r}="","",ROUND({MKT_MONTH}*C{cur_r}/'
               f'SUMIF({coef_sum_range},">"&0),0)),"")')
    sc(ws,cur_r,6, fml=mkt_fml,
       font=F(size=9,color='7C3AED'),
       fill=PURPLE_FILL if day_n%2==1 else PatternFill('solid',fgColor='F5F3FF'),
       align=aR(), nfmt='#,##0')

    # Обращений нужно = заказов_день / конверсия
    leads_fml = f'=IFERROR(IF(D{cur_r}="","",ROUND(D{cur_r}/{CONV_CELL},0)),"")'
    sc(ws,cur_r,7, fml=leads_fml,
       font=F(size=9,color='B45309'),
       fill=ORANGE_FILL if day_n%2==1 else PatternFill('solid',fgColor='FEF3C7'),
       align=aR(), nfmt='#,##0')

    ws.row_dimensions[cur_r].height = 16
    r+=1

table_end = r-1

# Итоги
sc(ws,r,1, val='ИТОГО', font=F(bold=True,size=10), fill=TOTAL_FILL,
   align=aL(), brd=brd_t)
sc(ws,r,2, val='', fill=TOTAL_FILL, brd=brd_t)
sc(ws,r,3, val='', fill=TOTAL_FILL, brd=brd_t)
for ci, col in enumerate(['D','E','F','G'], start=4):
    sc(ws,r,ci, fml=f'=SUMIF({col}{table_start}:{col}{table_end},">"&0)',
       font=F(bold=True,size=10), fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
ws.row_dimensions[r].height = 20
r+=2

ws.merge_cells(f'A{r}:G{r}')
sc(ws,r,1,
   val='ℹ  Выберите месяц в C2. Задайте план: заказы, ср.чек, маркетинг, конверсию. '
       'Дни распределяются пропорционально: Пн-Пт × 1.0, Сб × 0.6, Вс × 0.3. '
       'Несуществующие дни (29-31) скрываются автоматически.',
   font=F(size=9,color='64748B',italic=True),
   fill=GRAY_FILL, align=aL())
ws.row_dimensions[r].height = 28

print(f'✓ Лист "План продаж" создан, строк: {r}')

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ: КПД ЛОГИСТИКИ
# ═══════════════════════════════════════════════════════════════════════════
if 'КПД Логистики' in wb.sheetnames: del wb['КПД Логистики']
ws = wb.create_sheet('КПД Логистики')

# Параметры логистики:
# 3 машины × 40 адресов/день = 120 адресов макс
# 20 забор + 20 доставка на машину
# КПД логистики = факт_адресов / (3×40) × 100%
# КПД цеха = факт_кв.м / максимальная мощность цеха × 100%
# КПД оборота = факт_выручка / план_выручка × 100%

ws.column_dimensions['A'].width = 16
for col_l, w in zip(['B','C','D','E','F','G','H'], [14,14,14,14,14,14,14]):
    ws.column_dimensions[col_l].width = w

r = 1
ws.merge_cells('A1:H1')
sc(ws,r,1, val='КПД ЛОГИСТИКИ — Эффективность использования ресурсов',
   font=F(bold=True,size=13,color='FFFFFF'), fill=HEADER_FILL, align=aC())
ws.row_dimensions[r].height = 28
r+=1

# ── Параметры ────────────────────────────────────────────────────────────────
ws.merge_cells(f'A{r}:H{r}')
sc(ws,r,1, val='⚙️  ПАРАМЕТРЫ (введите свои значения)',
   font=F(bold=True,size=10,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='374151'), align=aL(1))
ws.row_dimensions[r].height = 18
r+=1

params = [
    ('Кол-во машин:', 3, 'шт'),
    ('Адресов/машина/день (план):', 40, 'адр'),
    ('из них: забор (план):', 20, 'адр'),
    ('из них: доставка (план):', 20, 'адр'),
    ('Макс. мощность цеха (кв.м/мес):', 5000, 'кв.м'),
    ('Конверсия заказ→выполнен:', 0.97, '%'),
]
param_rows = {}
for label, val, unit in params:
    ws.merge_cells(f'A{r}:C{r}')
    sc(ws,r,1, val=label, font=F(size=9,bold=True), fill=GRAY_FILL, align=aL(1))
    c_val = ws.cell(row=r, column=4)
    c_val.value = val
    c_val.font = Font(bold=True,size=11,color='1D4ED8',name='Calibri')
    c_val.fill = YELLOW_FILL
    c_val.alignment = Alignment(horizontal='center',vertical='center')
    c_val.border = S(l='medium',r='medium',t='thin',b='thin',lc='3B82F6',rc='3B82F6')
    if unit == '%': c_val.number_format = '0%'
    elif unit in ('кв.м','адр','шт'): c_val.number_format = '#,##0'
    sc(ws,r,5, val=unit, font=F(size=9,color='64748B'), fill=GRAY_FILL, align=aL())
    param_rows[label] = r
    ws.row_dimensions[r].height = 18
    r+=1

# Ссылочные ячейки параметров
TRUCKS_ROW   = param_rows['Кол-во машин:']
ADDR_ROW     = param_rows['Адресов/машина/день (план):']
PICKUP_ROW   = param_rows['из них: забор (план):']
DELIV_ROW    = param_rows['из них: доставка (план):']
CAP_SQM_ROW  = param_rows['Макс. мощность цеха (кв.м/мес):']

TRUCKS_CELL  = f'$D${TRUCKS_ROW}'
ADDR_CELL    = f'$D${ADDR_ROW}'
CAP_SQM_CELL = f'$D${CAP_SQM_ROW}'

# Расчётная максимальная мощность
ws.merge_cells(f'A{r}:C{r}')
sc(ws,r,1, val='Макс. адресов/день (расчёт):', font=F(size=9,bold=True), fill=GRAY_FILL, align=aL(1))
sc(ws,r,4, fml=f'={TRUCKS_CELL}*{ADDR_CELL}',
   font=F(bold=True,size=11,color='991B1B'), fill=GREEN_FILL,
   align=aC(), nfmt='#,##0')
sc(ws,r,5, val='адр/день = машины × адресов', font=F(size=9,color='64748B'), fill=GRAY_FILL, align=aL())
ws.row_dimensions[r].height = 18
MAX_ADDR_ROW = r
MAX_ADDR_CELL = f'$D${MAX_ADDR_ROW}'
r+=2

# ── Выбор месяца ─────────────────────────────────────────────────────────────
ws.merge_cells(f'A{r}:B{r}')
sc(ws,r,1, val='Выберите месяц:', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,r,3, val='Апрель', font=F(bold=True,size=11,color='1D4ED8'),
   fill=YELLOW_FILL, align=aC(),
   brd=S(l='medium',r='medium',t='medium',b='medium',lc='3B82F6',rc='3B82F6',tc='3B82F6',bc='3B82F6'))
dv_m2 = DataValidation(type='list',
    formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"',
    allow_blank=False)
ws.add_data_validation(dv_m2)
MONTH_SEL_CELL = f'C{r}'
dv_m2.sqref = MONTH_SEL_CELL
ws.row_dimensions[r].height = 22
r+=1

# Рабочих дней в месяце (ввод вручную)
ws.merge_cells(f'A{r}:B{r}')
sc(ws,r,1, val='Рабочих дней в месяце:', font=F(bold=True,size=10), fill=GRAY_FILL, align=aL(1))
sc(ws,r,3, val=26, font=F(bold=True,size=11,color='065F46'),
   fill=YELLOW_FILL, align=aC(), nfmt='#,##0',
   brd=S(l='medium',r='medium',t='medium',b='medium',lc='3B82F6',rc='3B82F6',tc='3B82F6',bc='3B82F6'))
WORK_DAYS_CELL = f'$C${r}'
ws.row_dimensions[r].height = 20
r+=2

# Разделитель
ws.merge_cells(f'A{r}:H{r}')
sc(ws,r,1, fill=PatternFill('solid',fgColor='1E293B'))
ws.row_dimensions[r].height = 5
r+=1

# ── Заголовок таблицы ────────────────────────────────────────────────────────
ws.merge_cells(f'A{r}:H{r}')
sc(ws,r,1, val='📋 ДНЕВНОЙ КПД — Вводите фактические данные',
   font=F(bold=True,size=10,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='1D4ED8'), align=aL(1))
ws.row_dimensions[r].height = 18
r+=1

HDR_KPD = ['День','Факт адресов','Заборов\n(факт)','Доставок\n(факт)',
           'КПД логист. %','Кв.м факт','КПД цеха %','КПД оборота %']
for ci, h in enumerate(HDR_KPD, start=1):
    sc(ws,r,ci, val=h, font=F(bold=True,size=9,color='FFFFFF'),
       fill=HEADER_FILL, align=aC(True), brd=brd_h)
ws.row_dimensions[r].height = 30
ws.freeze_panes = f'A{r+1}'
r+=1

kpd_start = r

# Выручка план на день: связываем с "План продаж" если хотим, или ввод
# Для простоты: дневной КПД оборота = кол-во заказов факт / план_заказов_день
# Используем 31 день
MONTH_NUM2 = f'MATCH({MONTH_SEL_CELL},{{"Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"}},0)'

for day_n in range(1, 32):
    cur_r = r
    row_fill = WHITE_FILL if day_n%2==1 else GRAY_FILL

    # День: скрываем если > кол-ва дней в месяце
    date_fml = f'=IFERROR(IF(DAY(DATE(2026,{MONTH_NUM2},{day_n}))={day_n},{day_n},""),"")'
    sc(ws,cur_r,1, fml=date_fml,
       font=F(size=9,bold=True), fill=row_fill, align=aC(), nfmt='#,##0')

    # Факт адресов (ввод — желтый)
    sc(ws,cur_r,2, val=None,
       font=F(size=9,color='5F4C0A'), fill=YELLOW_FILL, align=aR(), nfmt='#,##0',
       brd=S(l='medium',r='medium',lc='CA8A04',rc='CA8A04'))

    # Заборов факт (ввод)
    sc(ws,cur_r,3, val=None,
       font=F(size=9,color='5F4C0A'), fill=YELLOW_FILL, align=aR(), nfmt='#,##0',
       brd=S(l='medium',r='medium',lc='CA8A04',rc='CA8A04'))

    # Доставок факт (ввод)
    sc(ws,cur_r,4, val=None,
       font=F(size=9,color='5F4C0A'), fill=YELLOW_FILL, align=aR(), nfmt='#,##0',
       brd=S(l='medium',r='medium',lc='CA8A04',rc='CA8A04'))

    # КПД логистики = факт_адресов / (машин × адр_план) × 100%
    kpd_log_fml = f'=IFERROR(IF(B{cur_r}="","",B{cur_r}/{MAX_ADDR_CELL}),"")'
    # Цвет: зелёный если >80%, оранжевый 60-80%, красный <60%
    sc(ws,cur_r,5, fml=kpd_log_fml,
       font=F(size=9,bold=True,color='065F46'), fill=GREEN_FILL, align=aR(), nfmt='0.0%')

    # Кв.м факт (ввод)
    sc(ws,cur_r,6, val=None,
       font=F(size=9,color='7C3AED'), fill=PatternFill('solid',fgColor='EDE9FE'), align=aR(), nfmt='#,##0.0',
       brd=S(l='medium',r='medium',lc='7C3AED',rc='7C3AED'))

    # КПД цеха = факт_кв.м / (макс_мощность / раб_дней)
    kpd_cex_fml = (f'=IFERROR(IF(F{cur_r}="","",F{cur_r}/({CAP_SQM_CELL}/{WORK_DAYS_CELL})),"")')
    sc(ws,cur_r,7, fml=kpd_cex_fml,
       font=F(size=9,bold=True,color='7C3AED'), fill=PURPLE_FILL, align=aR(), nfmt='0.0%')

    # КПД оборота = факт_адресов / план_адресов_день
    # план = факт × (дни в мес) / заказы_месяц — упрощённо: план адресов = MAX_ADDR_CELL
    kpd_oborot_fml = f'=IFERROR(IF(B{cur_r}="","",B{cur_r}/{MAX_ADDR_CELL}),"")'
    sc(ws,cur_r,8, fml=kpd_oborot_fml,
       font=F(size=9,bold=True,color='1D4ED8'), fill=BLUE_FILL, align=aR(), nfmt='0.0%')

    ws.row_dimensions[cur_r].height = 16
    r+=1

kpd_end = r-1

# Итоги
sc(ws,r,1, val='ИТОГО', font=F(bold=True,size=10), fill=TOTAL_FILL,
   align=aL(), brd=brd_t)
for ci, col in enumerate(['B','C','D'], start=2):
    sc(ws,r,ci,
       fml=f'=SUMIF({col}{kpd_start}:{col}{kpd_end},">"&0)',
       font=F(bold=True,size=10), fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0')
# Средние КПД
sc(ws,r,5,
   fml=f'=IFERROR(B{r}/({MAX_ADDR_CELL}*COUNTIF(B{kpd_start}:B{kpd_end},">"&0)),"")',
   font=F(bold=True,size=10,color='065F46'), fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='0.0%')
sc(ws,r,6,
   fml=f'=SUMIF(F{kpd_start}:F{kpd_end},">"&0)',
   font=F(bold=True,size=10), fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='#,##0.0')
sc(ws,r,7,
   fml=f'=IFERROR(F{r}/{CAP_SQM_CELL},"")',
   font=F(bold=True,size=10,color='7C3AED'), fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='0.0%')
sc(ws,r,8,
   fml=f'=IFERROR(B{r}/({MAX_ADDR_CELL}*COUNTIF(B{kpd_start}:B{kpd_end},">"&0)),"")',
   font=F(bold=True,size=10,color='1D4ED8'), fill=TOTAL_FILL, align=aR(), brd=brd_t, nfmt='0.0%')
ws.row_dimensions[r].height = 20
r+=2

# Легенда КПД
ws.merge_cells(f'A{r}:H{r}')
sc(ws,r,1, val='ℹ  КПД Логистики = Факт адресов / (Машин × Адр/маш). '
               'КПД Цеха = Факт кв.м / (Мощность цеха / Раб.дней). '
               'КПД Оборота = то же что логистика (т.к. адрес = заказ). '
               '100% = полная загрузка. Цель: >80%.',
   font=F(size=9,color='64748B',italic=True),
   fill=GRAY_FILL, align=aL())
ws.row_dimensions[r].height = 28
r+=1

ws.merge_cells(f'A{r}:H{r}')
sc(ws,r,1, val='Жёлтые = ввод ежедневных данных. '
               'Зелёные = КПД логистики. Синие = КПД оборота. Фиолетовые = КПД цеха.',
   font=F(size=9,color='64748B',italic=True),
   fill=GRAY_FILL, align=aL())
ws.row_dimensions[r].height = 16

print(f'✓ Лист "КПД Логистики" создан, строк: {r}')

# ─── Сохраняем ───────────────────────────────────────────────────────────────
wb.save(SRC_PATH)
print(f'\n✓ Файл сохранён: {SRC_PATH}')
print('  Листы добавлены: Воронка, План продаж, КПД Логистики')
