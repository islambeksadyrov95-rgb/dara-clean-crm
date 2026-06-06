import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Константы ────────────────────────────────────────────────────────────────
SRC_PATH   = 'D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет.xlsx'
SRC_SHEET  = 'ДДС 2025'          # имя исходного листа
NEW_SHEET  = 'ДДС 2025 V2'

MONTHS_RU  = ['Январь','Февраль','Март','Апрель','Май','Июнь',
               'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

# Колонки значений в исходном листе (1-based Excel cols)
# A=1 (name), B=2 Jan, D=4 Feb, F=6 Mar, H=8 Apr, J=10 May, L=12 Jun,
# N=14 Jul, P=16 Aug, R=18 Sep, T=20 Oct, V=22 Nov, X=24 Dec, Z=26 Total
SRC_MONTH_EXCEL_COLS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24]

# Колонки в новом листе V2
# A=1 label, B=2 Jan ... M=13 Dec, N=14 ИТОГО
V2_DATA_COLS = list(range(2, 14))   # B..M = cols 2..13
V2_TOTAL_COL = 14                   # N = ИТОГО

BLOCKS_ORDER = ['ПРОИЗВОДСТВО', 'ЛОГИСТИКА', 'МАРКЕТИНГ', 'ПРОДАЖИ', 'НАЛОГИ', 'ОПЕРАЦИОННЫЕ']
BLOCK_COLORS = {
    'ПРОИЗВОДСТВО': '8B5CF6',
    'ЛОГИСТИКА':    'F59E0B',
    'МАРКЕТИНГ':    '3B82F6',
    'ПРОДАЖИ':      '10B981',
    'НАЛОГИ':       '6B7280',
    'ОПЕРАЦИОННЫЕ': 'EC4899',
}

# Маппинг статья → блок
BLOCK_MAP = {
    'ФОТ оклад ЦЕХ':'ПРОИЗВОДСТВО','Аванс':'ПРОИЗВОДСТВО',
    'ФОТ больничный':'ПРОИЗВОДСТВО','ФОТ аутсорс':'ПРОИЗВОДСТВО','Бонусы':'ПРОИЗВОДСТВО',
    'Шампунь для ковров':'ПРОИЗВОДСТВО','Химия для чистки':'ПРОИЗВОДСТВО',
    'Химия для цеха':'ПРОИЗВОДСТВО','Пакеты':'ПРОИЗВОДСТВО','Бирки':'ПРОИЗВОДСТВО',
    'Скобы':'ПРОИЗВОДСТВО','Накладные':'ПРОИЗВОДСТВО','Скотч':'ПРОИЗВОДСТВО',
    'Хозбыт':'ПРОИЗВОДСТВО','Инструменты':'ПРОИЗВОДСТВО','Бытовые':'ПРОИЗВОДСТВО',
    'Содержание цеха':'ПРОИЗВОДСТВО','Аренда Агбиса':'ПРОИЗВОДСТВО',
    'Свет':'ПРОИЗВОДСТВО','Газ':'ПРОИЗВОДСТВО','Вода':'ПРОИЗВОДСТВО',
    'Продукты цех':'ПРОИЗВОДСТВО','Приобретение для кухни цеха':'ПРОИЗВОДСТВО',
    'Вывоз мусора':'ПРОИЗВОДСТВО','Септик':'ПРОИЗВОДСТВО',
    'Документы рабочим':'ПРОИЗВОДСТВО','Ремонт ковров':'ПРОИЗВОДСТВО',
    'Оплата услуг штор и тюлей':'ПРОИЗВОДСТВО',
    'Оборудование (обновление)':'ПРОИЗВОДСТВО',
    'Оборудование (тех.обслуживание)':'ПРОИЗВОДСТВО',
    'Приобретение для цеха':'ПРОИЗВОДСТВО',
    'ФОТ оклад Логистика':'ЛОГИСТИКА','Обед водителям':'ЛОГИСТИКА',
    'Приобретение для курьеров':'ЛОГИСТИКА','Такси/доставка':'ЛОГИСТИКА',
    'Абонент трекер':'ЛОГИСТИКА','ГСМ':'ЛОГИСТИКА','Мойка':'ЛОГИСТИКА',
    'Запчасти':'ЛОГИСТИКА','Замена шин/Вулканизация':'ЛОГИСТИКА',
    'Ремонт ходовой части':'ЛОГИСТИКА','Ремонт двигителя и коробки':'ЛОГИСТИКА',
    'Ремонт':'ЛОГИСТИКА','Замена масла':'ЛОГИСТИКА','Замена колодок':'ЛОГИСТИКА',
    'Штрафы':'ЛОГИСТИКА','Техосмотр':'ЛОГИСТИКА','Страховка':'ЛОГИСТИКА',
    'Прочее расходы':'ЛОГИСТИКА',
    'Покупка рекламы Google':'МАРКЕТИНГ','Покупка рекламы Yandex':'МАРКЕТИНГ',
    'Покупка рекламы Instagram':'МАРКЕТИНГ','Покупка рекламы TikTok':'МАРКЕТИНГ',
    'Покупка рекламы 2Гис':'МАРКЕТИНГ','Наружняя реклама':'МАРКЕТИНГ',
    'СММ':'МАРКЕТИНГ','Констектолог':'МАРКЕТИНГ','Таргетолог':'МАРКЕТИНГ',
    'Подписка сайта':'МАРКЕТИНГ','Оплата модели':'МАРКЕТИНГ','Видеограф':'МАРКЕТИНГ',
    'Бюджет на съемки':'МАРКЕТИНГ','Купоны и баннеры':'МАРКЕТИНГ',
    'Бартер/Блогер':'МАРКЕТИНГ','Обзвон':'МАРКЕТИНГ','Реклама мебели':'МАРКЕТИНГ',
    'ФОТ оклад Отел продаж':'ПРОДАЖИ','Телефония Кар Тел':'ПРОДАЖИ',
    'Wazzup подписка':'ПРОДАЖИ','Канцелярия':'ПРОДАЖИ',
    'Приобретение для отдела продаж':'ПРОДАЖИ',
    'Налоги_1':'НАЛОГИ','Тариф':'НАЛОГИ',
    'Интернет':'ОПЕРАЦИОННЫЕ','Подписка Битрикс':'ОПЕРАЦИОННЫЕ',
    'Бухгалтер':'ОПЕРАЦИОННЫЕ','Содержание офиса':'ОПЕРАЦИОННЫЕ',
    'Услиги юриста':'ОПЕРАЦИОННЫЕ','Услуги типографии':'ОПЕРАЦИОННЫЕ',
    'Тимбилдинг':'ОПЕРАЦИОННЫЕ','Покупка консультации':'ОПЕРАЦИОННЫЕ',
    'Обьяление для сотрудников':'ОПЕРАЦИОННЫЕ',
    'Единоразовый расход (приобретение)':'ОПЕРАЦИОННЫЕ',
    'Прочие расходы':'ОПЕРАЦИОННЫЕ','Покупка ковра':'ОПЕРАЦИОННЫЕ',
    'Услуга аренды машины':'ОПЕРАЦИОННЫЕ',
    'Погашение займа':'__LOAN','Погашение кредита':'__LOAN',
    'Kaspi Pay':'__WD','Наличными':'__WD',
}

SKIP = {
    'Поставщик','Поставщик_1','Хоз.рас материалы_1','Маркетинг_1',
    'Транспортные расходы','Транспортные расходы_1','ФОТ_1',
    'Финансовые операции','Финансовые операции_1',
    'Доходы','Расходы','Итого','Чистый доход','EBITDA',
    'Вывод средств','Кассовый разрыв','Кол-во заказов','Кол-во кв.м',
    'Баланс компании','Остаточный капитал','Самовывоз','Услуги',
    '2025','Пополнение','Получение банковского кредита','Получение займа',
    'ФОТ Собственник',
}

# ── 1. Читаем исходный лист: строим карту name → excel_row ───────────────────
# Читаем значения отдельно (data_only нужен только для чтения)
_wb_data = openpyxl.load_workbook(SRC_PATH, data_only=True)
ws_src = _wb_data[SRC_SHEET]

# row_map: {stripped_name: excel_row_number}
row_map = {}
income_row = None   # строка "Итого" для доходов
wd_rows    = []     # строки вывода средств (Kaspi Pay, Наличными)
loan_rows_src = []  # строки погашения кредита

# Сначала найдём строку Итого доходов (row 15 по факту)
# Ищем все строки
all_src = list(ws_src.iter_rows(values_only=True))
income_itogo_row = None
in_income_section = False

for i, row in enumerate(all_src, start=1):
    name = row[0]
    if name is None:
        continue
    name_s = str(name).strip()

    if name_s == 'Доходы':
        in_income_section = True
    if name_s == 'Расходы':
        in_income_section = False

    if in_income_section and name_s == 'Итого':
        income_itogo_row = i
        continue

    if name_s in SKIP:
        continue

    row_map[name_s] = i

    # Проверяем значения
    vals = [row[c-1] if (c-1) < len(row) else None for c in SRC_MONTH_EXCEL_COLS]
    vals = [v if isinstance(v,(int,float)) else 0 for v in vals]

    block = BLOCK_MAP.get(name_s)
    # Вывод средств — только строки НИЖЕ строки 100 (после основного блока расходов)
    # Строки 7,8 — это Kaspi Pay/Наличными из секции Пополнение (доходы) = нули
    if block == '__WD' and i > 100:
        wd_rows.append(i)
    elif block == '__LOAN':
        loan_rows_src.append(i)

print(f'Итого доходов: строка {income_itogo_row}')
print(f'Строки кредита: {loan_rows_src}')
print(f'Строки вывода: {wd_rows}')

# ── 2. Формируем список строк по блокам (name, src_row) ──────────────────────
block_items = {b: [] for b in BLOCKS_ORDER}
for name_s, src_row in row_map.items():
    block = BLOCK_MAP.get(name_s)
    if block and not block.startswith('__') and block in block_items:
        # Получаем сумму для сортировки по убыванию
        src_row_data = all_src[src_row - 1]
        total = sum(
            src_row_data[c-1] if (c-1)<len(src_row_data) and isinstance(src_row_data[c-1],(int,float)) else 0
            for c in SRC_MONTH_EXCEL_COLS
        )
        block_items[block].append((name_s, src_row, total))

# Сортируем внутри блока по убыванию суммы
for b in BLOCKS_ORDER:
    block_items[b].sort(key=lambda x: -x[2])

# ── 3. Стили ─────────────────────────────────────────────────────────────────
def mk_border(l='thin',r='thin',t='thin',b='thin', lc='CBD5E1',rc='CBD5E1',tc='CBD5E1',bc='CBD5E1'):
    def s(st, c): return Side(style=st, color=c) if st else Side(style=None)
    return Border(left=s(l,lc), right=s(r,rc), top=s(t,tc), bottom=s(b,bc))

HEADER_FILL = PatternFill('solid', fgColor='1E293B')
WHITE_FILL  = PatternFill('solid', fgColor='FFFFFF')
GRAY_FILL   = PatternFill('solid', fgColor='F8FAFC')
TOTAL_FILL  = PatternFill('solid', fgColor='F1F5F9')
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'

def set_cell(ws, row, col, value=None, formula=None, font=None, fill=None,
             alignment=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col)
    if formula:
        c.value = formula
    elif value is not None:
        c.value = value
    if font:      c.font      = font
    if fill:      c.fill      = fill
    if alignment: c.alignment = alignment
    if border:    c.border    = border
    if num_fmt:   c.number_format = num_fmt
    return c

def hfont(size=10, bold=True, color='FFFFFF'):
    return Font(bold=bold, size=size, color=color, name='Calibri')
def dfont(size=9, bold=False, color='334155'):
    return Font(bold=bold, size=size, color=color, name='Calibri')
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def right():  return Alignment(horizontal='right',  vertical='center')
def left(indent=0): return Alignment(horizontal='left', vertical='center', indent=indent)

# ── 4. Формульный хелпер ─────────────────────────────────────────────────────
def src_ref(src_excel_row, src_excel_col):
    """Ссылка на ячейку исходного листа."""
    col_letter = get_column_letter(src_excel_col)
    return f"='{SRC_SHEET}'!{col_letter}{src_excel_row}"

def sum_range(col_letter, row_start, row_end):
    """=SUM(Brow_start:Brow_end)"""
    return f"=SUM({col_letter}{row_start}:{col_letter}{row_end})"

def row_sum(row_n):
    """=SUM(B{r}:M{r}) — сумма 12 месяцев по строке"""
    return f"=SUM(B{row_n}:M{row_n})"

# ── 5. Создаём лист ──────────────────────────────────────────────────────────
# Закрываем data_only воркбук, открываем чистый для сохранения формул
_wb_data.close()
wb = openpyxl.load_workbook(SRC_PATH)

if NEW_SHEET in wb.sheetnames:
    del wb[NEW_SHEET]
ws = wb.create_sheet(NEW_SHEET)

ws.column_dimensions['A'].width = 36
for i in range(2, 14):
    ws.column_dimensions[get_column_letter(i)].width = 12
ws.column_dimensions[get_column_letter(14)].width = 14

# ── 6. Строка 1: заголовок ────────────────────────────────────────────────────
r = 1
ws.merge_cells(f'A{r}:N{r}')
set_cell(ws, r, 1,
    value='ДДС 2025 — Группировка по блокам дашборда',
    font=hfont(13), fill=HEADER_FILL, alignment=center())
ws.row_dimensions[r].height = 28

# ── 7. Строка 2: названия месяцев ────────────────────────────────────────────
r = 2
brd_hdr = mk_border(t='medium',b='medium',tc='475569',bc='475569')
set_cell(ws, r, 1, value='Статья расходов',
    font=hfont(10), fill=HEADER_FILL, alignment=center(), border=brd_hdr)
for j, m in enumerate(MONTHS_RU, start=2):
    set_cell(ws, r, j, value=m,
        font=hfont(9), fill=HEADER_FILL, alignment=center(), border=brd_hdr)
set_cell(ws, r, 14, value='ИТОГО',
    font=hfont(10), fill=PatternFill('solid',fgColor='0F172A'),
    alignment=center(), border=brd_hdr)
ws.row_dimensions[r].height = 22

ws.freeze_panes = 'B3'

# ── хелпер: записать строку со ссылками на ДДС 2025 ─────────────────────────
def write_ref_row(ws, r, label, src_row, label_fill, val_fill, bold=False, indent=2):
    brd_l  = mk_border(l='medium', lc='94A3B8')
    brd_v  = mk_border()
    brd_vt = mk_border(r='medium', rc='94A3B8')
    font   = dfont(bold=bold)

    set_cell(ws, r, 1, value=label,
        font=font, fill=label_fill, alignment=left(indent), border=brd_l)
    for j, src_col in enumerate(SRC_MONTH_EXCEL_COLS, start=2):
        set_cell(ws, r, j, formula=src_ref(src_row, src_col),
            font=font, fill=val_fill, alignment=right(), border=brd_v, num_fmt=NUM_FMT)
    set_cell(ws, r, 14, formula=row_sum(r),
        font=dfont(bold=True), fill=val_fill, alignment=right(), border=brd_vt, num_fmt=NUM_FMT)
    ws.row_dimensions[r].height = 16

# хелпер: SUM-строка по диапазону строк (итого блока / COGS / etc.)
def write_sum_row(ws, r, label, data_row_start, data_row_end,
                  label_fill, val_fill, font_color='1E293B',
                  bold=True, size=9, top_border='medium', bot_border='medium'):
    brd_l = mk_border(l='medium', t=top_border, b=bot_border,
                      lc='94A3B8', tc='475569', bc='475569')
    brd_v = mk_border(t=top_border, b=bot_border, tc='475569', bc='475569')
    brd_vt= mk_border(r='medium', t=top_border, b=bot_border,
                      rc='94A3B8', tc='475569', bc='475569')
    font  = Font(bold=bold, size=size, color=font_color, name='Calibri')

    set_cell(ws, r, 1, value=label,
        font=font, fill=label_fill, alignment=left(1), border=brd_l)
    for j in range(2, 14):
        col_l = get_column_letter(j)
        set_cell(ws, r, j, formula=f'=SUM({col_l}{data_row_start}:{col_l}{data_row_end})',
            font=font, fill=val_fill, alignment=right(), border=brd_v, num_fmt=NUM_FMT)
    col_l = get_column_letter(14)
    set_cell(ws, r, 14, formula=f'=SUM({col_l}{data_row_start}:{col_l}{data_row_end})',
        font=Font(bold=True, size=size, color=font_color, name='Calibri'),
        fill=val_fill, alignment=right(), border=brd_vt, num_fmt=NUM_FMT)
    ws.row_dimensions[r].height = 20

# хелпер: формульная строка (произвольная формула на каждую колонку)
def write_formula_row(ws, r, label, col_formulas,
                      label_fill, val_fill, font_color='1E293B',
                      bold=True, size=10, num_fmt=NUM_FMT):
    """col_formulas: list of 13 formula strings [jan..dec, total]"""
    brd_l = mk_border(l='medium', t='medium', b='medium', lc='94A3B8', tc='475569', bc='475569')
    brd_v = mk_border(t='medium', b='medium', tc='475569', bc='475569')
    brd_vt= mk_border(r='medium', t='medium', b='medium', rc='94A3B8', tc='475569', bc='475569')
    font  = Font(bold=bold, size=size, color=font_color, name='Calibri')

    set_cell(ws, r, 1, value=label,
        font=font, fill=label_fill, alignment=left(1), border=brd_l)
    for j, fml in enumerate(col_formulas[:13], start=2):
        brd = brd_vt if j == 14 else brd_v
        set_cell(ws, r, j, formula=fml,
            font=font if j < 14 else Font(bold=True, size=size, color=font_color, name='Calibri'),
            fill=val_fill, alignment=right(), border=brd, num_fmt=num_fmt)
    ws.row_dimensions[r].height = 20

def write_spacer(ws, r):
    for col in range(1, 15):
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill('solid', fgColor='F1F5F9')
        c.border = Border()
    ws.row_dimensions[r].height = 5

# ── 8. ДОХОДЫ ────────────────────────────────────────────────────────────────
r += 1
write_spacer(ws, r); r += 1

# Доходы = ссылка на строку "Итого" из секции доходов ДДС 2025
rev_row = r
write_ref_row(ws, r,
    label='▶  ДОХОДЫ (Услуги)',
    src_row=income_itogo_row,
    label_fill=PatternFill('solid', fgColor='DCFCE7'),
    val_fill=PatternFill('solid', fgColor='F0FDF4'),
    bold=True, indent=1)
ws.cell(row=r, column=1).font = Font(bold=True, size=11, color='065F46', name='Calibri')
ws.row_dimensions[r].height = 20

r += 1
write_spacer(ws, r); r += 1

# ── 9. БЛОКИ РАСХОДОВ ─────────────────────────────────────────────────────────
block_total_rows = {}   # block → excel row of its total line

for block in BLOCKS_ORDER:
    color    = BLOCK_COLORS[block]
    b_fill   = PatternFill('solid', fgColor=color)
    items    = block_items[block]

    # -- Заголовок блока
    brd_hb = mk_border(l='medium', t='medium', lc='94A3B8', tc='475569')
    set_cell(ws, r, 1, value=f'▶  {block}',
        font=Font(bold=True, size=10, color='FFFFFF', name='Calibri'),
        fill=b_fill, alignment=left(1), border=brd_hb)
    for j in range(2, 15):
        brd = mk_border(t='medium', r='medium' if j==14 else 'thin',
                        tc='475569', rc='94A3B8' if j==14 else 'CBD5E1')
        ws.cell(row=r, column=j).fill   = b_fill
        ws.cell(row=r, column=j).border = brd
    ws.row_dimensions[r].height = 20
    r += 1

    # -- Детальные строки
    detail_start = r
    for idx, (name, src_row, _) in enumerate(items):
        row_fill = WHITE_FILL if (idx % 2 == 0) else GRAY_FILL
        write_ref_row(ws, r, f'     {name}', src_row,
                      label_fill=row_fill, val_fill=row_fill)
        r += 1
    detail_end = r - 1

    # -- Итого блока
    block_total_rows[block] = r
    write_sum_row(ws, r,
        label=f'   Итого {block.capitalize()}',
        data_row_start=detail_start, data_row_end=detail_end,
        label_fill=TOTAL_FILL, val_fill=TOTAL_FILL,
        font_color=color, bold=True, size=9)
    r += 1
    write_spacer(ws, r); r += 1

# ── 10. ИТОГО COGS (без вывода) ───────────────────────────────────────────────
cogs_row = r
bt_rows = list(block_total_rows.values())
cogs_formulas = []
for col_i in range(2, 15):
    col_l = get_column_letter(col_i)
    refs  = '+'.join(f'{col_l}{bt}' for bt in bt_rows)
    cogs_formulas.append(f'={refs}')
write_formula_row(ws, r,
    label='▶  ИТОГО COGS (без вывода)',
    col_formulas=cogs_formulas,
    label_fill=PatternFill('solid', fgColor='FEE2E2'),
    val_fill=PatternFill('solid', fgColor='FEF2F2'),
    font_color='991B1B', bold=True, size=11)
r += 1

# -- % Маржи (без вывода)
margin_row = r
margin_fmls = []
for col_i in range(2, 15):
    col_l = get_column_letter(col_i)
    margin_fmls.append(f'=IF({col_l}{rev_row}=0,0,({col_l}{rev_row}-{col_l}{cogs_row})/{col_l}{rev_row})')
write_formula_row(ws, r,
    label='     % Валовой маржи (без вывода)',
    col_formulas=margin_fmls,
    label_fill=PatternFill('solid', fgColor='F0FDF4'),
    val_fill=PatternFill('solid', fgColor='F0FDF4'),
    font_color='065F46', bold=False, size=9, num_fmt=PCT_FMT)
ws.row_dimensions[r].height = 16
r += 1

# -- Валовая прибыль (без вывода)
gp_row = r
gp_fmls = [f'={get_column_letter(j)}{rev_row}-{get_column_letter(j)}{cogs_row}'
           for j in range(2, 15)]
write_formula_row(ws, r,
    label='▶  ВАЛОВАЯ ПРИБЫЛЬ (без вывода)',
    col_formulas=gp_fmls,
    label_fill=PatternFill('solid', fgColor='D1FAE5'),
    val_fill=PatternFill('solid', fgColor='DCFCE7'),
    font_color='065F46', bold=True, size=11)
r += 1
write_spacer(ws, r); r += 1

# ── 11. ВЫВОД СРЕДСТВ — секция с чекбоксом ────────────────────────────────────
# Разделитель
sep_fill = PatternFill('solid', fgColor='F1F5F9')
ws.merge_cells(f'A{r}:N{r}')
set_cell(ws, r, 1,
    value='— ВЫВОД СРЕДСТВ И ФИНАНСОВЫЕ ОПЕРАЦИИ (под чертой) —',
    font=Font(bold=True, size=9, color='7C3AED', italic=True, name='Calibri'),
    fill=sep_fill,
    alignment=Alignment(horizontal='right', vertical='center'))
ws.row_dimensions[r].height = 16
r += 1

# Чекбокс строка
CHECKBOX_CELL = f'B{r}'
checkbox_fill = PatternFill('solid', fgColor='EDE9FE')
ws.merge_cells(f'A{r}:A{r}')
set_cell(ws, r, 1,
    value='⚙  Включить вывод средств в COGS?',
    font=Font(bold=True, size=10, color='5B21B6', name='Calibri'),
    fill=checkbox_fill,
    alignment=Alignment(horizontal='left', vertical='center'))
# Dropdown в столбце B (Январь)
set_cell(ws, r, 2,
    value='Нет',
    font=Font(bold=True, size=11, color='1D4ED8', name='Calibri'),
    fill=PatternFill('solid', fgColor='DBEAFE'),
    alignment=Alignment(horizontal='center', vertical='center'))
# Подсказка рядом
ws.merge_cells(f'C{r}:N{r}')
set_cell(ws, r, 3,
    value='← "Да" = включить в COGS  |  "Нет" = ниже черты',
    font=Font(size=9, color='6B7280', italic=True, name='Calibri'),
    fill=checkbox_fill,
    alignment=Alignment(horizontal='left', vertical='center'))
dv = DataValidation(type='list', formula1='"Да,Нет"', allow_blank=False)
ws.add_data_validation(dv)
dv.sqref = CHECKBOX_CELL
ws.row_dimensions[r].height = 22
r += 1

# Кредит
loan_detail_start = r
for src_row_n in loan_rows_src:
    label = str(all_src[src_row_n-1][0]).strip()
    write_ref_row(ws, r, f'     {label}', src_row_n,
        label_fill=PatternFill('solid', fgColor='FEF9C3'),
        val_fill=PatternFill('solid', fgColor='FEF9C3'))
    r += 1

# Вывод средств (ABS — хранятся отрицательными)
wd_detail_rows = []
for src_row_n in wd_rows:
    label = str(all_src[src_row_n-1][0]).strip()
    brd_l = mk_border(l='medium', lc='94A3B8')
    brd_v = mk_border()
    brd_vt= mk_border(r='medium', rc='94A3B8')
    font  = dfont()
    wd_fill = PatternFill('solid', fgColor='FEF9C3')
    set_cell(ws, r, 1, value=f'     {label}',
        font=font, fill=wd_fill, alignment=left(2), border=brd_l)
    for j, src_col in enumerate(SRC_MONTH_EXCEL_COLS, start=2):
        col_letter = get_column_letter(src_col)
        formula = f"=ABS('{SRC_SHEET}'!{col_letter}{src_row_n})"
        set_cell(ws, r, j, formula=formula,
            font=font, fill=wd_fill, alignment=right(), border=brd_v, num_fmt=NUM_FMT)
    set_cell(ws, r, 14, formula=row_sum(r),
        font=dfont(bold=True), fill=wd_fill, alignment=right(), border=brd_vt, num_fmt=NUM_FMT)
    ws.row_dimensions[r].height = 16
    wd_detail_rows.append(r)
    r += 1

loan_detail_end = r - 1

# Итого вывод + кредит
wd_total_row = r
all_detail_rows = list(range(loan_detail_start, loan_detail_end + 1))
wd_total_fmls = []
for col_i in range(2, 15):
    col_l = get_column_letter(col_i)
    refs = '+'.join(f'{col_l}{dr}' for dr in all_detail_rows)
    wd_total_fmls.append(f'={refs}')
write_formula_row(ws, r,
    label='   Итого вывод + кредит',
    col_formulas=wd_total_fmls,
    label_fill=PatternFill('solid', fgColor='FEF3C7'),
    val_fill=PatternFill('solid', fgColor='FEF3C7'),
    font_color='92400E', bold=True, size=9)
ws.row_dimensions[r].height = 16
r += 1
write_spacer(ws, r); r += 1

# ── 12. ИТОГО COGS (с учётом галочки) ─────────────────────────────────────────
cogs_checked_row = r
cogs_checked_fmls = []
for col_i in range(2, 15):
    col_l = get_column_letter(col_i)
    base = '+'.join(f'{col_l}{bt}' for bt in bt_rows)
    wd   = f'{col_l}{wd_total_row}'
    cogs_checked_fmls.append(f'=IF({CHECKBOX_CELL}="Да",{base}+{wd},{base})')
write_formula_row(ws, r,
    label='▶  ИТОГО COGS (с учётом галочки)',
    col_formulas=cogs_checked_fmls,
    label_fill=PatternFill('solid', fgColor='FEE2E2'),
    val_fill=PatternFill('solid', fgColor='FEF2F2'),
    font_color='991B1B', bold=True, size=11)
r += 1

# -- % Валовой маржи (с учётом галочки)
margin_checked_fmls = []
for col_i in range(2, 15):
    col_l = get_column_letter(col_i)
    margin_checked_fmls.append(
        f'=IF({col_l}{rev_row}=0,0,({col_l}{rev_row}-{col_l}{cogs_checked_row})/{col_l}{rev_row})')
write_formula_row(ws, r,
    label='     % Валовой маржи (с учётом галочки)',
    col_formulas=margin_checked_fmls,
    label_fill=PatternFill('solid', fgColor='FEF2F2'),
    val_fill=PatternFill('solid', fgColor='FEF2F2'),
    font_color='991B1B', bold=False, size=9, num_fmt=PCT_FMT)
ws.row_dimensions[r].height = 16
r += 1

# ── 13. ЧИСТЫЙ ДОХОД ──────────────────────────────────────────────────────────
net_row = r
net_fmls = []
for col_i in range(2, 15):
    col_l = get_column_letter(col_i)
    deductions = '+'.join(f'{col_l}{dr}' for dr in all_detail_rows)
    net_fmls.append(f'={col_l}{gp_row}-({deductions})')
write_formula_row(ws, r,
    label='▶  ЧИСТЫЙ ДОХОД',
    col_formulas=net_fmls,
    label_fill=PatternFill('solid', fgColor='DBEAFE'),
    val_fill=PatternFill('solid', fgColor='EFF6FF'),
    font_color='1E40AF', bold=True, size=11)

# Примечание
r += 2
ws.merge_cells(f'A{r}:N{r}')
set_cell(ws, r, 1,
    value='ℹ  Жёлтые ячейки = факт из ДДС 2025. Галочка в B{cb}: "Да" → вывод входит в COGS; "Нет" → вывод отражён отдельно ниже черты.'.replace('{cb}', str(net_row - 7)),
    font=Font(size=9, color='64748B', italic=True, name='Calibri'),
    fill=PatternFill('solid', fgColor='F8FAFC'),
    alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
ws.row_dimensions[r].height = 28

# ── 13. Сохраняем ─────────────────────────────────────────────────────────────
wb.save(SRC_PATH)
print(f'✓ Лист "{NEW_SHEET}" создан с формулами в {SRC_PATH}')
print(f'  Блоки:')
for b in BLOCKS_ORDER:
    print(f'    {b}: {len(block_items[b])} статей → строка итого {block_total_rows[b]}')
print(f'  ИТОГО COGS строка: {cogs_row}')
print(f'  Валовая прибыль строка: {gp_row}')
print(f'  Чистый доход строка: {net_row}')
print(f'  Всего строк: {r}')
