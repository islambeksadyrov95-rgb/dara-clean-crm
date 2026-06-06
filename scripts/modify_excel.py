# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILE = "D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет.xlsx"

wb = load_workbook(FILE)
ws = wb["ДДС 2026 V2"]

plan_cols = [6, 9, 12, 15, 18, 21, 24, 27, 30]  # F, I, L, O, R, U, X, AA, AD

# ============================================================
# STEP 1: CFO params AH11/AI11 and AH12/AI12
# ============================================================
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
label_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws["AH11"] = "Ежемес. вывод (план)"
ws["AH11"].font = Font(name="Calibri", size=9, bold=False)
ws["AH11"].fill = label_fill
ws["AH11"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

ws["AI11"] = 0
ws["AI11"].fill = yellow_fill
ws["AI11"].font = Font(name="Calibri", size=10, bold=True)
ws["AI11"].alignment = Alignment(horizontal="center", vertical="center")
ws["AI11"].number_format = "#,##0"

ws["AH12"] = "Нач. баланс (конец 2025)"
ws["AH12"].font = Font(name="Calibri", size=9, bold=False)
ws["AH12"].fill = label_fill
ws["AH12"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

ws["AI12"] = 0
ws["AI12"].fill = yellow_fill
ws["AI12"].font = Font(name="Calibri", size=10, bold=True)
ws["AI12"].alignment = Alignment(horizontal="center", vertical="center")
ws["AI12"].number_format = "#,##0"

print("STEP 1: CFO params added AH11/AI11 and AH12/AI12")

# ============================================================
# STEP 2: Row 121 - Плановый вывод средств
# ============================================================
print(f"Row 121 check - A: {repr(ws['A121'].value)}, B: {repr(ws['B121'].value)}")

ws["A121"] = "Плановый вывод средств"
ws["A121"].font = Font(name="Calibri", size=10, italic=True)
ws["A121"].alignment = Alignment(indent=2)

light_red_fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")

for col_idx in plan_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}121"]
    cell.value = "=$AI$11"
    cell.number_format = "#,##0"
    cell.fill = light_red_fill

print("STEP 2: Row 121 added")

# ============================================================
# STEP 3: Update Row 122 SUM to include row 121
# ============================================================
# Fact: B,C,D,E,H(8),K(11),N(14),Q(17),T(20),W(23),Z(26),AC(29)
# Plan: F(6),I(9),L(12),O(15),R(18),U(21),X(24),AA(27),AD(30)
all_data_cols = [2, 3, 4, 5, 6, 8, 9, 11, 12, 14, 15, 17, 18, 20, 21, 23, 24, 26, 27, 29, 30]

for col_idx in all_data_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}122"]
    cell.value = f"=SUM({col}117:{col}121)"

print("STEP 3: Row 122 updated to SUM(xxx117:xxx121)")

# ============================================================
# STEP 4: Fill missing plan formulas rows 111,112,113,124,125,126
# ============================================================
for col_idx in plan_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}111"]
    if cell.value is None or cell.value == 0:
        cell.value = f"={col}39+{col}60+{col}80+{col}88+{col}93+{col}109"
        print(f"  Added {col}111")

for col_idx in plan_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}112"]
    if cell.value is None or cell.value == 0:
        cell.value = f"=IF({col}5=0,0,({col}5-{col}111)/{col}5)"
        cell.number_format = "0.0%"
        print(f"  Added {col}112")

for col_idx in plan_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}113"]
    if cell.value is None or cell.value == 0:
        cell.value = f"={col}5-{col}111"
        print(f"  Added {col}113")

for col_idx in plan_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}124"]
    if cell.value is None or cell.value == 0:
        cell.value = (
            f'=IF($E$116="Да",'
            f'{col}39+{col}60+{col}80+{col}88+{col}93+{col}109+{col}122,'
            f'{col}39+{col}60+{col}80+{col}88+{col}93+{col}109)'
        )
        print(f"  Added {col}124")

for col_idx in plan_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}125"]
    if cell.value is None or cell.value == 0:
        cell.value = f"=IF({col}5=0,0,({col}5-{col}124)/{col}5)"
        cell.number_format = "0.0%"
        print(f"  Added {col}125")

for col_idx in plan_cols:
    col = get_column_letter(col_idx)
    cell = ws[f"{col}126"]
    if cell.value is None or cell.value == 0:
        cell.value = f"={col}5-{col}124"
        print(f"  Added {col}126")

print("STEP 4: Missing plan formulas filled")

# ============================================================
# STEP 5: Rows 128-130 - Header, Баланс, Кассовый разрыв
# ============================================================
ws["A128"] = "-- БАЛАНС КОМПАНИИ И КАССОВЫЙ РАЗРЫВ --"
ws["A128"].font = Font(name="Calibri", size=9, bold=True, color="444444", italic=True)
ws["A128"].alignment = Alignment(horizontal="center")
try:
    ws.merge_cells("A128:AF128")
except Exception as e:
    print(f"  merge warning: {e}")

ws["A129"] = "Баланс компании"
ws["A129"].font = Font(name="Calibri", size=10, bold=True)

ws["A130"] = "Кассовый разрыв"
ws["A130"].font = Font(name="Calibri", size=10, bold=True, color="CC0000")

blue_fill      = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
blue_plan_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
red_fill       = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")

fact_cols_sequence = [
    ("B", None),
    ("C", "B"),
    ("D", "C"),
    ("E", "D"),
    ("H", "E"),
    ("K", "H"),
    ("N", "K"),
    ("Q", "N"),
    ("T", "Q"),
    ("W", "T"),
    ("Z", "W"),
    ("AC", "Z"),
]

plan_cols_sequence = [
    ("F", "D"),
    ("I", "F"),
    ("L", "I"),
    ("O", "L"),
    ("R", "O"),
    ("U", "R"),
    ("X", "U"),
    ("AA", "X"),
    ("AD", "AA"),
]

for col, prev in fact_cols_sequence:
    cell = ws[f"{col}129"]
    cell.value = f"=$AI$12+{col}126" if prev is None else f"={prev}129+{col}126"
    cell.number_format = "#,##0"
    cell.fill = blue_fill
    cell.font = Font(name="Calibri", size=10, bold=True)

for col, prev in plan_cols_sequence:
    cell = ws[f"{col}129"]
    cell.value = f"=$AI$12+{col}126" if prev is None else f"={prev}129+{col}126"
    cell.number_format = "#,##0"
    cell.fill = blue_plan_fill
    cell.font = Font(name="Calibri", size=10, bold=True)

all_monthly_cols = [col for col, _ in fact_cols_sequence] + [col for col, _ in plan_cols_sequence]
for col in all_monthly_cols:
    cell = ws[f"{col}130"]
    cell.value = f"=MAX(0,-{col}129)"
    cell.number_format = "#,##0"
    cell.fill = red_fill
    cell.font = Font(name="Calibri", size=10, bold=True, color="CC0000")

print("STEP 5: Balance (row 129) and cash gap (row 130) rows added")

# ============================================================
# Save
# ============================================================
wb.save(FILE)
print(f"Saved: {FILE}")
print("DONE!")
