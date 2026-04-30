import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC_PATH  = 'D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет.xlsx'
SHEET_26  = 'ДДС 2026'
SHEET_25  = 'ДДС 2025'
SHEET_KVM = 'Кол-во Кв.м_Заказов_Ср.чек'

# Читаем значения отдельно (data_only=True нужен только для чтения)
_wb_data = openpyxl.load_workbook(SRC_PATH, data_only=True)
rows_26  = list(_wb_data[SHEET_26].iter_rows(values_only=True))
rows_25  = list(_wb_data[SHEET_25].iter_rows(values_only=True))
rows_kvm = list(_wb_data[SHEET_KVM].iter_rows(values_only=True))
_wb_data.close()

# Основной воркбук — БЕЗ data_only, чтобы сохранить формулы в существующих листах
wb = openpyxl.load_workbook(SRC_PATH)

# ─── Структура колонок ДДС 2026 V2 ─────────────────────────────────────────
# A=1: label
# B=2  Jan Факт
# C=3  Feb Факт
# D=4  Mar Факт
# Apr→Dec: 3 col per month (Факт | План | %)
#   Apr: E(5) F(6) G(7)
#   May: H(8) I(9) J(10)
#   Jun: K(11) L(12) M(13)
#   Jul: N(14) O(15) P(16)
#   Aug: Q(17) R(18) S(19)
#   Sep: T(20) U(21) V(22)
#   Oct: W(23) X(24) Y(25)
#   Nov: Z(26) AA(27) AB(28)
#   Dec: AC(29) AD(30) AE(31)
# AF=32: ИТОГО Факт
# AG=33: разделитель
# AH=34: параметр label
# AI=35: параметр value

MONTHS_RU = ['Январь','Февраль','Март','Апрель','Май','Июнь',
             'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

# Маппинг месяц (1-12) → колонки в V2
def v2_cols(month):   # 1-based
    if month <= 3:
        return (month + 1, None, None)      # B=2, C=3, D=4
    else:
        base = 4 + (month - 4) * 3 + 1     # Apr=5, May=8, Jun=11 ...
        return (base, base+1, base+2)

ИТОГО_COL = 32   # AF

# Маппинг месяц → колонка значений в ДДС 2026 (1-based)
def src26_col(month):
    return month * 2   # Jan=2, Feb=4, Mar=6, Apr=8 ...

# Маппинг месяц → колонка значений в ДДС 2025 (1-based)
# Jan=B(2), Feb=D(4), Mar=F(6), Apr=H(8), May=J(10), ...
def src25_col(month):
    return month * 2   # идентичная структура

# ─── CFO параметры планирования ─────────────────────────────────────────────
# Строки параметрической зоны (AH:AI, начиная с row 2)
# Row 2 = header
# Row 3 = ДОХОДЫ   → $AI$3
# Row 4 = ПРОИЗВОДСТВО → $AI$4
# Row 5 = ЛОГИСТИКА    → $AI$5
# Row 6 = МАРКЕТИНГ    → $AI$6
# Row 7 = ПРОДАЖИ      → $AI$7
# Row 8 = НАЛОГИ        → $AI$8
# Row 9 = ОПЕРАЦИОННЫЕ → $AI$9
BLOCK_GROWTH_CELL = {
    '__INCOME':      '$AI$3',
    'ПРОИЗВОДСТВО':  '$AI$4',
    'ЛОГИСТИКА':     '$AI$5',
    'МАРКЕТИНГ':     '$AI$6',
    'ПРОДАЖИ':       '$AI$7',
    'НАЛОГИ':        '$AI$8',
    'ОПЕРАЦИОННЫЕ':  '$AI$9',
}
BLOCK_GROWTH_DEFAULTS = {
    '__INCOME': 0.15, 'ПРОИЗВОДСТВО': 0.12, 'ЛОГИСТИКА': 0.10,
    'МАРКЕТИНГ': 0.10, 'ПРОДАЖИ': 0.08, 'НАЛОГИ': 0.08, 'ОПЕРАЦИОННЫЕ': 0.08,
}

# ─── Блок-маппинг ─────────────────────────────────────────────────────────
BLOCK_MAP = {
    'ФОТ оклад ЦЕХ':'ПРОИЗВОДСТВО','Аванс':'ПРОИЗВОДСТВО',
    'ФОТ больничный':'ПРОИЗВОДСТВО','ФОТ аутсорс':'ПРОИЗВОДСТВО','Бонусы':'ПРОИЗВОДСТВО',
    'Шампунь для ковров':'ПРОИЗВОДСТВО','Химия для чистки':'ПРОИЗВОДСТВО',
    'Химия для цеха':'ПРОИЗВОДСТВО','Пакеты':'ПРОИЗВОДСТВО','Бирки':'ПРОИЗВОДСТВО',
    'Скобы':'ПРОИЗВОДСТВО','Накладные':'ПРОИЗВОДСТВО','Скотч':'ПРОИЗВОДСТВО',
    'Хозбыт':'ПРОИЗВОДСТВО','Инструменты':'ПРОИЗВОДСТВО','Бытовые':'ПРОИЗВОДСТВО',
    'Содержание цеха':'ПРОИЗВОДСТВО','Аренда Агбиса':'ПРОИЗВОДСТВО',
    'Свет':'ПРОИЗВОДСТВО','Газ':'ПРОИЗВОДСТВО','Вода':'ПРОИЗВОДСТВО',
    'Продукты цех':'ПРОИЗВОДСТВО','Приобретение для кухни цеха':'ПРОИЗВОДСТВО',
    'Вывоз мусора':'ПРОИЗВОДСТВО','Септик':'ПРОИЗВОДСТВО',
    'Документы рабочим':'ПРОИЗВОДСТВО','Ремонт ковров':'ПРОИЗВОДСТВО',
    'Оплата услуг штор и тюлей':'ПРОИЗВОДСТВО',
    'Оборудование (обновление)':'ПРОИЗВОДСТВО',
    'Оборудование (тех.обслуживание)':'ПРОИЗВОДСТВО',
    'Приобретение для цеха':'ПРОИЗВОДСТВО',
    'ФОТ оклад Логистика':'ЛОГИСТИКА','Обед водителям':'ЛОГИСТИКА',
    'Приобретение для курьеров':'ЛОГИСТИКА','Такси/доставка':'ЛОГИСТИКА',
    'Абонент трекер':'ЛОГИСТИКА','ГСМ':'ЛОГИСТИКА','Мойка':'ЛОГИСТИКА',
    'Запчасти':'ЛОГИСТИКА','Замена шин/Вулканизация':'ЛОГИСТИКА',
    'Ремонт ходовой части':'ЛОГИСТИКА','Ремонт двигителя и коробки':'ЛОГИСТИКА',
    'Ремонт':'ЛОГИСТИКА','Замена масла':'ЛОГИСТИКА','Замена колодок':'ЛОГИСТИКА',
    'Штрафы':'ЛОГИСТИКА','Техосмотр':'ЛОГИСТИКА','Страховка':'ЛОГИСТИКА',
    'Прочее расходы':'ЛОГИСТИКА',
    'Покупка рекламы Google':'МАРКЕТИНГ','Покупка рекламы Yandex':'МАРКЕТИНГ',
    'Покупка рекламы Instagram':'МАРКЕТИНГ','Покупка рекламы TikTok':'МАРКЕТИНГ',
    'Покупка рекламы 2Гис':'МАРКЕТИНГ','Наружняя реклама':'МАРКЕТИНГ',
    'СММ':'МАРКЕТИНГ','Констектолог':'МАРКЕТИНГ','Таргетолог':'МАРКЕТИНГ',
    'Подписка сайта':'МАРКЕТИНГ','Оплата модели':'МАРКЕТИНГ','Видеограф':'МАРКЕТИНГ',
    'Бюджет на съемки':'МАРКЕТИНГ','Купоны и баннеры':'МАРКЕТИНГ',
    'Бартер/Блогер':'МАРКЕТИНГ','Обзвон':'МАРКЕТИНГ','Реклама мебели':'МАРКЕТИНГ',
    'ФОТ оклад Отел продаж':'ПРОДАЖИ','Телефония Кар Тел':'ПРОДАЖИ',
    'Wazzup подписка':'ПРОДАЖИ','Канцелярия':'ПРОДАЖИ',
    'Приобретение для отдела продаж':'ПРОДАЖИ',
    'Налоги_1':'НАЛОГИ','Тариф':'НАЛОГИ',
    'Интернет':'ОПЕРАЦИОННЫЕ','Подписка Битрикс':'ОПЕРАЦИОННЫЕ',
    'Бухгалтер':'ОПЕРАЦИОННЫЕ','Содержание офиса':'ОПЕРАЦИОННЫЕ',
    'Услиги юриста':'ОПЕРАЦИОННЫЕ','Услуги типографии':'ОПЕРАЦИОННЫЕ',
    'Тимбилдинг':'ОПЕРАЦИОННЫЕ','Покупка консультации':'ОПЕРАЦИОННЫЕ',
    'Обьяление для сотрудников':'ОПЕРАЦИОННЫЕ',
    'Единоразовый расход (приобретение)':'ОПЕРАЦИОННЫЕ',
    'Прочие расходы':'ОПЕРАЦИОННЫЕ','Покупка ковра':'ОПЕРАЦИОННЫЕ',
    'Услуга аренды машины':'ОПЕРАЦИОННЫЕ',
    'Погашение кредита':'__LOAN','Погашение займа':'__LOAN',
    'Kaspi Pay':'__WD','Наличными':'__WD',
}

SKIP = {
    'Поставщик','Поставщик_1','Хоз.рас материалы_1','Маркетинг_1',
    'Транспортные расходы','Транспортные расходы_1','ФОТ_1',
    'Финансовые операции','Финансовые операции_1',
    'Доходы','Расходы','Итого','Чистый доход','EBITDA',
    'Вывод средств','Кассовый разрыв','Кол-во заказов','Кол-во кв.м',
    'Баланс компании','Остаточный капитал','Самовывоз','Услуги',
    'тех.строка месяц','Пополнение','Получение банковского кредита',
    'Получение займа','ФОТ Собственник',
}

BLOCKS_ORDER = ['ПРОИЗВОДСТВО','ЛОГИСТИКА','МАРКЕТИНГ','ПРОДАЖИ','НАЛОГИ','ОПЕРАЦИОННЫЕ']
BLOCK_COLORS = {
    'ПРОИЗВОДСТВО':'8B5CF6','ЛОГИСТИКА':'F59E0B','МАРКЕТИНГ':'3B82F6',
    'ПРОДАЖИ':'10B981','НАЛОГИ':'6B7280','ОПЕРАЦИОННЫЕ':'EC4899',
}

# ─── Сканируем ДДС 2026: строим row_map ──────────────────────────────────────
row_map_26 = {}
income_itogo_row_26 = None
loan_src_rows  = []
wd_src_rows    = []
in_income = False

for i, row in enumerate(rows_26, start=1):
    name = row[0]
    if not name: continue
    name_s = str(name).strip()
    if name_s == 'Доходы':  in_income = True
    if name_s == 'Расходы': in_income = False
    if in_income and name_s == 'Итого':
        income_itogo_row_26 = i; continue
    if name_s in SKIP: continue
    row_map_26[name_s] = i
    blk = BLOCK_MAP.get(name_s)
    if blk == '__LOAN': loan_src_rows.append(i)
    if blk == '__WD' and i > 100: wd_src_rows.append(i)

block_items_26 = {b:[] for b in BLOCKS_ORDER}
for name_s, sr in row_map_26.items():
    blk = BLOCK_MAP.get(name_s)
    if blk and not blk.startswith('__') and blk in block_items_26:
        row_data = rows_26[sr-1]
        total = sum(row_data[src26_col(m)-1] or 0
                    for m in range(1,4)
                    if (src26_col(m)-1) < len(row_data) and isinstance(row_data[src26_col(m)-1],(int,float)))
        block_items_26[blk].append((name_s, sr, total))

for b in BLOCKS_ORDER:
    block_items_26[b].sort(key=lambda x: -x[2])

# ─── Сканируем ДДС 2025: строим row_map_25 ───────────────────────────────────
row_map_25 = {}
income_itogo_row_25 = None
in_income_25 = False

for i, row in enumerate(rows_25, start=1):
    name = row[0]
    if not name: continue
    name_s = str(name).strip()
    if name_s == 'Доходы':  in_income_25 = True
    if name_s == 'Расходы': in_income_25 = False
    if in_income_25 and name_s == 'Итого':
        income_itogo_row_25 = i; continue
    if name_s in SKIP: continue
    row_map_25[name_s] = i

print(f'  ДДС 2025: income_itogo_row={income_itogo_row_25}, items mapped={len(row_map_25)}')

# ─── Стили ───────────────────────────────────────────────────────────────────
def S(l='thin',r='thin',t='thin',b='thin',lc='CBD5E1',rc='CBD5E1',tc='CBD5E1',bc='CBD5E1'):
    def sd(s,c): return Side(style=s,color=c) if s else Side(style=None)
    return Border(left=sd(l,lc),right=sd(r,rc),top=sd(t,tc),bottom=sd(b,bc))

def F(bold=False,size=9,color='1E293B',italic=False):
    return Font(bold=bold,size=size,color=color,name='Calibri',italic=italic)
def aL(i=0): return Alignment(horizontal='left',  vertical='center',indent=i)
def aR():    return Alignment(horizontal='right', vertical='center')
def aC():    return Alignment(horizontal='center',vertical='center',wrap_text=True)

HEADER_FILL = PatternFill('solid',fgColor='1E293B')
WHITE_FILL  = PatternFill('solid',fgColor='FFFFFF')
GRAY_FILL   = PatternFill('solid',fgColor='F8FAFC')
PLAN_FILL   = PatternFill('solid',fgColor='FEFCE8')   # жёлтый = ввод плана
FACT_FILL   = PatternFill('solid',fgColor='F0FDF4')   # зелёный = факт
PCT_FILL    = PatternFill('solid',fgColor='EFF6FF')   # синий = %
TOTAL_FILL  = PatternFill('solid',fgColor='F1F5F9')
CHECK_FILL  = PatternFill('solid',fgColor='FEF9C3')
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'

# ─── Хелперы формул ──────────────────────────────────────────────────────────
def ref26(src_r, month):
    return f"='{SHEET_26}'!{get_column_letter(src26_col(month))}{src_r}"

def row_total_formula(row_n):
    parts = ['B','C','D']
    for m in range(4,13):
        fc,_,_ = v2_cols(m)
        parts.append(get_column_letter(fc))
    return '=' + '+'.join(f'{c}{row_n}' for c in parts)

def pct_formula(fact_col, plan_col, row_n):
    fc = get_column_letter(fact_col)
    pc = get_column_letter(plan_col)
    return f'=IF({pc}{row_n}>0,{fc}{row_n}/{pc}{row_n},"")'

def col_sum(col_letter, r_start, r_end):
    return f'=SUM({col_letter}{r_start}:{col_letter}{r_end})'

# ─── Запись ячейки ───────────────────────────────────────────────────────────
def sc(ws,row,col,val=None,fml=None,font=None,fill=None,align=None,brd=None,nfmt=None):
    c = ws.cell(row=row,column=col)
    c.value = fml if fml else val
    if font:  c.font  = font
    if fill:  c.fill  = fill
    if align: c.alignment = align
    if brd:   c.border= brd
    if nfmt:  c.number_format = nfmt
    return c

# ════════════════════════════════════════════════════════════════════════════
# ЛИСТ 1: ДДС 2026 V2
# ════════════════════════════════════════════════════════════════════════════
if 'ДДС 2026 V2' in wb.sheetnames: del wb['ДДС 2026 V2']
ws = wb.create_sheet('ДДС 2026 V2')

# ── Ширины колонок ──────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 36
for c in ['B','C','D']: ws.column_dimensions[c].width = 11
for m in range(4,13):
    fc, pc, vc = v2_cols(m)
    ws.column_dimensions[get_column_letter(fc)].width = 11
    ws.column_dimensions[get_column_letter(pc)].width = 11
    ws.column_dimensions[get_column_letter(vc)].width = 7
ws.column_dimensions[get_column_letter(ИТОГО_COL)].width = 13
ws.column_dimensions['AG'].width = 2    # разделитель
ws.column_dimensions['AH'].width = 30   # параметры label
ws.column_dimensions['AI'].width = 10   # параметры value

# ── Строка 1: Мега-заголовок ─────────────────────────────────────────────────
r = 1
ws.merge_cells(f'A{r}:{get_column_letter(ИТОГО_COL)}{r}')
sc(ws,r,1, val='ДДС 2026 — CFO Dashboard (Факт Q1 + Факт/План Q2-Q4)',
   font=F(bold=True,size=13,color='FFFFFF'), fill=HEADER_FILL, align=aC())
ws.row_dimensions[r].height = 28

# ── Строка 2: Группировка месяцев ────────────────────────────────────────────
r = 2
sc(ws,r,1, font=F(), fill=HEADER_FILL)
ws.merge_cells(f'B{r}:D{r}')
sc(ws,r,2, val='Q1 2026 — ФАКТ',
   font=F(bold=True,size=9,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='065F46'), align=aC())
q_ranges = [(4,6,'2Кв 2026'),(7,9,'3Кв 2026'),(10,12,'4Кв 2026')]
for (ms, me, ql) in q_ranges:
    fc_s,_,_ = v2_cols(ms)
    _,_,vc_e = v2_cols(me)
    ws.merge_cells(f'{get_column_letter(fc_s)}{r}:{get_column_letter(vc_e)}{r}')
    sc(ws,r,fc_s, val=ql,
       font=F(bold=True,size=9,color='FFFFFF'),
       fill=PatternFill('solid',fgColor='1D4ED8'), align=aC())
sc(ws,r,ИТОГО_COL, val='ИТОГО ФАКТ',
   font=F(bold=True,size=9,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='0F172A'), align=aC())
# Параметры: строка 2 — заголовок зоны
ws.merge_cells(f'AH{r}:AI{r}')
sc(ws,r,34, val='📊 CFO ПАРАМЕТРЫ',
   font=F(bold=True,size=9,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='1E3A5F'), align=aC())
ws.row_dimensions[r].height = 18

# ── Строка 3: Заголовки колонок + параметр ДОХОДЫ ────────────────────────────
r = 3
brd3 = S(t='medium',b='medium',tc='475569',bc='475569')
sc(ws,r,1, val='Статья', font=F(bold=True,size=10,color='FFFFFF'),
   fill=HEADER_FILL, align=aC(), brd=brd3)
q1_colors = ['166534','15803D','16A34A']
for idx, (m, qc) in enumerate(zip([1,2,3], q1_colors)):
    sc(ws,r, m+1, val=MONTHS_RU[m-1],
       font=F(bold=True,size=9,color='FFFFFF'),
       fill=PatternFill('solid',fgColor=qc), align=aC(), brd=brd3)
for m in range(4,13):
    fc, pc, vc = v2_cols(m)
    month_name = MONTHS_RU[m-1]
    sc(ws,r,fc, val=f'{month_name}\nФакт',
       font=F(bold=True,size=8,color='FFFFFF'),
       fill=PatternFill('solid',fgColor='065F46'), align=aC(), brd=brd3)
    sc(ws,r,pc, val=f'{month_name}\nПлан',
       font=F(bold=True,size=8,color='5F4C0A'),
       fill=PatternFill('solid',fgColor='CA8A04'), align=aC(), brd=brd3)
    sc(ws,r,vc, val='%',
       font=F(bold=True,size=8,color='1E3A5F'),
       fill=PatternFill('solid',fgColor='93C5FD'), align=aC(), brd=brd3)
sc(ws,r,ИТОГО_COL, val='ИТОГО',
   font=F(bold=True,size=10,color='FFFFFF'),
   fill=PatternFill('solid',fgColor='0F172A'), align=aC(), brd=brd3)
ws.row_dimensions[r].height = 30
ws.freeze_panes = 'B4'

# ── Записываем зону параметров CFO (строки 3-9, колонки AH:AI) ──────────────
PARAM_LABELS = [
    ('% роста ДОХОДЫ',        '__INCOME',      '065F46'),
    ('% роста ПРОИЗВОДСТВО',  'ПРОИЗВОДСТВО',  '7C3AED'),
    ('% роста ЛОГИСТИКА',     'ЛОГИСТИКА',     'B45309'),
    ('% роста МАРКЕТИНГ',     'МАРКЕТИНГ',     '1D4ED8'),
    ('% роста ПРОДАЖИ',       'ПРОДАЖИ',       '065F46'),
    ('% роста НАЛОГИ',        'НАЛОГИ',        '374151'),
    ('% роста ОПЕРАЦИОННЫЕ',  'ОПЕРАЦИОННЫЕ',  'BE185D'),
]
for pi, (label, blk, clr) in enumerate(PARAM_LABELS):
    pr = 3 + pi   # rows 3..9
    sc(ws, pr, 34, val=label,
       font=F(size=9, color='1E293B'),
       fill=PatternFill('solid', fgColor='F0F9FF'),
       align=aL(1),
       brd=S(l='medium', lc='3B82F6', r='thin', rc='CBD5E1'))
    c_val = ws.cell(row=pr, column=35)
    c_val.value = BLOCK_GROWTH_DEFAULTS[blk]
    c_val.number_format = '0%'
    c_val.font = Font(bold=True, size=11, color=clr, name='Calibri')
    c_val.fill = PLAN_FILL
    c_val.alignment = Alignment(horizontal='center', vertical='center')
    c_val.border = S(l='medium', r='medium', t='thin', b='thin',
                     lc='3B82F6', rc='3B82F6')
    ws.row_dimensions[pr].height = 18

# Подсказка под параметрами
pr_note = 10
ws.merge_cells(f'AH{pr_note}:AI{pr_note}')
sc(ws, pr_note, 34,
   val='↑ Введите % роста от ДДС 2025. Жёлтые = ввод.',
   font=F(size=8, color='64748B', italic=True),
   fill=PatternFill('solid', fgColor='F8FAFC'), align=aL())
ws.row_dimensions[pr_note].height = 14

# ── Хелпер: написать строку с данными (CFO план из ДДС 2025) ─────────────────
def write_data_row(ws, r, label, src_row_26, label_fill, fact_fill,
                   indent=2, bold=False, growth_cell=None, row_25_n=None):
    brd_l = S(l='medium',lc='94A3B8')
    brd_r = S(r='medium',rc='94A3B8')
    font  = F(bold=bold)

    sc(ws,r,1, val=label, font=font, fill=label_fill, align=aL(indent), brd=brd_l)

    # Q1: только факт (план не нужен — Q1 уже прошёл)
    for m in [1,2,3]:
        fc,_,_ = v2_cols(m)
        sc(ws,r,fc, fml=ref26(src_row_26,m),
           font=font, fill=FACT_FILL, align=aR(), brd=S(), nfmt=NUM_FMT)

    # Q2-Q4: факт | план (CFO формула) | %
    for m in range(4,13):
        fc,pc,vc = v2_cols(m)
        # Факт из ДДС 2026
        sc(ws,r,fc, fml=f"='{SHEET_26}'!{get_column_letter(src26_col(m))}{src_row_26}",
           font=font, fill=FACT_FILL, align=aR(), brd=S(), nfmt=NUM_FMT)
        # План — CFO: ДДС 2025 × (1 + % роста)
        if growth_cell and row_25_n:
            col_25 = get_column_letter(src25_col(m))
            plan_fml = f"='{SHEET_25}'!{col_25}{row_25_n}*(1+{growth_cell})"
            sc(ws,r,pc, fml=plan_fml,
               font=F(size=9,color='5F4C0A'), fill=PLAN_FILL, align=aR(), brd=S(), nfmt=NUM_FMT)
        else:
            sc(ws,r,pc, font=font, fill=PLAN_FILL, align=aR(), brd=S(), nfmt=NUM_FMT)
        # % выполнения
        sc(ws,r,vc, fml=pct_formula(fc,pc,r),
           font=F(size=9,color='1E40AF',italic=True), fill=PCT_FILL, align=aC(), nfmt=PCT_FMT)

    # ИТОГО
    sc(ws,r,ИТОГО_COL, fml=row_total_formula(r),
       font=F(bold=True), fill=TOTAL_FILL, align=aR(), brd=brd_r, nfmt=NUM_FMT)
    ws.row_dimensions[r].height = 16

# ── Хелпер: итоговая строка по диапазону (SUM) ─────────────────────────────
def write_block_total_row(ws, r, label, row_start, row_end, color, size=9):
    fill   = TOTAL_FILL
    brd_l  = S(l='medium',t='medium',b='medium',lc='94A3B8',tc='475569',bc='475569')
    brd_m  = S(t='medium',b='medium',tc='475569',bc='475569')
    brd_r  = S(r='medium',t='medium',b='medium',rc='94A3B8',tc='475569',bc='475569')
    font   = F(bold=True,size=size,color=color)

    sc(ws,r,1, val=label, font=font, fill=fill, align=aL(1), brd=brd_l)

    for m in [1,2,3]:
        fc,_,_ = v2_cols(m)
        cl = get_column_letter(fc)
        sc(ws,r,fc, fml=col_sum(cl,row_start,row_end),
           font=font, fill=fill, align=aR(), brd=brd_m, nfmt=NUM_FMT)

    for m in range(4,13):
        fc,pc,vc = v2_cols(m)
        cl_f = get_column_letter(fc)
        cl_p = get_column_letter(pc)
        sc(ws,r,fc, fml=col_sum(cl_f,row_start,row_end),
           font=font, fill=fill, align=aR(), brd=brd_m, nfmt=NUM_FMT)
        sc(ws,r,pc, fml=col_sum(cl_p,row_start,row_end),
           font=font, fill=PLAN_FILL, align=aR(), brd=brd_m, nfmt=NUM_FMT)
        sc(ws,r,vc, fml=pct_formula(fc,pc,r),
           font=F(size=9,color='1E40AF',bold=True), fill=PCT_FILL, align=aC(), nfmt=PCT_FMT)

    sc(ws,r,ИТОГО_COL, fml=row_total_formula(r),
       font=F(bold=True,size=size,color=color), fill=fill, align=aR(), brd=brd_r, nfmt=NUM_FMT)
    ws.row_dimensions[r].height = 20

# ── Хелпер: большая итоговая строка ──────────────────────────────────────────
def write_big_row(ws, r, label, formula_parts_fn, label_fill, val_fill, fcolor, fsize=11):
    brd_l  = S(l='medium',t='medium',b='medium',lc='94A3B8',tc='475569',bc='475569')
    brd_m  = S(t='medium',b='medium',tc='475569',bc='475569')
    brd_r  = S(r='medium',t='medium',b='medium',rc='94A3B8',tc='475569',bc='475569')
    font   = F(bold=True,size=fsize,color=fcolor)

    sc(ws,r,1, val=label, font=font, fill=label_fill, align=aL(1), brd=brd_l)
    for col_i in range(2, ИТОГО_COL+1):
        cl = get_column_letter(col_i)
        is_pct  = col_i >= 7 and ((col_i - 7) % 3 == 2)
        is_plan = col_i >= 6 and ((col_i - 6) % 3 == 0) and col_i < ИТОГО_COL
        if is_pct:
            sc(ws,r,col_i, fill=PCT_FILL, brd=brd_m)
            continue
        brd  = brd_r if col_i == ИТОГО_COL else brd_m
        fml  = formula_parts_fn(cl, r)
        sc(ws,r,col_i, fml=fml,
           font=font if not is_plan else F(bold=True,size=fsize,color='7C5100'),
           fill=label_fill if not is_plan else PLAN_FILL,
           align=aR(), brd=brd, nfmt=NUM_FMT)
    ws.row_dimensions[r].height = 22

def spacer(ws, r, color='F1F5F9'):
    for c in range(1,ИТОГО_COL+1):
        ws.cell(row=r,column=c).fill = PatternFill('solid',fgColor=color)
    ws.row_dimensions[r].height = 5

# ════ ЗАПИСЬ ЛИСТА ═══════════════════════════════════════════════════════════

# ── ДОХОДЫ ──────────────────────────────────────────────────────────────────
r = 4
spacer(ws,r); r+=1

rev_row = r
brd_l = S(l='medium',t='medium',b='medium',lc='94A3B8',tc='475569',bc='475569')
brd_m = S(t='medium',b='medium',tc='475569',bc='475569')
brd_r = S(r='medium',t='medium',b='medium',rc='94A3B8',tc='475569',bc='475569')
sc(ws,r,1, val='▶  ДОХОДЫ (Услуги)', font=F(bold=True,size=11,color='065F46'),
   fill=PatternFill('solid',fgColor='DCFCE7'), align=aL(1), brd=brd_l)

for m in [1,2,3]:
    fc,_,_ = v2_cols(m)
    sc(ws,r,fc, fml=f"='{SHEET_26}'!{get_column_letter(src26_col(m))}{income_itogo_row_26}",
       font=F(bold=True,size=11,color='065F46'),
       fill=PatternFill('solid',fgColor='F0FDF4'), align=aR(), brd=brd_m, nfmt=NUM_FMT)

for m in range(4,13):
    fc,pc,vc = v2_cols(m)
    sc(ws,r,fc, fml=f"='{SHEET_26}'!{get_column_letter(src26_col(m))}{income_itogo_row_26}",
       font=F(bold=True,size=11,color='065F46'),
       fill=PatternFill('solid',fgColor='F0FDF4'), align=aR(), brd=brd_m, nfmt=NUM_FMT)
    # CFO план доходов: ДДС 2025 × (1 + % роста доходов)
    if income_itogo_row_25:
        col_25 = get_column_letter(src25_col(m))
        plan_fml = f"='{SHEET_25}'!{col_25}{income_itogo_row_25}*(1+{BLOCK_GROWTH_CELL['__INCOME']})"
        sc(ws,r,pc, fml=plan_fml,
           font=F(bold=True,size=11,color='7C5100'),
           fill=PLAN_FILL, align=aR(), brd=brd_m, nfmt=NUM_FMT)
    else:
        sc(ws,r,pc, font=F(bold=True,size=11,color='7C5100'),
           fill=PLAN_FILL, align=aR(), brd=brd_m, nfmt=NUM_FMT)
    sc(ws,r,vc, fml=pct_formula(fc,pc,r),
       font=F(size=9,color='1E40AF',bold=True), fill=PCT_FILL, align=aC(), nfmt=PCT_FMT)

sc(ws,r,ИТОГО_COL, fml=row_total_formula(r),
   font=F(bold=True,size=11,color='065F46'),
   fill=PatternFill('solid',fgColor='DCFCE7'), align=aR(), brd=brd_r, nfmt=NUM_FMT)
ws.row_dimensions[r].height = 22
r+=1; spacer(ws,r); r+=1

# ── БЛОКИ РАСХОДОВ ──────────────────────────────────────────────────────────
block_total_rows_v2 = {}

for block in BLOCKS_ORDER:
    color = BLOCK_COLORS[block]
    items = block_items_26[block]
    cfill = PatternFill('solid',fgColor=color)
    growth_cell = BLOCK_GROWTH_CELL.get(block)

    # Заголовок блока
    brd_hb = S(l='medium',t='medium',lc='94A3B8',tc='475569')
    sc(ws,r,1, val=f'▶  {block}',
       font=F(bold=True,size=10,color='FFFFFF'), fill=cfill, align=aL(1), brd=brd_hb)
    for ci in range(2,ИТОГО_COL+1):
        ws.cell(row=r,column=ci).fill = cfill
        ws.cell(row=r,column=ci).border = S(t='medium',tc='475569')
    ws.row_dimensions[r].height = 20
    r+=1

    detail_start = r
    for idx,(name,src_row,_) in enumerate(items):
        row_fill = WHITE_FILL if idx%2==0 else GRAY_FILL
        row_25_n = row_map_25.get(name)
        write_data_row(ws, r, f'     {name}', src_row, row_fill, FACT_FILL,
                       growth_cell=growth_cell, row_25_n=row_25_n)
        r+=1
    detail_end = r-1

    block_total_rows_v2[block] = r
    write_block_total_row(ws,r, f'   Итого {block.capitalize()}',
                          detail_start, detail_end, color, size=9)
    r+=1; spacer(ws,r); r+=1

# ── ИТОГО COGS (без вывода) ──────────────────────────────────────────────────
cogs_row = r
bt_rows  = list(block_total_rows_v2.values())

def cogs_formula(cl, row_n):
    return '=' + '+'.join(f'{cl}{bt}' for bt in bt_rows)

write_big_row(ws,r,'▶  ИТОГО COGS (без вывода)',
              cogs_formula,
              PatternFill('solid',fgColor='FEE2E2'),
              PatternFill('solid',fgColor='FEF2F2'),
              '991B1B', fsize=11)
r+=1

# ── % Маржи (без вывода) ─────────────────────────────────────────────────────
margin_row = r
for col_i in range(2,ИТОГО_COL+1):
    cl = get_column_letter(col_i)
    is_pct_col  = col_i >= 7 and ((col_i - 7) % 3 == 2)
    is_plan_col = col_i >= 6 and ((col_i - 6) % 3 == 0) and col_i < ИТОГО_COL
    if is_pct_col or is_plan_col:
        ws.cell(row=r,column=col_i).fill = GRAY_FILL
        continue
    fml = f'=IF({cl}{rev_row}=0,0,({cl}{rev_row}-{cl}{cogs_row})/{cl}{rev_row})'
    sc(ws,r,col_i, fml=fml,
       font=F(size=9,color='065F46',italic=True),
       fill=PatternFill('solid',fgColor='F0FDF4'),
       align=aR(), nfmt=PCT_FMT)
sc(ws,r,1, val='     % Валовой маржи (без вывода)',
   font=F(size=9,color='065F46',italic=True),
   fill=PatternFill('solid',fgColor='F0FDF4'), align=aL(2),
   brd=S(l='medium',lc='94A3B8'))
ws.row_dimensions[r].height = 16
r+=1

# ── ВАЛОВАЯ ПРИБЫЛЬ (без вывода) ─────────────────────────────────────────────
gp_row = r
def gp_formula(cl, row_n): return f'={cl}{rev_row}-{cl}{cogs_row}'
write_big_row(ws,r,'▶  ВАЛОВАЯ ПРИБЫЛЬ (без вывода)',
              gp_formula,
              PatternFill('solid',fgColor='D1FAE5'),
              PatternFill('solid',fgColor='DCFCE7'),
              '065F46', fsize=11)
r+=1; spacer(ws,r); r+=1

# ════════════════════════════════════════════════════════════════════════════
# СЕКЦИЯ ВЫВОДА СРЕДСТВ + ЧЕКБОКС
# ════════════════════════════════════════════════════════════════════════════
ws.merge_cells(f'A{r}:{get_column_letter(ИТОГО_COL)}{r}')
sc(ws,r,1, val='── ВЫВОД СРЕДСТВ И ФИНАНСОВЫЕ ОПЕРАЦИИ (под чертой) ──',
   font=F(bold=True,size=10,color='7C3AED'),
   fill=PatternFill('solid',fgColor='EDE9FE'), align=aC())
ws.row_dimensions[r].height = 18
r+=1

# ── ЧЕКБОКС ─────────────────────────────────────────────────────────────────
ws.merge_cells(f'A{r}:C{r}')
sc(ws,r,1,
   val='⚙️  Включить вывод средств в COGS?',
   font=F(bold=True,size=10,color='7C3AED'),
   fill=PatternFill('solid',fgColor='F5F3FF'), align=aL(1))

CHECKBOX_CELL = f'E{r}'
sc(ws,r,5, val='Нет',
   font=F(bold=True,size=12,color='7C3AED'),
   fill=PatternFill('solid',fgColor='EDE9FE'), align=aC(),
   brd=S(l='medium',r='medium',t='medium',b='medium',
         lc='7C3AED',rc='7C3AED',tc='7C3AED',bc='7C3AED'))
sc(ws,r,6, val='← Да = включить в COGS | Нет = ниже черты',
   font=F(size=9,color='6B7280',italic=True), align=aL())
ws.row_dimensions[r].height = 22
checkbox_row = r
r+=1

dv = DataValidation(type='list', formula1='"Да,Нет"', allow_blank=False,
                    showDropDown=False)
dv.error = 'Выберите Да или Нет'
dv.errorTitle = 'Неверный ввод'
dv.prompt = 'Выберите Да или Нет'
dv.promptTitle = 'Включить в COGS?'
ws.add_data_validation(dv)
dv.sqref = CHECKBOX_CELL

# ── Строки Погашение кредита/займа ───────────────────────────────────────────
loan_detail_start = r
for src_row_n in loan_src_rows:
    name_s = str(rows_26[src_row_n-1][0]).strip()
    write_data_row(ws,r,f'     {name_s}',src_row_n,
                   PatternFill('solid',fgColor='FEF9C3'),
                   PatternFill('solid',fgColor='FEFCE8'))
    r+=1

# ── Строки Вывод средств ─────────────────────────────────────────────────────
wd_detail_start = r
for src_row_n in wd_src_rows:
    name_s = str(rows_26[src_row_n-1][0]).strip()
    brd_l  = S(l='medium',lc='94A3B8')
    brd_r2 = S(r='medium',rc='94A3B8')
    wd_fill = PatternFill('solid',fgColor='FEE2E2')
    sc(ws,r,1, val=f'     {name_s}',
       font=F(), fill=wd_fill, align=aL(2), brd=brd_l)
    for m in [1,2,3]:
        fc,_,_ = v2_cols(m)
        sc_col = src26_col(m)
        sc(ws,r,fc,
           fml=f"=ABS('{SHEET_26}'!{get_column_letter(sc_col)}{src_row_n})",
           font=F(), fill=wd_fill, align=aR(), brd=S(), nfmt=NUM_FMT)
    for m in range(4,13):
        fc,pc,vc = v2_cols(m)
        sc_col = src26_col(m)
        sc(ws,r,fc,
           fml=f"=ABS('{SHEET_26}'!{get_column_letter(sc_col)}{src_row_n})",
           font=F(), fill=wd_fill, align=aR(), brd=S(), nfmt=NUM_FMT)
        sc(ws,r,pc, fill=PLAN_FILL, align=aR(), brd=S(), nfmt=NUM_FMT)
        sc(ws,r,vc, fml=pct_formula(fc,pc,r),
           font=F(size=9,color='1E40AF'), fill=PCT_FILL, align=aC(), nfmt=PCT_FMT)
    sc(ws,r,ИТОГО_COL, fml=row_total_formula(r),
       font=F(bold=True), fill=wd_fill, align=aR(), brd=brd_r2, nfmt=NUM_FMT)
    ws.row_dimensions[r].height = 16
    r+=1

loan_wd_end = r-1
spacer(ws,r); r+=1

# ── ИТОГО ВЫВОДА ─────────────────────────────────────────────────────────────
wd_total_row = r
write_block_total_row(ws,r,'   Итого вывод + кредит',
                      loan_detail_start, loan_wd_end, '7C3AED', size=9)
r+=1; spacer(ws,r); r+=1

# ── ИТОГО COGS С УЧЁТОМ ГАЛОЧКИ ─────────────────────────────────────────────
cogs_with_wd_row = r

def cogs_full_formula(cl, row_n):
    base = '+'.join(f'{cl}{bt}' for bt in bt_rows)
    wd   = f'{cl}{wd_total_row}'
    return f'=IF({CHECKBOX_CELL}="Да",{base}+{wd},{base})'

write_big_row(ws,r,'▶  ИТОГО COGS (с учётом галочки)',
              cogs_full_formula,
              PatternFill('solid',fgColor='FEE2E2'),
              PatternFill('solid',fgColor='FEF2F2'),
              '991B1B', fsize=11)
r+=1

# ── % Валовой маржи (с учётом галочки) ──────────────────────────────────────
margin_checked_row = r
for col_i in range(2, ИТОГО_COL+1):
    cl = get_column_letter(col_i)
    is_pct_col  = col_i >= 7 and ((col_i - 7) % 3 == 2)
    is_plan_col = col_i >= 6 and ((col_i - 6) % 3 == 0) and col_i < ИТОГО_COL
    if is_pct_col or is_plan_col:
        ws.cell(row=r, column=col_i).fill = PatternFill('solid', fgColor='FEF2F2')
        continue
    fml = f'=IF({cl}{rev_row}=0,0,({cl}{rev_row}-{cl}{cogs_with_wd_row})/{cl}{rev_row})'
    sc(ws, r, col_i, fml=fml,
       font=F(size=9, color='991B1B', italic=True),
       fill=PatternFill('solid', fgColor='FEF2F2'),
       align=aR(), nfmt=PCT_FMT)
sc(ws, r, 1, val='     % Валовой маржи (с учётом галочки)',
   font=F(size=9, color='991B1B', italic=True),
   fill=PatternFill('solid', fgColor='FEF2F2'), align=aL(2),
   brd=S(l='medium', lc='94A3B8'))
ws.row_dimensions[r].height = 16
r += 1

# ── ЧИСТЫЙ ДОХОД ─────────────────────────────────────────────────────────────
net_row = r

def net_formula(cl, row_n):
    return f'={cl}{rev_row}-{cl}{cogs_with_wd_row}'

write_big_row(ws,r,'▶  ЧИСТЫЙ ДОХОД',
              net_formula,
              PatternFill('solid',fgColor='DBEAFE'),
              PatternFill('solid',fgColor='EFF6FF'),
              '1E40AF', fsize=11)
r+=1

# Пояснение
ws.merge_cells(f'A{r}:{get_column_letter(ИТОГО_COL)}{r}')
sc(ws,r,1,
   val=f'ℹ  Жёлтые ячейки (План) = расчёт авто по ДДС 2025 × % роста (зона AH:AI). Можно переписать вручную. Галочка {CHECKBOX_CELL}: "Да" → вывод входит в COGS.',
   font=F(size=9,color='64748B',italic=True),
   fill=PatternFill('solid',fgColor='F8FAFC'), align=aL())
ws.row_dimensions[r].height = 30

print(f'✓ Лист "ДДС 2026 V2" создан, строк: {r}, чекбокс: {CHECKBOX_CELL}')

# ════════════════════════════════════════════════════════════════════════════
# ЛИСТ 2: СЕБЕСТОИМОСТЬ V2
# ════════════════════════════════════════════════════════════════════════════
if 'Себестоимость V2' in wb.sheetnames: del wb['Себестоимость V2']
ws2 = wb.create_sheet('Себестоимость V2')

KVM_SHEET = SHEET_KVM
KVM_START = 2

ws2.column_dimensions['A'].width = 14
for ci, w in enumerate([12,12,16,14,14,12,14,14,12], start=2):
    ws2.column_dimensions[get_column_letter(ci)].width = w

r2 = 1
ws2.merge_cells(f'A{r2}:J{r2}')
sc(ws2,r2,1, val='СЕБЕСТОИМОСТЬ — Unit Economics по месяцам',
   font=F(bold=True,size=13,color='FFFFFF'), fill=HEADER_FILL, align=aC())
ws2.row_dimensions[r2].height = 28
r2+=1

ws2.merge_cells(f'A{r2}:J{r2}')
sc(ws2,r2,1,
   val=f'Данные: "{KVM_SHEET}" (кв.м, заказы) + "ДДС 2026 V2" (COGS)',
   font=F(size=9,color='64748B',italic=True),
   fill=PatternFill('solid',fgColor='F8FAFC'), align=aL())
ws2.row_dimensions[r2].height = 16
r2+=1

COLS_HDR = ['Месяц','Кол-во\nзаказов','Кол-во\nкв.м','Выручка (₸)',
            'COGS (₸)','Вал.прибыль','% Маржи',
            'Себест.\n₸/заказ','Себест.\n₸/кв.м','Ср.чек\n(₸/заказ)']
brd3 = S(t='medium',b='medium',tc='475569',bc='475569')
for ci, h in enumerate(COLS_HDR, start=1):
    sc(ws2,r2,ci, val=h, font=F(bold=True,size=9,color='FFFFFF'),
       fill=HEADER_FILL, align=aC(), brd=brd3)
ws2.row_dimensions[r2].height = 32
ws2.freeze_panes = 'B4'
r2+=1

MONTHS_LIST = ['Январь','Февраль','Март','Апрель','Май','Июнь',
               'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

for m_idx, month_name in enumerate(MONTHS_LIST):
    src_kvm_row = KVM_START + m_idx
    row_fill = WHITE_FILL if m_idx%2==0 else GRAY_FILL

    kvm_orders = f"='{KVM_SHEET}'!N{src_kvm_row}"
    kvm_sqm    = f"='{KVM_SHEET}'!L{src_kvm_row}"
    kvm_rev    = f"='{KVM_SHEET}'!M{src_kvm_row}"

    fc_v2,_,_ = v2_cols(m_idx+1)
    cogs_v2_ref = f"='ДДС 2026 V2'!{get_column_letter(fc_v2)}{cogs_with_wd_row}"

    sc(ws2,r2,1, val=month_name, font=F(bold=True,size=9), fill=row_fill, align=aL())
    sc(ws2,r2,2, fml=kvm_orders, font=F(size=9), fill=row_fill, align=aR(), nfmt='#,##0')
    sc(ws2,r2,3, fml=kvm_sqm,    font=F(size=9), fill=row_fill, align=aR(), nfmt='#,##0.0')
    sc(ws2,r2,4, fml=kvm_rev,    font=F(size=9), fill=row_fill, align=aR(), nfmt='#,##0')
    sc(ws2,r2,5, fml=cogs_v2_ref,font=F(size=9), fill=row_fill, align=aR(), nfmt='#,##0')

    cl_rev  = f'D{r2}'
    cl_cogs = f'E{r2}'
    cl_ord  = f'B{r2}'
    cl_sqm  = f'C{r2}'

    sc(ws2,r2,6, fml=f'=IF({cl_rev}>0,{cl_rev}-{cl_cogs},"")',
       font=F(size=9), fill=row_fill, align=aR(), nfmt='#,##0')
    sc(ws2,r2,7, fml=f'=IF({cl_rev}>0,({cl_rev}-{cl_cogs})/{cl_rev},"")',
       font=F(size=9,color='065F46'), fill=row_fill, align=aR(), nfmt='0.0%')
    sc(ws2,r2,8, fml=f'=IF({cl_ord}>0,{cl_cogs}/{cl_ord},"")',
       font=F(size=9), fill=row_fill, align=aR(), nfmt='#,##0')
    sc(ws2,r2,9, fml=f'=IF({cl_sqm}>0,{cl_cogs}/{cl_sqm},"")',
       font=F(size=9), fill=row_fill, align=aR(), nfmt='#,##0.0')
    sc(ws2,r2,10, fml=f'=IF({cl_ord}>0,{cl_rev}/{cl_ord},"")',
       font=F(size=9,color='1E40AF'), fill=row_fill, align=aR(), nfmt='#,##0')

    ws2.row_dimensions[r2].height = 18
    r2+=1

r2_total = r2
brd_t = S(t='medium',b='medium',tc='475569',bc='475569',l='medium',r='medium',lc='94A3B8',rc='94A3B8')
total_fill = PatternFill('solid',fgColor='E2E8F0')
sc(ws2,r2,1, val='ИТОГО', font=F(bold=True,size=10), fill=total_fill, align=aL(), brd=brd_t)
for ci, fml in enumerate([
    f'=SUM(B{r2_total-12}:B{r2_total-1})',
    f'=SUM(C{r2_total-12}:C{r2_total-1})',
    f'=SUM(D{r2_total-12}:D{r2_total-1})',
    f'=SUM(E{r2_total-12}:E{r2_total-1})',
    f'=IF(D{r2}>0,D{r2}-E{r2},"")',
    f'=IF(D{r2}>0,(D{r2}-E{r2})/D{r2},"")',
    f'=IF(B{r2}>0,E{r2}/B{r2},"")',
    f'=IF(C{r2}>0,E{r2}/C{r2},"")',
    f'=IF(B{r2}>0,D{r2}/B{r2},"")',
], start=2):
    nfmt = '0.0%' if ci==7 else ('#,##0.0' if ci in [3,9] else '#,##0')
    sc(ws2,r2,ci, fml=fml, font=F(bold=True,size=10),
       fill=total_fill, align=aR(), brd=brd_t, nfmt=nfmt)
ws2.row_dimensions[r2].height = 20

r2+=2
ws2.merge_cells(f'A{r2}:J{r2}')
sc(ws2,r2,1,
   val='ℹ  Выручка и Кол-во кв.м/заказов — из листа "Кол-во Кв.м_Заказов_Ср.чек". COGS — из "ДДС 2026 V2". Галочка в ДДС 2026 V2 влияет на включение вывода средств в COGS.',
   font=F(size=9,color='64748B',italic=True),
   fill=PatternFill('solid',fgColor='F8FAFC'), align=aL())
ws2.row_dimensions[r2].height = 28

print(f'✓ Лист "Себестоимость V2" создан, строк: {r2}')

# ─── Сохраняем ───────────────────────────────────────────────────────────────
wb.save(SRC_PATH)
print(f'\n✓ Файл сохранён: {SRC_PATH}')
print(f'  ДДС 2026 V2: чекбокс={CHECKBOX_CELL}, CFO зона AH:AI строки 3-9')
print(f'  Себестоимость V2: {len(MONTHS_LIST)} месяцев')
