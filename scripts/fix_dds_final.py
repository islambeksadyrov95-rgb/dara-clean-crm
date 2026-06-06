import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook
import re

FILE_BAK = "D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет_backup.xlsx"
FILE_OUT = "D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет.xlsx"

wb = load_workbook(FILE_BAK)
ws26 = wb["ДДС 2026 V2"]
ws25v2 = wb["ДДС 2025 V2"]

# Build name->row for ДДС 2025 V2
name_to_row_25v2 = {}
for row in range(1, ws25v2.max_row + 1):
    val = ws25v2.cell(row=row, column=1).value
    if val:
        name_to_row_25v2[val.strip()] = row

# Build row->name for ДДС 2026 V2
row_to_name_26v2 = {}
for row in range(1, ws26.max_row + 1):
    val = ws26.cell(row=row, column=1).value
    if val:
        row_to_name_26v2[row] = val.strip()

# Build row->name for ДДС 2025 V2
row_to_name_25v2 = {}
for row in range(1, ws25v2.max_row + 1):
    val = ws25v2.cell(row=row, column=1).value
    if val:
        row_to_name_25v2[row] = val.strip()

def find_in_25v2(name):
    clean = name.replace('\u25b6', '').replace('\u25ba', '').strip()
    for n25, r25 in name_to_row_25v2.items():
        if n25.replace('\u25b6', '').replace('\u25ba', '').strip() == clean:
            return r25
    return None

# Row mapping: 26v2 cell row -> 25v2 row
row_26_to_25v2 = {}
for r26, name in row_to_name_26v2.items():
    r25 = find_in_25v2(name)
    if r25 is not None:
        row_26_to_25v2[r26] = r25

print(f"Mapped {len(row_26_to_25v2)} rows")

# Month column mapping: old 'ДДС 2025' col -> ДДС 2025 V2 col
old_col_to_v2_col = {
    'H': 'E',  # April
    'J': 'F',  # May
    'L': 'G',  # June
    'N': 'H',  # July
    'P': 'I',  # August
    'R': 'J',  # September
    'T': 'K',  # October
    'V': 'L',  # November
    'X': 'M',  # December
}

# Test regex against known formula format
test_formula = "='ДДС 2025'!H15*(1+$AI$3)"
# The formula string stored in openpyxl does NOT have leading =
# openpyxl stores formulas as strings starting with =
# Let's check what the actual stored value looks like
print(f"\nTest formula: {repr(test_formula)}")

# Pattern for: ='ДДС 2025'!COL ROW*(1+$AI$N)
# Note: $ in the formula is a literal dollar sign
pat = re.compile(r"^='ДДС 2025'!([A-Z]+)(\d+)\*\(1\+\$AI\$(\d+)\)$")
m = pat.match(test_formula)
print(f"Pattern match test: {m}")

fixed_count = 0
skipped = []

print("\n=== Processing formulas ===")
for row in ws26.iter_rows():
    for cell in row:
        val = cell.value
        if not isinstance(val, str):
            continue
        if "'ДДС 2025'!" not in val:
            continue
        if "'ДДС 2025 V2'!" in val:
            continue

        m = pat.match(val)
        if not m:
            print(f"  UNMATCHED at {cell.coordinate}: {repr(val)}")
            skipped.append((cell.coordinate, val))
            continue

        old_col = m.group(1)
        ai_num  = m.group(3)

        # Get new column in ДДС 2025 V2
        new_col = old_col_to_v2_col.get(old_col)
        if new_col is None:
            print(f"  UNKNOWN col '{old_col}' at {cell.coordinate}: {val}")
            skipped.append((cell.coordinate, val))
            continue

        # Get new row by matching item name
        cell_row = cell.row
        new_row = row_26_to_25v2.get(cell_row)
        if new_row is None:
            print(f"  NO ROW MAP for row {cell_row} at {cell.coordinate}: {val}")
            skipped.append((cell.coordinate, val))
            continue

        new_formula = f"='ДДС 2025 V2'!{new_col}{new_row}*(1+$AI${ai_num})"
        cell.value = new_formula
        fixed_count += 1

        n26 = row_to_name_26v2.get(cell_row, "?")[:25]
        n25 = row_to_name_25v2.get(new_row, "?")[:25]
        print(f"  {cell.coordinate}: {val} -> {new_formula}  [{n26} -> {n25}]")

print(f"\n=== DONE: Fixed {fixed_count} formulas, Skipped {len(skipped)} ===")
if skipped:
    print("Skipped:")
    for coord, v in skipped:
        print(f"  {coord}: {v}")

# Save to output file
wb.save(FILE_OUT)
print(f"\nSaved to: {FILE_OUT}")
