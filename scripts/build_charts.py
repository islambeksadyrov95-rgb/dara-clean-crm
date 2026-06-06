# -*- coding: utf-8 -*-
"""
CFO Dashboard — создаёт лист "📊 Диаграммы" с charts по ДДС/Себестоимость/Воронка/ПланПродаж
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabel
from copy import deepcopy

SRC = "D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет — копия.xlsx"
DST = "D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет — DASHBOARD.xlsx"

MONTHS_RU = ['Январь','Февраль','Март','Апрель','Май','Июнь',
             'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

# ─── цвета CFO ───────────────────────────────────────────────────────────────
C_DARK   = "1F4E79"   # тёмно-синий
C_BLUE   = "2E75B6"   # синий
C_GREEN  = "70AD47"   # зелёный
C_RED    = "FF0000"   # красный
C_GRAY   = "A6A6A6"   # серый (план)
C_YELLOW = "FFC000"   # жёлтый
C_WHITE  = "FFFFFF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── загрузка данных ─────────────────────────────────────────────────────────
print("Загрузка файла...")
wb = load_workbook(SRC, data_only=True)

ws25 = wb['ДДС 2025 V2']
ws26 = wb['ДДС 2026 V2']
wsS  = wb['Себестоимость V2']
wsV  = wb['Воронка']
wsP  = wb['План продаж']

def val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None: return 0
    try: return float(v)
    except: return 0

# ─── Извлечение ДДС 2025 ─────────────────────────────────────────────────────
# Row 4 = Доходы, Row 124 = Чистый доход (cols 2-13 = Jan-Dec)
dds25_income  = [val(ws25, 4, c) for c in range(2, 14)]
dds25_net     = [val(ws25, 124, c) for c in range(2, 14)]
dds25_prod    = [val(ws25, 38, c) for c in range(2, 14)]
dds25_mkt     = [val(ws25, 79, c) for c in range(2, 14)]
dds25_log     = [val(ws25, 59, c) for c in range(2, 14)]

# ─── Извлечение ДДС 2026 ─────────────────────────────────────────────────────
# Структура: Jan=2, Feb=3, Mar=4, Apr_fact=5, Apr_plan=6, %=7, May_fact=8, May_plan=9, %=10...
# Факт: 2,3,4, 5,8,11,14,17,20,23,26,29
# План: 0,0,0, 6,9,12,15,18,21,24,27,30  (0 = нет плана, только факт)
fact_cols = [2,3,4,5,8,11,14,17,20,23,26,29]
plan_cols = [0,0,0,6,9,12,15,18,21,24,27,30]

dds26_income_fact = [val(ws26, 5, c) if c else 0 for c in fact_cols]
dds26_income_plan = [val(ws26, 5, c) if c else 0 for c in plan_cols]
dds26_net_fact    = [val(ws26, 126, c) if c else 0 for c in fact_cols]
dds26_net_plan    = [val(ws26, 126, c) if c else 0 for c in plan_cols]

# ─── Извлечение Себестоимость V2 ────────────────────────────────────────────
# Rows 4-15 = Jan-Dec, cols: Месяц=1, Заказы=2, Кв.м=3, Выручка=4, COGS=5, ВалПриб=6, %Маржи=7, СебЗак=8, СебКвм=9, СрЧек=10
seb_months = []
seb_revenue, seb_cogs, seb_margin_pct, seb_orders = [], [], [], []
seb_cac_order, seb_avg_check = [], []
for r in range(4, 16):
    m = wsS.cell(row=r, column=1).value
    if not m: continue
    seb_months.append(str(m))
    seb_orders.append(val(wsS, r, 2))
    seb_revenue.append(val(wsS, r, 4))
    seb_cogs.append(val(wsS, r, 5))
    gross = val(wsS, r, 6)
    seb_margin_pct.append(round(val(wsS, r, 7) * 100, 1) if val(wsS, r, 7) else 0)
    seb_cac_order.append(val(wsS, r, 8))
    seb_avg_check.append(val(wsS, r, 10))

# ─── Извлечение Воронка ──────────────────────────────────────────────────────
# Fact: rows 5-7 (Янв-Мар), Plan: rows 11-19 (Апр-Дек)
# Cols: Месяц=1, Маркетинг=2, Обращений=3, Продаж=4, Конверсия=5, CAC=6, СрЧек=7, LTV=8, Выручка=9, ROMI=10
funnel_months, funnel_leads, funnel_sales, funnel_conv, funnel_cac, funnel_revenue, funnel_romi, funnel_mkt = [], [], [], [], [], [], [], []

# Q1 fact rows 5-7
for r in [5, 6, 7]:
    m = wsV.cell(row=r, column=1).value
    if not m or 'ИТОГО' in str(m): continue
    funnel_months.append(str(m))
    funnel_mkt.append(val(wsV, r, 2))
    funnel_leads.append(val(wsV, r, 3))
    funnel_sales.append(val(wsV, r, 4))
    funnel_conv.append(round(val(wsV, r, 5)*100, 1))
    funnel_cac.append(val(wsV, r, 6))
    funnel_revenue.append(val(wsV, r, 9))
    funnel_romi.append(round(val(wsV, r, 10)*100, 1))

# Plan rows 11-19 (Апр-Дек)
for r in range(11, 21):
    m = wsV.cell(row=r, column=1).value
    if not m or 'ИТОГО' in str(m): continue
    funnel_months.append(str(m))
    funnel_mkt.append(val(wsV, r, 2))
    funnel_leads.append(val(wsV, r, 3))
    funnel_sales.append(val(wsV, r, 4))
    funnel_conv.append(round(val(wsV, r, 5)*100, 1))
    funnel_cac.append(val(wsV, r, 6))
    funnel_revenue.append(val(wsV, r, 9))
    funnel_romi.append(round(val(wsV, r, 10)*100, 1))

# ─── Извлечение План продаж (по месяцам) ────────────────────────────────────
# 8 блоков, каждый по 8 столбцов, начинаются с col 1, 9, 17, 25, 33, 41, 49, 57
# В каждом блоке: col+0=Дата, col+4=Выручка план, данные в rows 11-41
plan_months_names = ['Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
plan_start_cols = [1, 9, 17, 25, 33, 41, 49, 57]
plan_monthly_revenue = []   # суммарная выручка за месяц
plan_monthly_orders  = []   # суммарно заказов

for mi, sc in enumerate(plan_start_cols):
    rev_col = sc + 4   # Выручка план ₸
    ord_col = sc + 3   # Заказов план
    total_rev = 0
    total_ord = 0
    for r in range(11, 42):
        d = wsP.cell(row=r, column=sc).value
        if d is None: break
        v = wsP.cell(row=r, column=rev_col).value
        o = wsP.cell(row=r, column=ord_col).value
        if v: total_rev += float(v)
        if o: total_ord += float(o)
    plan_monthly_revenue.append(total_rev)
    plan_monthly_orders.append(total_ord)

print(f"  Воронка: {len(funnel_months)} месяцев")
print(f"  Себестоимость: {len(seb_months)} месяцев")
print(f"  Plan продаж (Май-Дек): {plan_monthly_revenue}")

# ─── Создание листа "📊 Диаграммы" ──────────────────────────────────────────
SHEET_NAME = "📊 Диаграммы"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]

ws = wb.create_sheet(SHEET_NAME, 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 18
for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N']:
    ws.column_dimensions[col_letter].width = 13
ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 20

# ─── ШАПКА ───────────────────────────────────────────────────────────────────
ws.merge_cells('A1:N1')
c = ws['A1']
c.value = "CFO DASHBOARD — DaraClean 2025–2026"
c.font = Font(name="Calibri", size=18, bold=True, color=C_WHITE)
c.fill = fill(C_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:N2')
c = ws['A2']
c.value = "📊 ДДС • Себестоимость • Воронка • План продаж"
c.font = Font(name="Calibri", size=11, color="A0C4E8")
c.fill = fill(C_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')

# ─────────────────────────────────────────────────────────────────────────────
# БЛОК 1: Данные ДДС для графиков (строки 5–30)
# ─────────────────────────────────────────────────────────────────────────────
data_start = 5

# Заголовок блока
ws.merge_cells(f'A{data_start}:N{data_start}')
hdr = ws[f'A{data_start}']
hdr.value = "1. ДДС 2025 vs 2026 — Доходы и Чистый результат"
hdr.font = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
hdr.fill = fill(C_BLUE)
hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[data_start].height = 28

# Заголовки столбцов
row = data_start + 1
headers = ['Показатель'] + MONTHS_RU + ['ИТОГО']
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=ci, value=h)
    c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()
ws.row_dimensions[row].height = 30

def write_data_row(ws, row, label, values, fmt='money', label_color=C_DARK, val_color="000000", bold_label=False):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Calibri", size=9, bold=bold_label, color=C_WHITE if label_color != "000000" else "000000")
    c.fill = fill(label_color) if label_color != "000000" else PatternFill()
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    c.border = thin_border()
    total = 0
    for ci, v in enumerate(values, 2):
        cell = ws.cell(row=row, column=ci, value=v if v != 0 else None)
        cell.font = Font(name="Calibri", size=9, color=val_color if v >= 0 else C_RED)
        cell.fill = PatternFill("solid", fgColor="F5F9FF") if (ci % 2 == 0) else PatternFill("solid", fgColor="EAF2FF")
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border()
        if fmt == 'money':
            cell.number_format = '#,##0'
        elif fmt == 'pct':
            cell.number_format = '0.0%'
        total += v if v else 0
    # ИТОГО col (col 14)
    tot_cell = ws.cell(row=row, column=14, value=total if total != 0 else None)
    tot_cell.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE if label_color != "000000" else ("000000" if total >= 0 else C_RED))
    tot_cell.fill = fill(C_DARK)
    tot_cell.alignment = Alignment(horizontal='right', vertical='center')
    tot_cell.border = thin_border()
    if fmt == 'money':
        tot_cell.number_format = '#,##0'
    ws.row_dimensions[row].height = 18

# ДДС 2025
r = data_start + 2
write_data_row(ws, r,   "Доходы 2025 (факт)", dds25_income, bold_label=True, label_color=C_DARK)
write_data_row(ws, r+1, "Чистый ДДС 2025 (факт)", dds25_net,  label_color=C_GREEN)
write_data_row(ws, r+2, "— Производство 2025",  dds25_prod,  label_color="4A4A4A")
write_data_row(ws, r+3, "— Маркетинг 2025",      dds25_mkt,   label_color="4A4A4A")
write_data_row(ws, r+4, "— Логистика 2025",      dds25_log,   label_color="4A4A4A")

# ДДС 2026
write_data_row(ws, r+6, "Доходы 2026 (факт)",  dds26_income_fact, bold_label=True, label_color=C_DARK)
write_data_row(ws, r+7, "Доходы 2026 (план)",  dds26_income_plan, bold_label=False, label_color=C_GRAY)
write_data_row(ws, r+8, "Чистый ДДС 2026 (факт)", dds26_net_fact, label_color=C_GREEN)
write_data_row(ws, r+9, "Чистый ДДС 2026 (план)", dds26_net_plan, label_color=C_GRAY)

# Разделитель
sep_row = r + 11
ws.row_dimensions[sep_row].height = 6

# ─────────────────────────────────────────────────────────────────────────────
# БЛОК 2: Себестоимость данные
# ─────────────────────────────────────────────────────────────────────────────
seb_start = sep_row + 1
ws.merge_cells(f'A{seb_start}:N{seb_start}')
hdr = ws[f'A{seb_start}']
hdr.value = "2. Себестоимость V2 — Unit Economics по месяцам"
hdr.font = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
hdr.fill = fill(C_BLUE)
hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[seb_start].height = 28

# Заголовки
seb_hdr_row = seb_start + 1
seb_cols = ['Месяц','Заказов','Выручка ₸','COGS ₸','Вал.Прибыль','Маржа %','Себ-ть/заказ','Ср.чек']
for ci, h in enumerate(seb_cols, 1):
    c = ws.cell(row=seb_hdr_row, column=ci, value=h)
    c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
ws.row_dimensions[seb_hdr_row].height = 24

for i, m in enumerate(seb_months):
    r2 = seb_hdr_row + 1 + i
    gross_p = seb_revenue[i] - seb_cogs[i]
    vals = [m, int(seb_orders[i]) if seb_orders[i] else 0,
            seb_revenue[i], seb_cogs[i], gross_p, seb_margin_pct[i]/100,
            seb_cac_order[i], seb_avg_check[i]]
    fmts = ['text','#,##0','#,##0','#,##0','#,##0','0.0%','#,##0','#,##0']
    bg = "F5F9FF" if i % 2 == 0 else "EAF2FF"
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws.cell(row=r2, column=ci, value=v)
        cell.font = Font(name="Calibri", size=9,
                         color=C_RED if (ci == 5 and isinstance(v, (int,float)) and v < 0) else "000000")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='right' if ci > 1 else 'left', vertical='center',
                                   indent=1 if ci == 1 else 0)
        cell.border = thin_border()
        if fmt != 'text':
            cell.number_format = fmt
    ws.row_dimensions[r2].height = 18

seb_data_end = seb_hdr_row + len(seb_months)

# ─────────────────────────────────────────────────────────────────────────────
# БЛОК 3: Воронка данные
# ─────────────────────────────────────────────────────────────────────────────
fun_sep = seb_data_end + 2
ws.row_dimensions[fun_sep].height = 6
fun_start = fun_sep + 1

ws.merge_cells(f'A{fun_start}:N{fun_start}')
hdr = ws[f'A{fun_start}']
hdr.value = "3. Воронка продаж — Факт Q1 2026 + План Apr–Dec"
hdr.font = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
hdr.fill = fill(C_BLUE)
hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[fun_start].height = 28

fun_hdr_row = fun_start + 1
fun_cols = ['Месяц','Маркетинг ₸','Обращений','Продаж','Конверсия %','CAC ₸','Выручка ₸','ROMI %']
for ci, h in enumerate(fun_cols, 1):
    c = ws.cell(row=fun_hdr_row, column=ci, value=h)
    c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
ws.row_dimensions[fun_hdr_row].height = 24

q1_months = ['Январь','Февраль','Март']
for i, m in enumerate(funnel_months):
    r2 = fun_hdr_row + 1 + i
    is_fact = m in q1_months
    bg = "E8F5E9" if is_fact else "FFF9E6"   # зелёный = факт, жёлтый = план
    vals = [m, funnel_mkt[i], funnel_leads[i], funnel_sales[i],
            funnel_conv[i]/100, funnel_cac[i], funnel_revenue[i], funnel_romi[i]/100]
    fmts = ['text','#,##0','#,##0','#,##0','0.0%','#,##0','#,##0','0.0%']
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws.cell(row=r2, column=ci, value=v)
        cell.font = Font(name="Calibri", size=9, bold=(i < 3))
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='right' if ci > 1 else 'left', vertical='center',
                                   indent=1 if ci == 1 else 0)
        cell.border = thin_border()
        if fmt != 'text': cell.number_format = fmt
    ws.row_dimensions[r2].height = 18

fun_data_end = fun_hdr_row + len(funnel_months)

# ─────────────────────────────────────────────────────────────────────────────
# БЛОК 4: План продаж (Май–Декабрь)
# ─────────────────────────────────────────────────────────────────────────────
plan_sep = fun_data_end + 2
ws.row_dimensions[plan_sep].height = 6
plan_start = plan_sep + 1

ws.merge_cells(f'A{plan_start}:N{plan_start}')
hdr = ws[f'A{plan_start}']
hdr.value = "4. План продаж Май–Декабрь 2026 — Ежемесячный таргет"
hdr.font = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
hdr.fill = fill(C_BLUE)
hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[plan_start].height = 28

plan_hdr_row = plan_start + 1
plan_hdr_cols = ['Месяц','Выручка план ₸','Заказов план','Ср. чек ₸']
for ci, h in enumerate(plan_hdr_cols, 1):
    c = ws.cell(row=plan_hdr_row, column=ci, value=h)
    c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
ws.row_dimensions[plan_hdr_row].height = 24

for i, m in enumerate(plan_months_names):
    r2 = plan_hdr_row + 1 + i
    avg_check = (plan_monthly_revenue[i] / plan_monthly_orders[i]) if plan_monthly_orders[i] else 0
    vals = [m, plan_monthly_revenue[i], round(plan_monthly_orders[i], 0), round(avg_check, 0)]
    fmts = ['text','#,##0','#,##0','#,##0']
    bg = "F5F9FF" if i % 2 == 0 else "EAF2FF"
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws.cell(row=r2, column=ci, value=v)
        cell.font = Font(name="Calibri", size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='right' if ci > 1 else 'left', vertical='center',
                                   indent=1 if ci == 1 else 0)
        cell.border = thin_border()
        if fmt != 'text': cell.number_format = fmt
    ws.row_dimensions[r2].height = 18

plan_data_end = plan_hdr_row + len(plan_months_names)

print("Данные записаны. Создаю диаграммы...")

# =============================================================================
# ДИАГРАММЫ
# =============================================================================
CHART_COL   = 'E'   # начинаем с колонки E
CHART_COL_I = 5     # числовой индекс
CHART_WIDTH  = 22   # единицы Excel (ширина)
CHART_HEIGHT = 14   # единицы Excel (высота)

def make_chart_title(text, size=12):
    from openpyxl.chart.title import Title
    from openpyxl.drawing.text import RichText, RichTextProperties, Paragraph, Run, RegularTextRun, ParagraphProperties, CharacterProperties
    # Простой способ — через chart.title
    return text

# ─────────────────────────────────────────────────────────────────────────────
# CHART 1: Доходы 2025 (факт) vs Доходы 2026 (факт + план)
# ─────────────────────────────────────────────────────────────────────────────
c1 = BarChart()
c1.type    = "col"
c1.grouping = "clustered"
c1.title   = "Доходы: 2025 факт vs 2026 факт/план (₸)"
c1.style   = 10
c1.y_axis.title = "₸"
c1.x_axis.title = "Месяц"
c1.width   = CHART_WIDTH
c1.height  = CHART_HEIGHT

# Месяцы (категории) — из строки data_start+1, cols B:M
cats = Reference(ws, min_col=2, max_col=13, min_row=data_start+1)

# Серия 1: Доходы 2025
ref25 = Reference(ws, min_col=2, max_col=13, min_row=data_start+2)
s1 = Series(ref25, title="Доходы 2025 (факт)")
s1.graphicalProperties.solidFill = C_BLUE
c1.append(s1)

# Серия 2: Доходы 2026 факт
ref26f = Reference(ws, min_col=2, max_col=13, min_row=data_start+8)
s2 = Series(ref26f, title="Доходы 2026 (факт)")
s2.graphicalProperties.solidFill = C_GREEN
c1.append(s2)

# Серия 3: Доходы 2026 план
ref26p = Reference(ws, min_col=2, max_col=13, min_row=data_start+9)
s3 = Series(ref26p, title="Доходы 2026 (план)")
s3.graphicalProperties.solidFill = C_GRAY
c1.append(s3)

c1.set_categories(cats)
ws.add_chart(c1, f"E{data_start}")

# ─────────────────────────────────────────────────────────────────────────────
# CHART 2: Чистый ДДС 2025 vs 2026 (факт и план)
# ─────────────────────────────────────────────────────────────────────────────
c2 = BarChart()
c2.type     = "col"
c2.grouping = "clustered"
c2.title    = "Чистый ДДС: 2025 факт vs 2026 факт/план (₸)"
c2.style    = 10
c2.y_axis.title = "₸"
c2.x_axis.title = "Месяц"
c2.width    = CHART_WIDTH
c2.height   = CHART_HEIGHT

ref_net25 = Reference(ws, min_col=2, max_col=13, min_row=data_start+3)
sn1 = Series(ref_net25, title="Чистый ДДС 2025 (факт)")
sn1.graphicalProperties.solidFill = C_BLUE
c2.append(sn1)

ref_net26f = Reference(ws, min_col=2, max_col=13, min_row=data_start+10)
sn2 = Series(ref_net26f, title="Чистый ДДС 2026 (факт)")
sn2.graphicalProperties.solidFill = C_GREEN
c2.append(sn2)

ref_net26p = Reference(ws, min_col=2, max_col=13, min_row=data_start+11)
sn3 = Series(ref_net26p, title="Чистый ДДС 2026 (план)")
sn3.graphicalProperties.solidFill = C_GRAY
c2.append(sn3)

c2.set_categories(cats)

# Размещаем chart 2 правее chart 1
c2_col = 'K'
ws.add_chart(c2, f"{c2_col}{data_start}")

# ─────────────────────────────────────────────────────────────────────────────
# CHART 3: Себестоимость — Выручка vs COGS по месяцам
# ─────────────────────────────────────────────────────────────────────────────
c3 = BarChart()
c3.type     = "col"
c3.grouping = "clustered"
c3.title    = "Себестоимость: Выручка vs COGS по месяцам (₸)"
c3.style    = 10
c3.y_axis.title = "₸"
c3.x_axis.title = "Месяц"
c3.width    = CHART_WIDTH
c3.height   = CHART_HEIGHT

cats_seb = Reference(ws, min_col=1, max_col=1, min_row=seb_hdr_row+1, max_row=seb_data_end)

ref_rev = Reference(ws, min_col=3, max_col=3, min_row=seb_hdr_row, max_row=seb_data_end)
sr1 = Series(ref_rev, title="Выручка")
sr1.graphicalProperties.solidFill = C_BLUE
c3.append(sr1)

ref_cogs = Reference(ws, min_col=4, max_col=4, min_row=seb_hdr_row, max_row=seb_data_end)
sr2 = Series(ref_cogs, title="COGS")
sr2.graphicalProperties.solidFill = C_RED
c3.append(sr2)

ref_gross = Reference(ws, min_col=5, max_col=5, min_row=seb_hdr_row, max_row=seb_data_end)
sr3 = Series(ref_gross, title="Вал. прибыль")
sr3.graphicalProperties.solidFill = C_GREEN
c3.append(sr3)

c3.set_categories(cats_seb)
ws.add_chart(c3, f"E{seb_start}")

# ─────────────────────────────────────────────────────────────────────────────
# CHART 4: Себестоимость — Маржа % по месяцам (линия)
# ─────────────────────────────────────────────────────────────────────────────
c4 = LineChart()
c4.title    = "Маржинальность % по месяцам"
c4.style    = 10
c4.y_axis.title = "%"
c4.x_axis.title = "Месяц"
c4.width    = CHART_WIDTH
c4.height   = CHART_HEIGHT

ref_margin = Reference(ws, min_col=6, max_col=6, min_row=seb_hdr_row, max_row=seb_data_end)
sm1 = Series(ref_margin, title="Маржа %")
sm1.graphicalProperties.line.solidFill = C_BLUE
sm1.graphicalProperties.line.width = 25000
sm1.smooth = True
c4.append(sm1)

c4.set_categories(cats_seb)
ws.add_chart(c4, f"K{seb_start}")

# ─────────────────────────────────────────────────────────────────────────────
# CHART 5: Воронка — Лиды vs Продажи по месяцам
# ─────────────────────────────────────────────────────────────────────────────
c5 = BarChart()
c5.type     = "col"
c5.grouping = "clustered"
c5.title    = "Воронка: Лиды vs Продажи (факт Q1 + план Apr–Dec)"
c5.style    = 10
c5.y_axis.title = "Кол-во"
c5.x_axis.title = "Месяц"
c5.width    = CHART_WIDTH
c5.height   = CHART_HEIGHT

cats_fun = Reference(ws, min_col=1, max_col=1, min_row=fun_hdr_row+1, max_row=fun_data_end)

ref_leads = Reference(ws, min_col=3, max_col=3, min_row=fun_hdr_row, max_row=fun_data_end)
sf1 = Series(ref_leads, title="Обращений")
sf1.graphicalProperties.solidFill = C_BLUE
c5.append(sf1)

ref_sales = Reference(ws, min_col=4, max_col=4, min_row=fun_hdr_row, max_row=fun_data_end)
sf2 = Series(ref_sales, title="Продаж")
sf2.graphicalProperties.solidFill = C_GREEN
c5.append(sf2)

c5.set_categories(cats_fun)
ws.add_chart(c5, f"E{fun_start}")

# ─────────────────────────────────────────────────────────────────────────────
# CHART 6: Воронка — Выручка + Маркетинг + ROMI
# ─────────────────────────────────────────────────────────────────────────────
c6 = BarChart()
c6.type     = "col"
c6.grouping = "clustered"
c6.title    = "Маркетинг vs Выручка (факт Q1 + план)"
c6.style    = 10
c6.y_axis.title = "₸"
c6.x_axis.title = "Месяц"
c6.width    = CHART_WIDTH
c6.height   = CHART_HEIGHT

ref_mkt = Reference(ws, min_col=2, max_col=2, min_row=fun_hdr_row, max_row=fun_data_end)
sm_1 = Series(ref_mkt, title="Маркетинг ₸")
sm_1.graphicalProperties.solidFill = C_YELLOW
c6.append(sm_1)

ref_frev = Reference(ws, min_col=7, max_col=7, min_row=fun_hdr_row, max_row=fun_data_end)
sm_2 = Series(ref_frev, title="Выручка ₸")
sm_2.graphicalProperties.solidFill = C_BLUE
c6.append(sm_2)

c6.set_categories(cats_fun)
ws.add_chart(c6, f"K{fun_start}")

# ─────────────────────────────────────────────────────────────────────────────
# CHART 7: План продаж — Выручка по месяцам (Май–Дек)
# ─────────────────────────────────────────────────────────────────────────────
c7 = BarChart()
c7.type     = "col"
c7.grouping = "clustered"
c7.title    = "План продаж: Выручка по месяцам Май–Дек 2026 (₸)"
c7.style    = 10
c7.y_axis.title = "₸"
c7.x_axis.title = "Месяц"
c7.width    = CHART_WIDTH
c7.height   = CHART_HEIGHT

cats_plan = Reference(ws, min_col=1, max_col=1, min_row=plan_hdr_row+1, max_row=plan_data_end)

ref_plan_rev = Reference(ws, min_col=2, max_col=2, min_row=plan_hdr_row, max_row=plan_data_end)
sp1 = Series(ref_plan_rev, title="Выручка план ₸")
sp1.graphicalProperties.solidFill = C_BLUE
c7.append(sp1)

ref_plan_ord = Reference(ws, min_col=3, max_col=3, min_row=plan_hdr_row, max_row=plan_data_end)
# Заказы на вторичной оси не делаем — только выручка
# Добавим линию заказов как lineChart поверх
c7.set_categories(cats_plan)
ws.add_chart(c7, f"E{plan_start}")

# ─────────────────────────────────────────────────────────────────────────────
# CHART 8: План продаж — Заказов по месяцам (линия)
# ─────────────────────────────────────────────────────────────────────────────
c8 = LineChart()
c8.title    = "План продаж: Кол-во заказов Май–Дек 2026"
c8.style    = 10
c8.y_axis.title = "Заказов"
c8.x_axis.title = "Месяц"
c8.width    = CHART_WIDTH
c8.height   = CHART_HEIGHT

ref_ord = Reference(ws, min_col=3, max_col=3, min_row=plan_hdr_row, max_row=plan_data_end)
so1 = Series(ref_ord, title="Заказов план")
so1.graphicalProperties.line.solidFill = C_GREEN
so1.graphicalProperties.line.width = 25000
so1.smooth = True
c8.append(so1)

ref_check = Reference(ws, min_col=4, max_col=4, min_row=plan_hdr_row, max_row=plan_data_end)
so2 = Series(ref_check, title="Ср. чек ₸")
so2.graphicalProperties.line.solidFill = C_YELLOW
so2.graphicalProperties.line.width = 20000
so2.smooth = True
c8.append(so2)

c8.set_categories(cats_plan)
ws.add_chart(c8, f"K{plan_start}")

print("Диаграммы созданы. Финальная настройка листа...")

# ─────────────────────────────────────────────────────────────────────────────
# Закрепить первые 2 строки
ws.freeze_panes = 'A3'

# Переместить лист на первое место
wb.move_sheet(SHEET_NAME, offset=0)

# ─────────────────────────────────────────────────────────────────────────────
# Сохранение
# ─────────────────────────────────────────────────────────────────────────────
wb.save(DST)
print()
print("=" * 60)
print(f"✅ Файл сохранён: {DST}")
print(f"   Лист '{SHEET_NAME}' перемещён на первое место")
print(f"   Диаграмм создано: 8")
print("=" * 60)
