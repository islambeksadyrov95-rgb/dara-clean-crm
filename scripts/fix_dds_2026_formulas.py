import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook
import re

FILE = "D:/Mind map/Dara Clean/Консультация/Финансы/Плановый расчет.xlsx"

wb = load_workbook(FILE)
ws26 = wb["ДДС 2026 V2"]
ws25v2 = wb["ДДС 2025 V2"]

# -----------------------------------------------------------------------
# 1. Build name→row mapping for ДДС 2025 V2 (the TARGET sheet)
# -----------------------------------------------------------------------
name_to_row_25v2 = {}
for row in range(1, ws25v2.max_row + 1):
    val = ws25v2.cell(row=row, column=1).value
    if val:
        name_to_row_25v2[val.strip()] = row

# -----------------------------------------------------------------------
# 2. Build row→name mapping for old 'ДДС 2025' sheet (the SOURCE of formulas)
#    Since we don't have the old sheet, we derive it from the KNOWN item list
#    provided in the prompt — items with their old row numbers.
#    BUT we don't know old rows! We need another approach:
#    - The formulas reference old 'ДДС 2025' rows
#    - We know the mapping: 26v2_row → item_name → 25v2_row
#    - So: use the CELL ROW in ДДС 2026 V2 to find item name, then look up in 25v2
# -----------------------------------------------------------------------

# Build name→row for ДДС 2026 V2
name_to_row_26v2 = {}
for row in range(1, ws26.max_row + 1):
    val = ws26.cell(row=row, column=1).value
    if val:
        name_to_row_26v2[val.strip()] = row

# Build row→name for ДДС 2026 V2
row_to_name_26v2 = {}
for row in range(1, ws26.max_row + 1):
    val = ws26.cell(row=row, column=1).value
    if val:
        row_to_name_26v2[row] = val.strip()

# Build row→name for ДДС 2025 V2
row_to_name_25v2 = {}
for row in range(1, ws25v2.max_row + 1):
    val = ws25v2.cell(row=row, column=1).value
    if val:
        row_to_name_25v2[row] = val.strip()

# -----------------------------------------------------------------------
# 3. Row mapping: ДДС 2026 V2 cell row → ДДС 2025 V2 target row
#    Strategy: for each row in 26v2, find item name, then look up in 25v2
# -----------------------------------------------------------------------
def find_in_25v2(name):
    """Find row in ДДС 2025 V2 for a given item name."""
    # Exact match
    if name in name_to_row_25v2:
        return name_to_row_25v2[name]
    # Strip whitespace
    name_stripped = name.strip()
    for n25, r25 in name_to_row_25v2.items():
        if n25.strip() == name_stripped:
            return r25
    # Remove special chars
    name_clean = name.replace('▶', '').replace('►', '').strip()
    for n25, r25 in name_to_row_25v2.items():
        n25_clean = n25.replace('▶', '').replace('►', '').strip()
        if name_clean == n25_clean:
            return r25
    return None

row_26_to_25v2 = {}
for r26, name in row_to_name_26v2.items():
    r25 = find_in_25v2(name)
    if r25 is not None:
        row_26_to_25v2[r26] = r25

print("=== Row mapping (DDS 2026 V2 row -> DDS 2025 V2 row) ===")
for r26, r25 in sorted(row_26_to_25v2.items()):
    n26 = row_to_name_26v2.get(r26, "")[:35]
    n25 = row_to_name_25v2.get(r25, "")[:35]
    print(f"  row {r26:3d} ({n26}) -> {r25:3d} ({n25})")
print(f"Mapped: {len(row_26_to_25v2)} rows")

# -----------------------------------------------------------------------
# 4. Month column mapping
#    ДДС 2026 V2 layout columns (where plan formulas live):
#      F=Apr, I=May, L=Jun, O=Jul, R=Aug, U=Sep, X=Oct, AA=Nov, AD=Dec
#    Old 'ДДС 2025' columns referenced in formulas:
#      H=Apr, J=May, L=Jun, N=Jul, P=Aug, R=Sep, T=Oct, V=Nov, X=Dec
#    ДДС 2025 V2 month columns (where we now reference):
#      E=Apr, F=May, G=Jun, H=Jul, I=Aug, J=Sep, K=Oct, L=Nov, M=Dec
# -----------------------------------------------------------------------
old_col_to_v2_col = {
    'H': 'E',   # April
    'J': 'F',   # May
    'L': 'G',   # June
    'N': 'H',   # July
    'P': 'I',   # August
    'R': 'J',   # September
    'T': 'K',   # October
    'V': 'L',   # November
    'X': 'M',   # December
}

# -----------------------------------------------------------------------
# 5. Pattern matching
#    Actual formula examples:
#      ='ДДС 2025'!H15*(1+$AI$3)
#      ='ДДС 2025'!J107*(1+$AI$4)
#    Pattern: ='ДДС 2025'!{COL}{ROW}*(1+$AI${N})
# -----------------------------------------------------------------------
# Use a flexible pattern that captures the whole thing
pattern = re.compile(r"^='ДДС 2025'!([A-Z]+)(\d+)\*\(1\+\$AI\$(\d+)\)$")

fixed_count = 0
skipped = []

print("\n=== Processing formulas ===")
for row in ws26.iter_rows():
    for cell in row:
        val = cell.value
        if not isinstance(val, str):
            continue
        if "'ДДС 2025'!" not in val:
            continue
        if "'ДДС 2025 V2'!" in val:
            continue

        m = pattern.match(val)
        if not m:
            print(f"  UNMATCHED formula at {cell.coordinate}: {repr(val)}")
            skipped.append((cell.coordinate, val))
            continue

        old_col = m.group(1)     # e.g. 'H', 'J', 'L', 'N', etc.
        old_row = int(m.group(2)) # old row number in 'ДДС 2025'
        ai_num  = m.group(3)      # e.g. '3', '4'

        # Get new column in ДДС 2025 V2
        new_col = old_col_to_v2_col.get(old_col)
        if new_col is None:
            print(f"  UNKNOWN old column '{old_col}' at {cell.coordinate}: {val}")
            skipped.append((cell.coordinate, val))
            continue

        # Get new row: use the CELL'S OWN ROW to find item name → 25v2 row
        cell_row = cell.row
        new_row = row_26_to_25v2.get(cell_row)
        if new_row is None:
            # Try to find by the old row number (cross-reference through item name)
            # The old_row is from 'ДДС 2025' — try to find matching item
            print(f"  NO ROW MAPPING for 26v2 row {cell_row} at {cell.coordinate}: {val}")
            skipped.append((cell.coordinate, val))
            continue

        # Build new formula
        new_formula = f"='ДДС 2025 V2'!{new_col}{new_row}*(1+$AI${ai_num})"
        cell.value = new_formula
        fixed_count += 1
        n26 = row_to_name_26v2.get(cell_row, "?")
        n25 = row_to_name_25v2.get(new_row, "?")
        print(f"  Fixed {cell.coordinate}: {val}")
        print(f"         -> {new_formula}  [{n26[:25]} -> {n25[:25]}]")

print(f"\n=== DONE: Fixed {fixed_count} formulas, Skipped {len(skipped)} ===")
if skipped:
    print("Skipped cells:")
    for coord, v in skipped:
        print(f"  {coord}: {v}")

wb.save(FILE)
print("\nFile saved!")
